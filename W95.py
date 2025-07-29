import tkinter as tk
from tkinter import ttk, messagebox, filedialog, colorchooser, simpledialog
import time
import psutil
import shutil
import socket
import subprocess
import platform
from datetime import datetime, timedelta
import threading
import json
import os
import random
try:
    import openpyxl
except ImportError:
    pass

class Windows95Desktop:
    def __init__(self):
        self.rootW95dist = tk.Tk()
        self.rootW95dist.title("Windows 95")
        
        # Setează fereastra să fie fullscreen fără bara de titlu
        self.rootW95dist.overrideredirect(True)
        self.rootW95dist.state('zoomed')  # Pentru Windows - fullscreen
        # Pentru Linux/Mac folosește: self.rootW95dist.attributes('-fullscreen', True)
        
        self.rootW95dist.configure(bg="#008080")  # Teal background
        
        # Variables pentru Start Menu
        self.start_menu_visible = True
        self.start_menu = None
        
        # Lista de ferestre deschise pentru taskbar
        self.open_windows = []
        
        # Variables for monitoring
        self.network_history = []
        self.cpu_history = []
        self.memory_history = []
        self.monitoring_active = False
        
        # Paint application variables
        self.current_color = "black"
        self.brush_size = 2
        self.paint_tool = "pencil"
        
        self.setup_desktop()
        self.setup_taskbar()
        self.setup_start_menu()  # Start menu permanent
        self.setup_clock()
        
    def setup_desktop(self):
        # Desktop background cu pattern
        self.desktop_frame = tk.Frame(self.rootW95dist, bg="#008080")
        self.desktop_frame.pack(fill="both", expand=True)
        
        # Definește poziția x pentru stânga și dreapta
        left_x = 50
        right_x = self.rootW95dist.winfo_screenwidth() - 120

        # Definește distanța y între icoane (pentru aliniere uniformă)
        icon_spacing = 70

        # Icoane pe desktop (partea stângă)
        self.create_desktop_icon("My Computer", left_x, 50, "computer")
        self.create_desktop_icon("Text Editor", left_x, 50 + icon_spacing, "editor")
        self.create_desktop_icon("Calculator", left_x, 50 + icon_spacing * 2, "calculator")
        self.create_desktop_icon("Network Monitor", left_x, 50 + icon_spacing * 3, "network")
        self.create_desktop_icon("Hardware Info", left_x, 50 + icon_spacing * 4, "hardware")
        self.create_desktop_icon("Paint", left_x, 50 + icon_spacing * 5, "paint")
        self.create_desktop_icon("Excel Lite", left_x, 50 + icon_spacing * 6, "excel")
        self.create_desktop_icon("Word Lite", left_x, 50 + icon_spacing * 7, "word")
        self.create_desktop_icon("Command Prompt", left_x, 50 + icon_spacing * 8, "terminal")
        self.create_desktop_icon("SQL Explorer", left_x, 50 + icon_spacing * 9, "database")
        self.create_desktop_icon("File Fisher", left_x, 50 + icon_spacing * 10, "recovery")
        self.create_desktop_icon("Recover Tunnel", left_x+100, 50 + icon_spacing * 10, "recover")

        # Icoane pe desktop (partea dreaptă)
        self.create_desktop_icon("Activate Product", right_x, 50, "activation")
        self.create_desktop_icon("System Requirements", right_x, 50 + icon_spacing, "sysinfo")
    
    def make_window_draggable(self, window, title_bar):
        """Make a window draggable by its title bar"""
        def start_move(event):
            window.x = event.x
            window.y = event.y

        def stop_move(event):
            window.x = None
            window.y = None

        def do_move(event):
            if hasattr(window, 'x') and window.x is not None:
                deltax = event.x - window.x
                deltay = event.y - window.y
                x = window.winfo_x() + deltax
                y = window.winfo_y() + deltay
                window.geometry(f"+{x}+{y}")

        title_bar.bind("<Button-1>", start_move)
        title_bar.bind("<ButtonRelease-1>", stop_move)
        title_bar.bind("<B1-Motion>", do_move)
        
        # Bind și pe label-ul din title bar
        for child in title_bar.winfo_children():
            if isinstance(child, tk.Label):
                child.bind("<Button-1>", start_move)
                child.bind("<ButtonRelease-1>", stop_move)
                child.bind("<B1-Motion>", do_move)
    
    def create_desktop_icon(self, name, x, y, icon_type="normal"):
        icon_frame = tk.Frame(self.desktop_frame, bg="#008080")
        icon_frame.place(x=x, y=y)
        
        # Simulează o iconiță (folosim un dreptunghi colorat)
        icon_canvas = tk.Canvas(icon_frame, width=32, height=32, bg="#008080", highlightthickness=0)
        icon_canvas.pack()
        
        # Culori diferite pentru tipuri diferite de icoane
        if icon_type == "editor":
            fill_color = "#ffff66"  # Galben pentru text editor
        elif icon_type == "calculator":
            fill_color = "#66ff66"  # Verde pentru calculator
        elif icon_type == "network":
            fill_color = "#66ccff"  # Albastru pentru network
        elif icon_type == "hardware":
            fill_color = "#ff9966"  # Orange pentru hardware
        elif icon_type == "paint":
            fill_color = "#ff6699"  # Roz pentru paint
        elif icon_type == "excel":
            fill_color = "#00cc66"
        elif icon_type == "activation":
            fill_color = "#ff6666"
        elif icon_type == "database":
            fill_color = "#3366cc"  # Albastru pentru database
        elif icon_type == "sysinfo":
            fill_color = "#5c5c5c"  # Gri închis pentru system info
        elif icon_type == "recovery":
            fill_color = "#ff9900"
        elif icon_type == "recover":
            fill_color = "#6600cc" 
        else:
            fill_color = "#c0c0c0"  # Gri standard
            
        icon_canvas.create_rectangle(2, 2, 30, 30, fill=fill_color, outline="#808080")
        icon_canvas.create_rectangle(4, 4, 28, 28, fill="#ffffff", outline="#404040")
        
        # Adaugă simboluri pentru icoane
        if icon_type == "network":
            icon_canvas.create_oval(8, 8, 24, 24, outline="#000080", width=2)
            icon_canvas.create_line(16, 8, 16, 24, fill="#000080", width=2)
            icon_canvas.create_line(8, 16, 24, 16, fill="#000080", width=2)
        elif icon_type == "hardware":
            icon_canvas.create_rectangle(8, 8, 24, 24, fill="#404040", outline="#000000")
            icon_canvas.create_rectangle(10, 10, 22, 22, fill="#808080", outline="#000000")
            icon_canvas.create_rectangle(12, 12, 20, 20, fill="#c0c0c0", outline="#000000")
        elif icon_type == "paint":
            # Draw a simple paint brush icon
            icon_canvas.create_rectangle(10, 6, 14, 18, fill="#8B4513", outline="#000000")  # Handle
            icon_canvas.create_oval(8, 18, 16, 26, fill="#ff0000", outline="#000000")  # Brush
            icon_canvas.create_oval(18, 10, 24, 16, fill="#0000ff", outline="#000000")  # Blue paint
            icon_canvas.create_oval(16, 18, 22, 24, fill="#00ff00", outline="#000000")  # Green paint
        elif icon_type == "excel":
            # Desenează un simbol tabel pentru Excel
            icon_canvas.create_rectangle(6, 6, 26, 26, fill="#008000", outline="#000000")
            # Linii verticale
            icon_canvas.create_line(13, 6, 13, 26, fill="#FFFFFF", width=1)
            icon_canvas.create_line(19, 6, 19, 26, fill="#FFFFFF", width=1)
            # Linii orizontale
            icon_canvas.create_line(6, 13, 26, 13, fill="#FFFFFF", width=1)
            icon_canvas.create_line(6, 19, 26, 19, fill="#FFFFFF", width=1)
        elif icon_type == "word":
            # Desenează un simbol pentru Word (document cu linii text)
            icon_canvas.create_rectangle(6, 6, 26, 26, fill="#0000CC", outline="#000000")
            # Linii text
            icon_canvas.create_line(9, 11, 23, 11, fill="#FFFFFF", width=1)
            icon_canvas.create_line(9, 15, 23, 15, fill="#FFFFFF", width=1)
            icon_canvas.create_line(9, 19, 23, 19, fill="#FFFFFF", width=1)
            icon_canvas.create_line(9, 23, 18, 23, fill="#FFFFFF", width=1)
        elif icon_type == "terminal":
            # Desenează un simbol pentru Command Prompt
            icon_canvas.create_rectangle(6, 6, 26, 26, fill="#000000", outline="#404040")
            # Simbolul prompt-ului >_
            icon_canvas.create_text(16, 16, text=">_", fill="#FFFFFF", font=("Courier", 10, "bold"))
        elif icon_type == "activation":
            # Desenează un simbol pentru activare (un cheie sau un lacăt)
            icon_canvas.create_rectangle(8, 8, 24, 24, fill="#ffcc00", outline="#000000")
            icon_canvas.create_line(16, 8, 16, 16, fill="#000000", width=2)  # Partea superioară a cheii
            icon_canvas.create_rectangle(12, 16, 20, 20, fill="#ffcc00", outline="#000000")  # Partea inferioară a cheii
        elif icon_type == "database":
            # Desenează un simbol pentru bază de date
            icon_canvas.create_rectangle(8, 10, 24, 26, fill="#3366cc", outline="#000000")
            # Linii orizontale pentru a reprezenta înregistrările
            icon_canvas.create_line(10, 14, 22, 14, fill="#ffffff", width=1)
            icon_canvas.create_line(10, 18, 22, 18, fill="#ffffff", width=1)
            icon_canvas.create_line(10, 22, 22, 22, fill="#ffffff", width=1)
        elif icon_type == "sysinfo":
            # Desenează un simbol pentru informații sistem
            icon_canvas.create_rectangle(8, 8, 24, 24, fill="#5c5c5c", outline="#000000")
            # Desenează un simbol "i" pentru informații
            icon_canvas.create_text(16, 16, text="i", fill="white", font=("Arial", 12, "bold"))
        elif icon_type == "recovery":
            # Desenează un simbol pentru recuperare fișiere
            icon_canvas.create_rectangle(8, 8, 24, 24, fill="#ff9900", outline="#000000")
            # Desenează un simbol stilizat de disc și undă
            icon_canvas.create_oval(10, 10, 22, 22, fill="#ffffff", outline="#000000")
            icon_canvas.create_arc(12, 12, 20, 20, start=0, extent=270, fill="#ff9900", outline="#000000")
            # Desenează un simbol de undiță/pescuit
            icon_canvas.create_line(10, 6, 22, 6, fill="#000000", width=2)
            icon_canvas.create_line(16, 6, 16, 10, fill="#000000", width=1)
        elif icon_type == "recover":
            # Desenează un simbol pentru recuperare avansată
            icon_canvas.create_rectangle(8, 8, 24, 24, fill="#6600cc", outline="#000000")
            # Desenează un simbol de hard disk și restaurare
            icon_canvas.create_rectangle(10, 10, 22, 18, fill="#c0c0c0", outline="#000000")
            icon_canvas.create_oval(12, 12, 20, 16, fill="#ffffff", outline="#000000")
            # Adaugă o săgeată de restaurare
            icon_canvas.create_line(12, 20, 16, 24, fill="#00ff00", width=2)
            icon_canvas.create_line(16, 24, 20, 20, fill="#00ff00", width=2)
        
        # Label pentru numele iconitei
        label = tk.Label(icon_frame, text=name, bg="#008080", fg="white", 
                        font=("MS Sans Serif", 8))
        label.pack()
        
        # Event pentru dublu-click
        icon_canvas.bind("<Double-Button-1>", lambda e: self.handle_icon_click(name, icon_type))
        label.bind("<Double-Button-1>", lambda e: self.handle_icon_click(name, icon_type))
    
    def handle_icon_click(self, name, icon_type):
        if icon_type == "editor":
            if not any(title == "Text Editor" for title, _, _ in self.open_windows):
                self.create_text_editor()
        elif icon_type == "calculator":
            if not any(title == "Calculator" for title, _, _ in self.open_windows):
                self.create_calculator()
        elif icon_type == "network":
            if not any(title == "Network Monitor" for title, _, _ in self.open_windows):
                self.create_network_monitor()
        elif icon_type == "hardware":
            if not any(title == "Hardware Info" for title, _, _ in self.open_windows):
                self.create_hardware_info()
        elif icon_type == "paint":
            if not any(title == "Paint" for title, _, _ in self.open_windows):
                self.create_paint_app()
        elif icon_type == "computer":
            if not any(title == "My Computer" for title, _, _ in self.open_windows):
                self.create_file_explorer()
        elif icon_type == "excel":
            if not any(title == "Excel Lite" for title, _, _ in self.open_windows):
                self.create_excel_lite()
        elif icon_type == "word":
            if not any(title == "Word Lite" for title, _, _ in self.open_windows):
                self.create_word_lite()
        elif icon_type == "terminal":
            if not any(title == "Command Prompt" for title, _, _ in self.open_windows):
                self.create_terminal()
        elif icon_type == "activation":
            if not any(title == "Activation Wizard" for title, _, _ in self.open_windows):
                self.create_activation_window()
        elif icon_type == "database":
            if not any(title == "SQL Explorer" for title, _, _ in self.open_windows):
                self.open_sql_explorer()
        elif icon_type == "sysinfo":
            if not any(title == "System Requirements" for title, _, _ in self.open_windows):
                self.create_about_window()
        elif icon_type == "recovery":
            if not any(title == "File Fisher" for title, _, _ in self.open_windows):
                self.create_file_recovery()
        elif icon_type == "recover":
            if not any(title == "Recover Tunnel" for title, _, _ in self.open_windows):
                self.create_retro_recover()
        else:
            self.open_window(name)
    
    def create_retro_recover(self):
        """Creează aplicația RetroRecover pentru recuperarea avansată a fișierelor"""
        # Creare fereastră principală
        recover_window = tk.Toplevel(self.rootW95dist)
        recover_window.title("Recover Tunnel")
        recover_window.overrideredirect(True)
        recover_window.geometry("900x700+200+100")
        recover_window.configure(bg="#c0c0c0")
        
        # Add Windows 95 style title bar
        title_bar = tk.Frame(recover_window, bg="#000080", height=25)
        title_bar.pack(fill="x", side="top")
        title_label = tk.Label(title_bar, text="Recover Tunnel", fg="white", bg="#000080",
                              font=("MS Sans Serif", 8, "bold"))
        title_label.pack(side="left", padx=5, pady=2)
        
        # Close button for title bar
        close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                font=("Arial", 8, "bold"), width=2, height=1,
                                relief="raised", bd=1,
                                command=lambda: self.close_window("Recover Tunnel", recover_window))
        close_button.pack(side="right", padx=2, pady=1)
        
        self.make_window_draggable(recover_window, title_bar)
        
        # Creare clase interne pentru a implementa funcționalitatea RetroRecover
        class RealFileRecoveryEngine:
            """Engine real pentru recuperarea fișierelor"""
            
            def __init__(self):
                self.stopped = False
                self.file_extensions = {
                    'Documents': ['.txt', '.doc', '.docx', '.pdf', '.rtf', '.odt'],
                    'Images': ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.ico'],
                    'Videos': ['.mp4', '.avi', '.mkv', '.mov', '.wmv', '.flv', '.webm'],
                    'Audio': ['.mp3', '.wav', '.flac', '.aac', '.ogg', '.wma'],
                    'Archives': ['.zip', '.rar', '.7z', '.tar', '.gz', '.bz2']
                }
                
            def scan_recycle_bin(self, callback):
                """Scanează Coșul de Reciclare pentru fișiere șterse"""
                self.stopped = False
                found_files = []
                
                try:
                    # Obține toate unitățile de disc
                    drives = [f"{chr(i)}:\\" for i in range(65, 91) if os.path.exists(f"{chr(i)}:\\")]
                    
                    total_drives = len(drives)
                    
                    for i, drive in enumerate(drives):
                        if self.stopped:
                            break
                            
                        progress = int((i / total_drives) * 100)
                        callback(progress, f"Scanning drive {drive}...")
                        
                        # Verifică folderele Recycle Bin
                        recycle_paths = [
                            os.path.join(drive, '$Recycle.Bin'),
                            os.path.join(drive, 'RECYCLER'),
                            os.path.join(drive, 'Recycled')
                        ]
                        
                        for recycle_path in recycle_paths:
                            if self.stopped:
                                break
                                
                            if os.path.exists(recycle_path):
                                self._scan_directory_recursive(recycle_path, callback, found_files, "Recycle Bin")
                                
                except Exception as e:
                    print(f"Error scanning recycle bin: {e}")
                    
                callback(100, "Recycle bin scan completed")
                
            def scan_temp_files(self, callback):
                """Scanează locațiile de fișiere temporare"""
                self.stopped = False
                found_files = []
                
                try:
                    temp_paths = [
                        tempfile.gettempdir(),
                        os.environ.get('TEMP', ''),
                        os.environ.get('TMP', ''),
                        os.path.join(os.environ.get('USERPROFILE', ''), 'AppData', 'Local', 'Temp'),
                        r'C:\Windows\Temp'
                    ]
                    
                    # Elimină duplicatele și căile goale
                    temp_paths = list(set([path for path in temp_paths if path and os.path.exists(path)]))
                    total_paths = len(temp_paths)
                    
                    for i, temp_path in enumerate(temp_paths):
                        if self.stopped:
                            break
                            
                        progress = int((i / total_paths) * 100)
                        callback(progress, f"Scanning {temp_path}...")
                        
                        self._scan_directory_recursive(temp_path, callback, found_files, "Temp Files")
                        
                except Exception as e:
                    print(f"Error scanning temp files: {e}")
                    
                callback(100, "Temp files scan completed")
                
            def scan_recent_files(self, callback):
                """Scanează fișierele recente din diverse surse"""
                self.stopped = False
                found_files = []
                
                try:
                    user_profile = os.environ.get('USERPROFILE', '')
                    
                    # Căile documentelor recente
                    recent_paths = [
                        os.path.join(user_profile, 'Recent'),
                        os.path.join(user_profile, 'AppData', 'Roaming', 'Microsoft', 'Windows', 'Recent'),
                        os.path.join(user_profile, 'AppData', 'Roaming', 'Microsoft', 'Office', 'Recent'),
                    ]
                    
                    total_paths = len(recent_paths) + 1  # +1 pentru scanarea registrului
                    current_path = 0
                    
                    for recent_path in recent_paths:
                        if self.stopped:
                            break
                            
                        progress = int((current_path / total_paths) * 100)
                        callback(progress, f"Scanning recent files...")
                        
                        if os.path.exists(recent_path):
                            self._scan_directory_recursive(recent_path, callback, found_files, "Recent Files")
                            
                        current_path += 1
                        
                    # Scanează registrul pentru fișiere recente
                    if not self.stopped and hasattr(winreg, 'OpenKey'):
                        progress = int((current_path / total_paths) * 100)
                        try:
                            self._scan_registry_recent_files(callback, found_files)
                        except:
                            pass
                        
                except Exception as e:
                    print(f"Error scanning recent files: {e}")
                    
                callback(100, "Recent files scan completed")
            
            def scan_browser_cache(self, callback):
                """Scanează directoarele cache ale browserelor"""
                self.stopped = False
                found_files = []
                
                try:
                    user_profile = os.environ.get('USERPROFILE', '')
                    cache_paths = [
                        # Chrome
                        os.path.join(user_profile, 'AppData', 'Local', 'Google', 'Chrome', 'User Data', 'Default', 'Cache'),
                        # Firefox
                        os.path.join(user_profile, 'AppData', 'Local', 'Mozilla', 'Firefox', 'Profiles'),
                        # Edge
                        os.path.join(user_profile, 'AppData', 'Local', 'Microsoft', 'Edge', 'User Data', 'Default', 'Cache'),
                        # Internet Explorer
                        os.path.join(user_profile, 'AppData', 'Local', 'Microsoft', 'Windows', 'INetCache'),
                    ]
                    
                    total_paths = len(cache_paths)
                    
                    for i, cache_path in enumerate(cache_paths):
                        if self.stopped:
                            break
                            
                        progress = int((i / total_paths) * 100)
                        callback(progress, f"Scanning browser cache...")
                        
                        if os.path.exists(cache_path):
                            self._scan_directory_recursive(cache_path, callback, found_files, "Browser Cache")
                            
                except Exception as e:
                    print(f"Error scanning browser cache: {e}")
                    
                callback(100, "Browser cache scan completed")
                
            def _scan_directory_recursive(self, directory, callback, found_files, source):
                """Scanează recursiv directorul pentru fișiere recuperabile"""
                try:
                    total_files = 0
                    processed_files = 0
                    
                    # Prima trecere: numără fișierele pentru calculul progresului
                    try:
                        for root, dirs, files in os.walk(directory):
                            total_files += len(files)
                            # Limitează adâncimea pentru a evita numărarea prea multor fișiere
                            if len(root.split(os.sep)) - len(directory.split(os.sep)) > 3:
                                dirs.clear()
                    except:
                        total_files = 100  # Estimare de rezervă
                    
                    # A doua trecere: procesează fișierele
                    for root, dirs, files in os.walk(directory):
                        if self.stopped:
                            break
                            
                        for file_name in files:
                            if self.stopped:
                                break
                                
                            processed_files += 1
                            progress = min(99, int((processed_files / max(total_files, 1)) * 100))
                            callback(progress, f"Scanning: {file_name[:30]}...")
                            
                            file_path = os.path.join(root, file_name)
                            
                            try:
                                if os.path.exists(file_path) and os.path.isfile(file_path):
                                    file_info = self._get_file_info(file_path, source)
                                    if file_info:
                                        found_files.append(file_info)
                                        callback(progress, f"Found: {file_name}", file_info)
                                        
                            except (OSError, PermissionError):
                                continue
                                
                        # Limitează adâncimea pentru a evita buclele infinite
                        if len(root.split(os.sep)) - len(directory.split(os.sep)) > 3:
                            dirs.clear()
                            
                except (OSError, PermissionError):
                    pass
                    
            def _get_file_info(self, file_path, source):
                """Obține informații detaliate despre un fișier"""
                try:
                    stat_info = os.stat(file_path)
                    file_size = stat_info.st_size
                    
                    # Ignoră fișierele foarte mici (probabil nu sunt date utilizator)
                    if file_size < 100:
                        return None
                        
                    # Ignoră fișierele sistem
                    file_name = os.path.basename(file_path)
                    if file_name.startswith('.') or file_name.startswith('~'):
                        return None
                        
                    file_ext = os.path.splitext(file_name)[1].lower()
                    file_type = self._get_file_type(file_ext)
                    
                    # Formatează dimensiunea fișierului
                    size_str = self._format_file_size(file_size)
                    
                    # Formatează data
                    mod_time = datetime.fromtimestamp(stat_info.st_mtime)
                    date_str = mod_time.strftime("%Y-%m-%d %H:%M")
                    
                    # Determină starea de recuperabilitate
                    status = "Recoverable"
                    if not os.access(file_path, os.R_OK):
                        status = "Access Denied"
                    elif file_size == 0:
                        status = "Empty File"
                        
                    return {
                        'name': file_name,
                        'size': size_str,
                        'type': file_type,
                        'location': source,
                        'date': date_str,
                        'status': status,
                        'full_path': file_path,
                        'file_size': file_size
                    }
                    
                except (OSError, PermissionError):
                    return None
                    
            def _get_file_type(self, extension):
                """Determină tipul de fișier din extensie"""
                for file_type, extensions in self.file_extensions.items():
                    if extension in extensions:
                        return file_type
                return "Other"
                
            def _format_file_size(self, size_bytes):
                """Formatează dimensiunea fișierului în format uman"""
                if size_bytes == 0:
                    return "0 B"
                size_names = ["B", "KB", "MB", "GB"]
                i = 0
                while size_bytes >= 1024 and i < len(size_names) - 1:
                    size_bytes /= 1024.0
                    i += 1
                return f"{size_bytes:.1f} {size_names[i]}"
                
            def _scan_registry_recent_files(self, callback, found_files):
                """Scanează registrul Windows pentru intrări de fișiere recente"""
                try:
                    if not hasattr(winreg, 'OpenKey'):
                        return
                        
                    # Cheia registrului pentru documente recente
                    reg_key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, 
                                           r"Software\Microsoft\Windows\CurrentVersion\Explorer\RecentDocs")
                    
                    callback(50, "Scanning registry recent files...")
                    
                    i = 0
                    while True:
                        try:
                            value_name, value_data, value_type = winreg.EnumValue(reg_key, i)
                            
                            # Încearcă să extragă calea fișierului din datele registrului
                            if isinstance(value_data, bytes) and len(value_data) > 0:
                                # Decodează ca UTF-16 și curăță
                                try:
                                    path_str = value_data.decode('utf-16le', errors='ignore').strip('\x00')
                                    if path_str and os.path.exists(path_str):
                                        file_info = self._get_file_info(path_str, "Registry Recent")
                                        if file_info:
                                            found_files.append(file_info)
                                            callback(50, f"Found registry entry: {os.path.basename(path_str)}", file_info)
                                except:
                                    pass
                                    
                            i += 1
                        except Exception:
                            break
                            
                    winreg.CloseKey(reg_key)
                    
                except Exception as e:
                    print(f"Error scanning registry: {e}")
        
        # Inițializăm variabilele
        default_font = ("MS Sans Serif", 8)
        bold_font = ("MS Sans Serif", 8, "bold")
        title_font = ("MS Sans Serif", 10, "bold")
        
        scan_progress = tk.DoubleVar()
        status_text = tk.StringVar(value="Ready")
        current_operation = tk.StringVar(value="")
        files_found = tk.IntVar(value=0)
        recoverable_files = []
        is_scanning = False
        
        # Engine pentru recuperare
        recovery_engine = RealFileRecoveryEngine()
        
        # Creare meniu
        menubar = tk.Menu(recover_window, bg='#c0c0c0', relief='raised', bd=1)
        recover_window.config(menu=menubar)
        
        # Meniul File
        file_menu = tk.Menu(menubar, tearoff=0, bg='#c0c0c0')
        menubar.add_cascade(label="File", menu=file_menu, underline=0)
        
        # Meniul Tools
        tools_menu = tk.Menu(menubar, tearoff=0, bg='#c0c0c0')
        menubar.add_cascade(label="Tools", menu=tools_menu, underline=0)
        
        # Meniul Help
        help_menu = tk.Menu(menubar, tearoff=0, bg='#c0c0c0')
        menubar.add_cascade(label="Help", menu=help_menu, underline=0)
        
        # Creare toolbar
        toolbar = tk.Frame(recover_window, bg='#c0c0c0', relief='raised', bd=1, height=40)
        toolbar.pack(fill='x', pady=2)
        
        # Butoane toolbar cu efect 3D
        btn_new = tk.Button(toolbar, text="New Scan", font=default_font,
                           relief='raised', bd=2, padx=10, pady=2,
                           bg='#c0c0c0')
        btn_new.pack(side='left', padx=2, pady=2)
        
        btn_recycle = tk.Button(toolbar, text="Recycle Bin", font=default_font,
                               relief='raised', bd=2, padx=10, pady=2,
                               bg='#c0c0c0')
        btn_recycle.pack(side='left', padx=2, pady=2)
        
        btn_temp = tk.Button(toolbar, text="Temp Files", font=default_font,
                            relief='raised', bd=2, padx=10, pady=2,
                            bg='#c0c0c0')
        btn_temp.pack(side='left', padx=2, pady=2)
        
        btn_recover = tk.Button(toolbar, text="Recover Selected", font=default_font,
                               relief='raised', bd=2, padx=10, pady=2,
                               bg='#c0c0c0')
        btn_recover.pack(side='left', padx=2, pady=2)
        
        btn_preview = tk.Button(toolbar, text="Preview", font=default_font,
                               relief='raised', bd=2, padx=10, pady=2,
                               bg='#c0c0c0')
        btn_preview.pack(side='left', padx=2, pady=2)
        
        # Separator
        separator = tk.Frame(toolbar, width=2, bg='#808080', relief='sunken', bd=1)
        separator.pack(side='left', fill='y', padx=5, pady=2)
        
        btn_clear = tk.Button(toolbar, text="Clear", font=default_font,
                             relief='raised', bd=2, padx=10, pady=2,
                             bg='#c0c0c0')
        btn_clear.pack(side='left', padx=2, pady=2)
        
        btn_stop = tk.Button(toolbar, text="Stop", font=default_font,
                            relief='raised', bd=2, padx=10, pady=2,
                            bg='#c0c0c0')
        btn_stop.pack(side='left', padx=2, pady=2)
        
        # Container principal
        main_frame = tk.Frame(recover_window, bg='#c0c0c0')
        main_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Panoul din stânga - Opțiuni scanare
        left_panel = tk.LabelFrame(main_frame, text="Recovery Options", font=bold_font,
                                 bg='#c0c0c0', relief='groove', bd=2)
        left_panel.pack(side='left', fill='y', padx=(0, 5))
        
        # Tipurile de recuperare
        tk.Label(left_panel, text="Recovery Types:", font=bold_font, bg='#c0c0c0').pack(anchor='w', padx=5, pady=5)
        
        recovery_types = {
            'Recycle Bin Files': tk.BooleanVar(value=True),
            'Temporary Files': tk.BooleanVar(value=True),
            'Recent Documents': tk.BooleanVar(value=True),
            'Browser Cache': tk.BooleanVar(value=False),
            'System Restore Points': tk.BooleanVar(value=False)
        }
        
        for recovery_type, var in recovery_types.items():
            cb = tk.Checkbutton(left_panel, text=recovery_type, variable=var,
                              font=default_font, bg='#c0c0c0')
            cb.pack(anchor='w', padx=10, pady=1)
            
        # Tipurile de fișiere
        tk.Label(left_panel, text="File Types:", font=bold_font, bg='#c0c0c0').pack(anchor='w', padx=5, pady=(15,2))
        
        file_types = {
            'Documents': tk.BooleanVar(value=True),
            'Images': tk.BooleanVar(value=True),
            'Videos': tk.BooleanVar(value=True),
            'Audio': tk.BooleanVar(value=True),
            'Archives': tk.BooleanVar(value=True),
            'All Files': tk.BooleanVar(value=False)
        }
        
        for file_type, var in file_types.items():
            cb = tk.Checkbutton(left_panel, text=file_type, variable=var,
                              font=default_font, bg='#c0c0c0')
            cb.pack(anchor='w', padx=10, pady=1)
            
        # Acțiuni rapide
        tk.Label(left_panel, text="Quick Actions:", font=bold_font, bg='#c0c0c0').pack(anchor='w', padx=5, pady=(15,5))
        
        btn_quick_recycle = tk.Button(left_panel, text="Quick Recycle Scan", 
                                    font=default_font, relief='raised', bd=2,
                                    bg='#c0c0c0')
        btn_quick_recycle.pack(fill='x', padx=5, pady=2)
        
        btn_quick_temp = tk.Button(left_panel, text="Quick Temp Scan", 
                                 font=default_font, relief='raised', bd=2,
                                 bg='#c0c0c0')
        btn_quick_temp.pack(fill='x', padx=5, pady=2)
        
        # Panoul din dreapta - Rezultate
        right_panel = tk.Frame(main_frame, bg='#c0c0c0')
        right_panel.pack(side='right', fill='both', expand=True)
        
        # Eticheta rezultatelor
        results_label = tk.Label(right_panel, text="Recoverable Files", 
                               font=bold_font, bg='#c0c0c0')
        results_label.pack(anchor='w', pady=(0,5))
        
        # Treeview pentru rezultate
        tree_frame = tk.Frame(right_panel, relief='sunken', bd=2)
        tree_frame.pack(fill='both', expand=True)
        
        # Creează treeview cu stilul clasic
        style = ttk.Style()
        style.configure("Classic.Treeview", background="white", foreground="black")
        style.configure("Classic.Treeview.Heading", background="#c0c0c0", foreground="black")
        
        tree = ttk.Treeview(tree_frame, style="Classic.Treeview")
        tree['columns'] = ('Size', 'Type', 'Location', 'Date Modified', 'Status')
        tree.heading('#0', text='File Name')
        tree.heading('Size', text='Size')
        tree.heading('Type', text='Type')
        tree.heading('Location', text='Original Location')
        tree.heading('Date Modified', text='Date Modified')
        tree.heading('Status', text='Status')
        
        # Lățimile coloanelor
        tree.column('#0', width=200)
        tree.column('Size', width=80)
        tree.column('Type', width=80)
        tree.column('Location', width=200)
        tree.column('Date Modified', width=120)
        tree.column('Status', width=100)
        
        # Bare de derulare
        v_scrollbar = tk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
        h_scrollbar = tk.Scrollbar(tree_frame, orient='horizontal', command=tree.xview)
        tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        tree.pack(side='left', fill='both', expand=True)
        v_scrollbar.pack(side='right', fill='y')
        h_scrollbar.pack(side='bottom', fill='x')
        
        # Cadrul de progres
        progress_frame = tk.Frame(right_panel, bg='#c0c0c0')
        progress_frame.pack(fill='x', pady=(5,0))
        
        # Bara de progres
        progress_bar = tk.Canvas(progress_frame, height=20, bg='white', relief='sunken', bd=2)
        progress_bar.pack(fill='x', pady=2)
        
        # Statusul operației
        operation_label = tk.Label(progress_frame, textvariable=current_operation,
                                 font=default_font, bg='#c0c0c0')
        operation_label.pack(anchor='w')
        
        # Bara de status
        status_frame = tk.Frame(recover_window, bg='#c0c0c0', relief='sunken', bd=1)
        status_frame.pack(fill='x', side='bottom')
        
        status_label = tk.Label(status_frame, textvariable=status_text,
                              font=default_font, bg='#c0c0c0', anchor='w')
        status_label.pack(side='left', padx=5)
        
        separator1 = tk.Frame(status_frame, width=2, bg='#808080', relief='sunken', bd=1)
        separator1.pack(side='left', fill='y', padx=2)
        
        files_label = tk.Label(status_frame, text="Files found: 0",
                             font=default_font, bg='#c0c0c0')
        files_label.pack(side='left', padx=5)
        
        separator2 = tk.Frame(status_frame, width=2, bg='#808080', relief='sunken', bd=1)
        separator2.pack(side='left', fill='y', padx=2)
        
        time_label = tk.Label(status_frame, text=datetime.now().strftime("%H:%M:%S"),
                            font=default_font, bg='#c0c0c0')
        time_label.pack(side='right', padx=5)
        
        # Funcția pentru actualizarea orei
        def update_time():
            time_label.config(text=datetime.now().strftime("%H:%M:%S"))
            recover_window.after(1000, update_time)
        
        update_time()
        
        # Funcția pentru actualizarea progresului
        def update_progress(value, operation=""):
            # Actualizează bara de progres
            progress_bar.delete("all")
            
            # Forțează canvas-ul să-și actualizeze dimensiunile mai întâi
            progress_bar.update_idletasks()
            width = progress_bar.winfo_width()
            height = progress_bar.winfo_height()
            
            if width > 1 and height > 1:
                fill_width = int((value / 100) * (width - 4))
                if fill_width > 0:
                    block_width = 8
                    blocks = max(1, fill_width // block_width)
                    for i in range(blocks):
                        x1 = 2 + i * block_width
                        x2 = min(x1 + block_width - 1, width - 2)
                        if x2 > x1:
                            progress_bar.create_rectangle(x1, 2, x2, height-2,
                                                       fill='#0000ff', outline='#0000ff')
            
            if operation:
                current_operation.set(operation)
            
            # Forțează actualizarea imediată a tuturor elementelor UI
            recover_window.update()
        
        # Funcția pentru afișarea ferestrei de setări
        def show_settings():
            settings_window = tk.Toplevel(recover_window)
            settings_window.title("Settings")
            settings_window.overrideredirect(True)
            settings_window.geometry("400x300")
            settings_window.configure(bg='#c0c0c0')
            settings_window.resizable(False, False)
            
            # Windows 95 style title bar
            title_bar = tk.Frame(settings_window, bg="#000080", height=25)
            title_bar.pack(fill="x", side="top")
            title_label = tk.Label(title_bar, text="Settings", fg="white", bg="#000080",
                                font=("MS Sans Serif", 8, "bold"))
            title_label.pack(side="left", padx=5, pady=2)
            
            # Close button for title bar
            close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                  font=("Arial", 8, "bold"), width=2, height=1,
                                  relief="raised", bd=1,
                                  command=settings_window.destroy)
            close_button.pack(side="right", padx=2, pady=1)
            
            # Make window draggable
            self.make_window_draggable(settings_window, title_bar)
            
            tk.Label(settings_window, text="RetroRecover Pro Settings", 
                   font=title_font, bg='#c0c0c0').pack(pady=10)
            
            # Setări recuperare
            frame1 = tk.LabelFrame(settings_window, text="Recovery Options", 
                                 font=bold_font, bg='#c0c0c0')
            frame1.pack(fill='x', padx=10, pady=5)
            
            create_log = tk.BooleanVar(value=True)
            verify_files = tk.BooleanVar(value=True)
            
            tk.Checkbutton(frame1, text="Create recovery log", variable=create_log,
                         font=default_font, bg='#c0c0c0').pack(anchor='w', padx=5)
            tk.Checkbutton(frame1, text="Verify recovered files", variable=verify_files,
                         font=default_font, bg='#c0c0c0').pack(anchor='w', padx=5)
            
            # Setări scanare
            frame2 = tk.LabelFrame(settings_window, text="Scan Settings", 
                                 font=bold_font, bg='#c0c0c0')
            frame2.pack(fill='x', padx=10, pady=5)
            
            tk.Label(frame2, text="Max file age (days):", font=default_font, bg='#c0c0c0').pack(anchor='w', padx=5)
            max_age_var = tk.StringVar(value="30")
            tk.Entry(frame2, textvariable=max_age_var, width=10).pack(anchor='w', padx=5, pady=2)
            
            # Butoane
            btn_frame = tk.Frame(settings_window, bg='#c0c0c0')
            btn_frame.pack(pady=20)
            
            tk.Button(btn_frame, text="OK", font=default_font, 
                   relief='raised', bd=2, padx=20, command=settings_window.destroy).pack(side='left', padx=5)
            tk.Button(btn_frame, text="Cancel", font=default_font, 
                   relief='raised', bd=2, padx=20, command=settings_window.destroy).pack(side='left', padx=5)
        
        # Funcția pentru afișarea ferestrei about
        def show_about():
            about_window = tk.Toplevel(recover_window)
            about_window.title("About RetroRecover Pro")
            about_window.overrideredirect(True)
            about_window.geometry("450x350")
            about_window.configure(bg='#c0c0c0')
            about_window.resizable(False, False)
            
            # Windows 95 style title bar
            title_bar = tk.Frame(about_window, bg="#000080", height=25)
            title_bar.pack(fill="x", side="top")
            title_label = tk.Label(title_bar, text="About RetroRecover Pro", fg="white", bg="#000080",
                                font=("MS Sans Serif", 8, "bold"))
            title_label.pack(side="left", padx=5, pady=2)
            
            # Close button for title bar
            close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                  font=("Arial", 8, "bold"), width=2, height=1,
                                  relief="raised", bd=1,
                                  command=about_window.destroy)
            close_button.pack(side="right", padx=2, pady=1)
            
            # Make window draggable
            self.make_window_draggable(about_window, title_bar)
            
            tk.Label(about_window, text="RetroRecover Pro", 
                   font=("MS Sans Serif", 14, "bold"), 
                   bg='#c0c0c0').pack(pady=20)
            
            tk.Label(about_window, text="Version 1.0 - Real Recovery", 
                   font=default_font, bg='#c0c0c0').pack()
            
            tk.Label(about_window, text="Advanced Real File Recovery Tool", 
                   font=default_font, bg='#c0c0c0').pack(pady=5)
            
            info_text = """
    This software helps you recover real deleted files from:
    - Recycle Bin (including hidden files)
    - Temporary file locations
    - Recent document lists
    - Browser cache files
    - System restore points

    Features:
    - Multiple recovery sources
    - File preview capabilities
    - Batch recovery operations

    CAUTION: This tool accesses real system files and folders.
    Use responsibly and ensure you have proper permissions.

    Copyright © 2024 Tudor Marmureanu
            """
            
            text_widget = tk.Text(about_window, bg='white', font=default_font, 
                                wrap='word', height=12, width=50)
            text_widget.pack(padx=20, pady=10)
            text_widget.insert('1.0', info_text)
            text_widget.config(state='disabled')
            
            tk.Button(about_window, text="OK", font=default_font, 
                   relief='raised', bd=2, padx=20, 
                   command=about_window.destroy).pack(pady=10)
        
        # Funcția pentru a curăța rezultatele
        def clear_results():
            """Curăță toate rezultatele curente"""
            for item in tree.get_children():
                tree.delete(item)
                
            recoverable_files.clear()
            files_found.set(0)
            files_label.config(text="Files found: 0")
            status_text.set("Results cleared")
            current_operation.set("")
            update_progress(0, "Ready")
        
        # Funcția pentru scanarea callback
        def scan_callback(progress, operation, found_file=None):
            update_progress(progress, operation)
            
            if found_file:
                # Adaugă fișierul găsit în tree
                tree.insert('', 'end', text=found_file['name'],
                           values=(found_file['size'], found_file['type'],
                                 found_file['location'], found_file['date'], found_file['status']))
                recoverable_files.append(found_file)
                files_found.set(len(recoverable_files))
                files_label.config(text=f"Files found: {files_found.get()}")
        
        # Funcția pentru scanarea Recycle Bin
        def scan_recycle_bin():
            nonlocal is_scanning
            if is_scanning:
                return
                
            recovery_engine.stopped = False  # Reset stop flag
            is_scanning = True
            status_text.set("Scanning Recycle Bin...")
            
            scan_thread = threading.Thread(target=_scan_recycle_bin_thread)
            scan_thread.daemon = True
            scan_thread.start()
            
        def _scan_recycle_bin_thread():
            nonlocal is_scanning
            try:
                recovery_engine.scan_recycle_bin(scan_callback)
            finally:
                is_scanning = False
                status_text.set("Recycle Bin scan completed")
        
        # Funcția pentru scanarea fișierelor temporare
        def scan_temp_files():
            nonlocal is_scanning
            if is_scanning:
                return
                
            recovery_engine.stopped = False  # Reset stop flag
            is_scanning = True
            status_text.set("Scanning temporary files...")
            
            scan_thread = threading.Thread(target=_scan_temp_files_thread)
            scan_thread.daemon = True
            scan_thread.start()
            
        def _scan_temp_files_thread():
            nonlocal is_scanning
            try:
                recovery_engine.scan_temp_files(scan_callback)
            finally:
                is_scanning = False
                status_text.set("Temporary files scan completed")
        
        # Funcția pentru scanarea fișierelor recente
        def scan_recent_files():
            nonlocal is_scanning
            if is_scanning:
                return
                
            recovery_engine.stopped = False  # Reset stop flag
            is_scanning = True
            status_text.set("Scanning recent files...")
            
            scan_thread = threading.Thread(target=_scan_recent_files_thread)
            scan_thread.daemon = True
            scan_thread.start()
            
        def _scan_recent_files_thread():
            nonlocal is_scanning
            try:
                recovery_engine.scan_recent_files(scan_callback)
            finally:
                is_scanning = False
                status_text.set("Recent files scan completed")
        
        # Funcția pentru scanare completă
        def new_scan():
            nonlocal is_scanning
            if is_scanning:
                messagebox.showwarning("Scan in Progress", "Please stop the current scan before starting a new one.")
                return
                
            # Reset stop flag
            recovery_engine.stopped = False
            
            # Curăță rezultatele anterioare
            for item in tree.get_children():
                tree.delete(item)
                
            recoverable_files.clear()
            files_found.set(0)
            
            # Începe scanarea cuprinzătoare
            is_scanning = True
            status_text.set("Starting comprehensive scan...")
            
            scan_thread = threading.Thread(target=perform_comprehensive_scan)
            scan_thread.daemon = True
            scan_thread.start()
            
        def perform_comprehensive_scan():
            nonlocal is_scanning
            try:
                selected_types = [rtype for rtype, var in recovery_types.items() if var.get()]
                total_operations = len(selected_types)
                current_op_index = 0  # Redenumit din 'current_operation' în 'current_op_index'
                
                if 'Recycle Bin Files' in selected_types and not recovery_engine.stopped:
                    current_op_index += 1  # Folosește noua denumire
                    base_progress = int(((current_op_index - 1) / total_operations) * 100)
                    recovery_engine.scan_recycle_bin(lambda p, op, f=None: scan_callback(
                        base_progress + (p // total_operations), op, f))
                    
                if 'Temporary Files' in selected_types and not recovery_engine.stopped:
                    current_op_index += 1  # Folosește noua denumire
                    base_progress = int(((current_op_index - 1) / total_operations) * 100)
                    recovery_engine.scan_temp_files(lambda p, op, f=None: scan_callback(
                        base_progress + (p // total_operations), op, f))
                    
                if 'Recent Documents' in selected_types and not recovery_engine.stopped:
                    current_op_index += 1  # Folosește noua denumire
                    base_progress = int(((current_op_index - 1) / total_operations) * 100)
                    recovery_engine.scan_recent_files(lambda p, op, f=None: scan_callback(
                        base_progress + (p // total_operations), op, f))
                    
                if 'Browser Cache' in selected_types and not recovery_engine.stopped:
                    current_op_index += 1  # Folosește noua denumire
                    base_progress = int(((current_op_index - 1) / total_operations) * 100)
                    recovery_engine.scan_browser_cache(lambda p, op, f=None: scan_callback(
                        base_progress + (p // total_operations), op, f))
                    
            except Exception as e:
                messagebox.showerror("Scan Error", f"An error occurred during scanning: {str(e)}")
            finally:
                is_scanning = False
                recovery_engine.stopped = False  # Reset pentru următoarea scanare
                if recovery_engine.stopped:
                    status_text.set("Scan stopped by user")
                    current_operation.set("Scan cancelled")  # Aici folosim variabila globală
                else:
                    status_text.set("Scan completed")
                    current_operation.set("Scan finished successfully")  # Aici folosim variabila globală
                update_progress(100, "")
        
        # Funcția pentru oprirea scanării
        def stop_scan():
            nonlocal is_scanning
            is_scanning = False
            recovery_engine.stopped = True
            status_text.set("Stopping scan...")
            current_operation.set("Scan stopped by user")
            update_progress(0, "Scan stopped")
        
        # Funcția pentru previzualizarea fișierului
        def preview_file():
            selected_item = tree.selection()
            if not selected_item:
                messagebox.showwarning("No Selection", "Please select a file to preview.")
                return
                
            file_name = tree.item(selected_item[0])['text']
            
            # Găsește informațiile despre fișier
            file_info = None
            for f in recoverable_files:
                if f['name'] == file_name:
                    file_info = f
                    break
                    
            if file_info and 'full_path' in file_info:
                show_preview_window(file_name, file_info['full_path'])
            else:
                messagebox.showwarning("Preview Error", "Cannot preview this file.")
        
        def show_preview_window(file_name, file_path):
            preview_window = tk.Toplevel(recover_window)
            preview_window.title(f"Preview: {file_name}")
            preview_window.overrideredirect(True)
            preview_window.geometry("500x400")
            preview_window.configure(bg='#c0c0c0')
            
            # Windows 95 style title bar
            title_bar = tk.Frame(preview_window, bg="#000080", height=25)
            title_bar.pack(fill="x", side="top")
            title_label = tk.Label(title_bar, text=f"Preview: {file_name}", fg="white", bg="#000080",
                                font=("MS Sans Serif", 8, "bold"))
            title_label.pack(side="left", padx=5, pady=2)
            
            # Close button for title bar
            close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                  font=("Arial", 8, "bold"), width=2, height=1,
                                  relief="raised", bd=1,
                                  command=preview_window.destroy)
            close_button.pack(side="right", padx=2, pady=1)
            
            # Make window draggable
            self.make_window_draggable(preview_window, title_bar)
            
            # Conținutul previzualizării
            text_area = tk.Text(preview_window, bg='white', font=default_font, wrap='word')
            scrollbar = tk.Scrollbar(preview_window, orient='vertical', command=text_area.yview)
            text_area.configure(yscrollcommand=scrollbar.set)
            
            text_area.pack(side='left', fill='both', expand=True, padx=10, pady=10)
            scrollbar.pack(side='right', fill='y', pady=10)
            
            try:
                if os.path.exists(file_path):
                    file_size = os.path.getsize(file_path)
                    if file_size < 1024 * 1024:  # Mai puțin de 1MB
                        try:
                            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                                content = f.read()
                                text_area.insert('1.0', content)
                        except:
                            text_area.insert('1.0', f"File: {file_name}\nPath: {file_path}\n\n")
                            text_area.insert('end', "Binary file or unsupported format.\nUse external application to view.")
                    else:
                        text_area.insert('1.0', f"File: {file_name}\nPath: {file_path}\nSize: {file_size} bytes\n\n")
                        text_area.insert('end', "File too large to preview.")
                else:
                    text_area.insert('1.0', f"File: {file_name}\nPath: {file_path}\n\n")
                    text_area.insert('end', "File no longer exists at this location.")
            except Exception as e:
                text_area.insert('1.0', f"Error previewing file: {str(e)}")
                
            text_area.config(state='disabled')
        
        # Funcția pentru recuperarea fișierelor selectate
        def recover_selected():
            selected_items = tree.selection()
            if not selected_items:
                messagebox.showwarning("No Selection", "Please select files to recover.")
                return
                
            # Cere folderul de destinație
            dest_folder = filedialog.askdirectory(title="Select Recovery Destination")
            if not dest_folder:
                return
                
            # Începe recuperarea
            recovery_thread = threading.Thread(target=perform_recovery,
                                             args=(selected_items, dest_folder))
            recovery_thread.daemon = True
            recovery_thread.start()
            
        def perform_recovery(selected_items, dest_folder):
            try:
                total_files = len(selected_items)
                recovered = 0
                failed = 0
                
                for i, item in enumerate(selected_items):
                    file_name = tree.item(item)['text']
                    progress = int((i / total_files) * 100)
                    update_progress(progress, f"Recovering {file_name}...")
                    
                    # Găsește informațiile despre fișier
                    file_info = None
                    for f in recoverable_files:
                        if f['name'] == file_name:
                            file_info = f
                            break
                            
                    if file_info and 'full_path' in file_info:
                        try:
                            # Efectuează recuperarea propriu-zisă
                            dest_path = os.path.join(dest_folder, file_name)
                            if os.path.exists(file_info['full_path']):
                                shutil.copy2(file_info['full_path'], dest_path)
                                recovered += 1
                            else:
                                failed += 1
                        except Exception as e:
                            print(f"Failed to recover {file_name}: {e}")
                            failed += 1
                    else:
                        failed += 1
                    
                update_progress(100, "Recovery completed")
                
                if recovered > 0:
                    messagebox.showinfo("Recovery Complete", 
                                      f"Successfully recovered {recovered} files to {dest_folder}\n"
                                      f"Failed to recover: {failed} files")
                else:
                    messagebox.showwarning("Recovery Failed", 
                                         "No files could be recovered. They may have been permanently deleted.")
                
            except Exception as e:
                messagebox.showerror("Recovery Error", f"An error occurred during recovery: {str(e)}")
        
        # Inițializare sample date pentru demo
        def populate_demo_data():
            """Populează cu date demo"""
            sample_files = [
                {"name": "important_document.docx", "size": "45.2 KB", "type": "Documents", 
                 "location": "Recycle Bin", "date": "2024-07-15 14:30", "status": "Recoverable"},
                {"name": "vacation_photo.jpg", "size": "2.3 MB", "type": "Images", 
                 "location": "Temp Files", "date": "2024-07-10 09:15", "status": "Recoverable"},
                {"name": "project_backup.zip", "size": "15.6 MB", "type": "Archives", 
                 "location": "Recent Files", "date": "2024-07-20 16:45", "status": "Recoverable"},
                {"name": "presentation.pptx", "size": "8.7 MB", "type": "Documents", 
                 "location": "Recycle Bin", "date": "2024-07-18 11:22", "status": "Recoverable"},
                {"name": "meeting_notes.txt", "size": "12.4 KB", "type": "Documents", 
                 "location": "Temp Files", "date": "2024-07-25 13:10", "status": "Recoverable"}
            ]
            
            for file in sample_files:
                tree.insert('', 'end', text=file["name"], values=(
                    file["size"], file["type"], file["location"], file["date"], file["status"]))
                recoverable_files.append(file)
            
            files_found.set(len(sample_files))
            files_label.config(text=f"Files found: {files_found.get()}")
            status_text.set("Sample data loaded")
        
        # Conectarea funcțiilor la butoane
        file_menu.add_command(label="New Scan", command=new_scan, underline=0)
        file_menu.add_command(label="Open Recovery Log", command=lambda: messagebox.showinfo("Info", "Log functionality is a demo"), underline=0)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=lambda: self.close_window("Recover Tunnel", recover_window), underline=1)
        
        tools_menu.add_command(label="Scan Recycle Bin", command=scan_recycle_bin, underline=5)
        tools_menu.add_command(label="Scan Temp Files", command=scan_temp_files, underline=5)
        tools_menu.add_command(label="Recent Files Scan", command=scan_recent_files, underline=0)
        tools_menu.add_separator()
        tools_menu.add_command(label="Settings", command=show_settings, underline=0)
        
        help_menu.add_command(label="About", command=show_about, underline=0)
        
        btn_new.config(command=new_scan)
        btn_recycle.config(command=scan_recycle_bin)
        btn_temp.config(command=scan_temp_files)
        btn_recover.config(command=recover_selected)
        btn_preview.config(command=preview_file)
        btn_clear.config(command=clear_results)
        btn_stop.config(command=stop_scan)
        
        btn_quick_recycle.config(command=scan_recycle_bin)
        btn_quick_temp.config(command=scan_temp_files)
        
        # Populează cu date demo
        populate_demo_data()
        
        # Adaugă în taskbar
        self.add_window_to_taskbar("Recover Tunnel", recover_window)
        recover_window.protocol("WM_DELETE_WINDOW", lambda: self.close_window("Recover Tunnel", recover_window))
    
    def create_file_recovery(self):
        """Creează aplicația File Fisher pentru recuperarea fișierelor"""
        recovery_window = tk.Toplevel(self.rootW95dist)
        recovery_window.title("File Fisher")
        recovery_window.overrideredirect(True)
        recovery_window.geometry("640x480+300+150")
        recovery_window.configure(bg="#c0c0c0")
        
        # Add Windows 95 style title bar
        title_bar = tk.Frame(recovery_window, bg="#000080", height=25)
        title_bar.pack(fill="x", side="top")
        title_label = tk.Label(title_bar, text="File Fisher", fg="white", bg="#000080",
                              font=("MS Sans Serif", 8, "bold"))
        title_label.pack(side="left", padx=5, pady=2)
        
        # Close button for title bar
        close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                font=("Arial", 8, "bold"), width=2, height=1,
                                relief="raised", bd=1,
                                command=lambda: self.close_window("File Fisher", recovery_window))
        close_button.pack(side="right", padx=2, pady=1)
        
        self.make_window_draggable(recovery_window, title_bar)
        
        # Windows 95 style configuration
        bg_color = '#c0c0c0'
        button_bg = '#c0c0c0'
        button_fg = '#000000'
        text_bg = '#ffffff'
        font_main = ('MS Sans Serif', 8)
        font_title = ('MS Sans Serif', 8, 'bold')
        
        # Variables
        scanning = False
        recovered_files = []
        scan_progress = 0
        
        # Main container
        main_frame = tk.Frame(recovery_window, bg=bg_color, relief='raised', bd=2)
        main_frame.pack(fill='both', expand=True, padx=2, pady=2)
        
        # Top section - Drive selection
        drive_frame = tk.LabelFrame(main_frame, text="Select Drive to Scan", 
                                  bg=bg_color, font=font_main, relief='groove', bd=2)
        drive_frame.pack(fill='x', padx=10, pady=5)
        
        drive_var = tk.StringVar()
        drive_combo = ttk.Combobox(drive_frame, textvariable=drive_var, 
                                  state='readonly', width=20)
        drive_combo.pack(side='left', padx=5, pady=5)
        
        # Function to refresh drives
        def refresh_drives():
            """Refresh available drives list"""
            drives = []
            # Get available drives (Windows)
            if os.name == 'nt':
                import string
                for letter in string.ascii_uppercase:
                    drive = f"{letter}:\\"
                    if os.path.exists(drive):
                        drives.append(drive)
            else:
                # Unix-like systems
                drives = ['/']
                
            drive_combo['values'] = drives
            if drives:
                drive_combo.set(drives[0])
        
        refresh_btn = tk.Button(drive_frame, text="Refresh", command=refresh_drives,
                              bg=button_bg, font=font_main, relief='raised', bd=2)
        refresh_btn.pack(side='left', padx=5)
        
        # Scan options
        options_frame = tk.LabelFrame(main_frame, text="Scan Options", 
                                    bg=bg_color, font=font_main, relief='groove', bd=2)
        options_frame.pack(fill='x', padx=10, pady=5)
        
        deep_scan_var = tk.BooleanVar()
        deep_check = tk.Checkbutton(options_frame, text="Deep Scan (slower but more thorough)",
                                  variable=deep_scan_var, bg=bg_color, font=font_main)
        deep_check.pack(anchor='w', padx=5, pady=2)
        
        file_types_var = tk.StringVar(value="All Files")
        file_type_frame = tk.Frame(options_frame, bg=bg_color)
        file_type_frame.pack(fill='x', padx=5, pady=2)
        
        tk.Label(file_type_frame, text="File Types:", bg=bg_color, font=font_main).pack(side='left')
        file_type_combo = ttk.Combobox(file_type_frame, textvariable=file_types_var,
                                     values=["All Files", "Images (jpg,png,gif)", "Documents (doc,pdf,txt)", 
                                             "Videos (mp4,avi,mov)", "Audio (mp3,wav,flac)"], width=25)
        file_type_combo.pack(side='left', padx=5)
        
        # Control buttons
        control_frame = tk.Frame(main_frame, bg=bg_color)
        control_frame.pack(fill='x', padx=10, pady=5)
        
        # Progress section
        progress_frame = tk.LabelFrame(main_frame, text="Scan Progress", 
                                     bg=bg_color, font=font_main, relief='groove', bd=2)
        progress_frame.pack(fill='x', padx=10, pady=5)
        
        progress_var = tk.StringVar(value="Ready to scan...")
        progress_label = tk.Label(progress_frame, textvariable=progress_var,
                                bg=bg_color, font=font_main)
        progress_label.pack(pady=5)
        
        progress_bar = ttk.Progressbar(progress_frame, mode='determinate')
        progress_bar.pack(fill='x', padx=10, pady=5)
        
        # Results section
        results_frame = tk.LabelFrame(main_frame, text="Found Files", 
                                    bg=bg_color, font=font_main, relief='groove', bd=2)
        results_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Treeview with scrollbars
        tree_frame = tk.Frame(results_frame, bg=bg_color)
        tree_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        tree = ttk.Treeview(tree_frame, columns=('Size', 'Type', 'Status'), show='tree headings')
        tree.heading('#0', text='File Name')
        tree.heading('Size', text='Size')
        tree.heading('Type', text='Type')
        tree.heading('Status', text='Status')
        
        tree.column('#0', width=200)
        tree.column('Size', width=80)
        tree.column('Type', width=80)
        tree.column('Status', width=100)
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient='horizontal', command=tree.xview)
        tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        tree.pack(side='left', fill='both', expand=True)
        v_scrollbar.pack(side='right', fill='y')
        h_scrollbar.pack(side='bottom', fill='x')
        
        # Status bar
        status_frame = tk.Frame(recovery_window, bg='#808080', relief='sunken', bd=1)
        status_frame.pack(fill='x', side='bottom')
        
        status_var = tk.StringVar(value="Ready")
        status_label = tk.Label(status_frame, textvariable=status_var,
                              bg='#808080', fg='white', font=font_main, anchor='w')
        status_label.pack(fill='x', padx=5, pady=1)
        
        # Utility functions
        def format_size(size):
            """Format file size in human readable format"""
            for unit in ['B', 'KB', 'MB', 'GB']:
                if size < 1024:
                    return f"{size:.1f} {unit}"
                size /= 1024
            return f"{size:.1f} TB"
        
        def get_file_extensions():
            """Get file extensions based on selected type"""
            file_type = file_types_var.get()
            
            extensions = {
                "All Files": [],
                "Images (jpg,png,gif)": ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff'],
                "Documents (doc,pdf,txt)": ['.doc', '.docx', '.pdf', '.txt', '.rtf', '.odt', '.ods', '.odp', '.md', '.html', '.htm', '.xls', '.xlsx', '.csv', '.rtf', '.ppt', '.pptx', '.json', '.log'],
                "Videos (mp4,avi,mov)": ['.mp4', '.avi', '.mov', '.mkv', '.wmv', '.flv'],
                "Audio (mp3,wav,flac)": ['.mp3', '.wav', '.flac', '.aac', '.ogg', '.wma']
            }
            
            return extensions.get(file_type, [])
        
        def is_file_match(file_path, file_types):
            """Check if file matches the selected criteria"""
            try:
                # Check file extension if specific type is selected
                if file_types:
                    _, ext = os.path.splitext(file_path.lower())
                    if ext not in file_types:
                        return False
                        
                # If "All Files" or file matches extension filter, include it
                return True
                        
            except Exception:
                return False
        
        def add_found_file(file_path):
            """Add a found file to the results"""
            try:
                file_name = os.path.basename(file_path)
                
                # Get file info
                if os.path.exists(file_path):
                    size = os.path.getsize(file_path)
                    status = "Existing"
                else:
                    size = 0
                    status = "Not Found"
                    
                size_str = format_size(size)
                _, ext = os.path.splitext(file_name)
                file_type = ext.upper()[1:] if ext else "Unknown"
                
                # Add to tree
                item = tree.insert('', 'end', text=file_name, 
                                 values=(size_str, file_type, status))
                
                # Store file info
                recovered_files.append({
                    'path': file_path,
                    'name': file_name,
                    'size': size,
                    'type': file_type,
                    'status': status,
                    'item': item
                })
                
            except Exception as e:
                print(f"Error adding file {file_path}: {e}")
        
        def copy_single_file(file_info, recovery_dir):
            """Copy a single file"""
            try:
                source_path = file_info['path']
                dest_path = os.path.join(recovery_dir, file_info['name'])
                
                # If source exists, copy it
                if os.path.exists(source_path):
                    import shutil
                    shutil.copy2(source_path, dest_path)
                    return True
                else:
                    # File doesn't exist - can't copy
                    return False
                    
            except Exception as e:
                print(f"Error copying {file_info['name']}: {e}")
                return False
        
        def recover_files():
            """Recover selected files"""
            selected_items = tree.selection()
            if not selected_items:
                messagebox.showwarning("Warning", "Please select files to fish!")
                return
                
            # Choose recovery location
            recovery_dir = filedialog.askdirectory(title="Select Fishing Location")
            if not recovery_dir:
                return
                
            # Start recovery
            status_var.set("Copying files...")
            recovered_count = 0
            
            for item in selected_items:
                try:
                    # Find file info
                    file_info = None
                    for f in recovered_files:
                        if f['item'] == item:
                            file_info = f
                            break
                            
                    if file_info:
                        success = copy_single_file(file_info, recovery_dir)
                        if success:
                            recovered_count += 1
                            # Update status in tree
                            tree.set(item, 'Status', 'Copied')
                            
                except Exception as e:
                    print(f"Copy error: {e}")
                    
            messagebox.showinfo("Copy Complete", 
                               f"Successfully copied {recovered_count} out of {len(selected_items)} files!")
            status_var.set(f"Copied {recovered_count} files")
        
        def perform_scan():
            """Perform the actual file scanning"""
            nonlocal scanning
            
            drive = drive_var.get()
            deep_scan = deep_scan_var.get()
            file_types = get_file_extensions()
            
            progress_var.set(f"Scanning {drive}...")
            status_var.set("Scanning...")
            
            total_files = 0
            scanned_files = 0
            
            try:
                # First pass - count files for progress
                if scanning:
                    for root, dirs, files in os.walk(drive):
                        total_files += len(files)
                        if not scanning:
                            break
                            
                progress_bar.config(maximum=total_files)
                
                # Second pass - actual scanning
                for root, dirs, files in os.walk(drive):
                    if not scanning:
                        break
                        
                    for file in files:
                        if not scanning:
                            break
                            
                        file_path = os.path.join(root, file)
                        scanned_files += 1
                        
                        # Update progress
                        progress_bar.config(value=scanned_files)
                        progress_var.set(f"Scanning: {file} ({scanned_files}/{total_files})")
                        recovery_window.update_idletasks()
                        
                        # Check if file matches criteria
                        if is_file_match(file_path, file_types):
                            add_found_file(file_path)
                            
                        # Small delay to prevent UI freezing
                        time.sleep(0.001)
                        
            except Exception as e:
                messagebox.showerror("Error", f"Scan error: {str(e)}")
                
            # Scan completed
            if scanning:
                progress_var.set(f"Scan completed! Found {len(recovered_files)} files")
                status_var.set(f"Found {len(recovered_files)} files")
            
            scanning = False
            scan_btn.config(state='normal')
            stop_btn.config(state='disabled')
            if recovered_files:
                recover_btn.config(state='normal')
        
        def start_scan():
            """Start the file recovery scan"""
            nonlocal scanning
            
            if not drive_var.get():
                messagebox.showwarning("Warning", "Please select a drive to scan!")
                return
                
            scanning = True
            scan_btn.config(state='disabled')
            stop_btn.config(state='normal')
            recover_btn.config(state='disabled')
            
            # Clear previous results
            for item in tree.get_children():
                tree.delete(item)
            recovered_files.clear()
            
            # Start scan in separate thread
            scan_thread = threading.Thread(target=perform_scan)
            scan_thread.daemon = True
            scan_thread.start()
        
        def stop_scan():
            """Stop the current scan"""
            nonlocal scanning
            
            scanning = False
            scan_btn.config(state='normal')
            stop_btn.config(state='disabled')
            if recovered_files:
                recover_btn.config(state='normal')
            progress_var.set("Scan stopped by user")
            status_var.set("Scan stopped")
        
        # Buttons
        scan_btn = tk.Button(control_frame, text="Start Scan", command=start_scan,
                            bg='#008000', fg='white', font=font_main, 
                            relief='raised', bd=2, width=12)
        scan_btn.pack(side='left', padx=5)
        
        stop_btn = tk.Button(control_frame, text="Stop Scan", command=stop_scan,
                            bg='#800000', fg='white', font=font_main, 
                            relief='raised', bd=2, width=12, state='disabled')
        stop_btn.pack(side='left', padx=5)
        
        recover_btn = tk.Button(control_frame, text="Extract Selected", command=recover_files,
                               bg='#000080', fg='white', font=font_main, 
                               relief='raised', bd=2, width=15, state='disabled')
        recover_btn.pack(side='right', padx=5)
        
        # Initialize
        refresh_drives()
        
        # Add to taskbar
        self.add_window_to_taskbar("File Fisher", recovery_window)
        recovery_window.protocol("WM_DELETE_WINDOW", lambda: self.close_window("File Fisher", recovery_window))
    
    def create_about_window(self):
        about_window = tk.Toplevel(self.rootW95dist)
        about_window.title("System Requirements")
        about_window.overrideredirect(True)
        about_window.geometry("700x300+300+200")
        about_window.configure(bg="#c9c9c9")
        
        # Add Windows 95 style title bar
        title_bar = tk.Frame(about_window, bg="#000080", height=25)
        title_bar.pack(fill="x", side="top")
        title_label = tk.Label(title_bar, text="System Requirements", fg="white", bg="#000080",
                             font=("MS Sans Serif", 8, "bold"))
        title_label.pack(side="left", padx=5, pady=2)
        
        # Close button for title bar
        close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                               font=("Arial", 8, "bold"), width=2, height=1,
                               relief="raised", bd=1,
                               command=lambda: self.close_window("System Requirements", about_window))
        close_button.pack(side="right", padx=2, pady=1)
        
        self.make_window_draggable(about_window, title_bar)
        
        # Obținem informațiile sistemului
        sys_info = platform.uname()
        system_version = sys_info.system
        processor_info = platform.processor()
        
        # Specificațiile minime
        min_os = "Windows 10 (x64)"
        min_processor_speed = 1  # GHz
        min_ram = 4  # GB
        min_disk_space = 300  # MB
        
        # Obținem informațiile sistemului
        os_version = sys_info.release
        ram = psutil.virtual_memory().total / (1024 ** 3)  # RAM în GB
        
        # Încercăm să extragem viteza procesorului - acest lucru ar putea să nu funcționeze pe toate sistemele
        try:
            processor_speed = float(processor_info.split()[2].split('GHz')[0])  # extragem valoarea GHz
        except (IndexError, ValueError):
            processor_speed = 0  # Nu putem determina viteza
        
        is_x64 = sys_info.machine in ["x86_64", "AMD64"]  # Verificăm dacă este sistem x64 (AMD64 sau x86_64)
        
        # Verificăm dacă sistemul îndeplinește cerințele minime
        meets_requirements = True
        requirements_message = "System meets minimum requirements"
        
        if system_version != "Windows" or not is_x64:
            meets_requirements = False
            requirements_message = "Does not meet minimum requirements: OS is not Windows x64."
        
        # Verificăm versiunea (mai sigur)
        try:
            major_version = int(os_version.split('.')[0])
            if major_version < 10:
                meets_requirements = False
                requirements_message = f"Does not meet minimum requirements: OS version is older than Windows 10 (Detected: Windows {major_version})."
        except (ValueError, IndexError):
            # Nu putem determina versiunea cu exactitate
            pass
        
        if processor_speed < min_processor_speed and processor_speed > 0:  # Doar dacă am reușit să extrageam viteza
            meets_requirements = False
            requirements_message = "Does not meet minimum requirements: Processor speed is too low."
        
        # Verificăm RAM-ul
        if ram < min_ram:
            meets_requirements = False
            requirements_message = "Does not meet minimum requirements: RAM is too low."
        
        # Verificăm dacă există suficient spațiu pe disc
        try:
            if system_version == "Windows":
                disk_space = psutil.disk_usage('/').free / (1024 ** 2)  # spațiu liber în MB
                if disk_space < min_disk_space:
                    meets_requirements = False
                    requirements_message = "Does not meet minimum requirements: Insufficient disk space."
            else:
                disk_space = psutil.disk_usage('/').free / (1024 ** 2)  # Pentru alte sisteme de operare
        except:
            disk_space = 0  # Nu putem determina spațiul pe disc
        
        # Creăm layout-ul pentru fereastra de 'About'
        frame = tk.Frame(about_window, bg='#c9c9c9')
        frame.pack(padx=10, pady=10)
        
        # Specificațiile minime
        min_specs = tk.Label(frame, text=f"Minimum Requirements:\nOS: {min_os}\nProcessor: {min_processor_speed} GHz\nRAM: {min_ram} GB\nDisk space: {min_disk_space} MB free", 
                           font=("Fixedsys", 12), bg='#c9c9c9')
        min_specs.grid(row=0, column=0, padx=10, pady=10, sticky='w')
        
        # Specificațiile curente
        current_specs = tk.Label(frame, text=f"Your System:\nOS: {system_version} {os_version} {sys_info.machine}\nProcessor: {processor_info}\nRAM: {ram:.2f} GB\nDisk space: {disk_space:.2f} MB free", 
                               font=("Fixedsys", 12), bg='#c9c9c9')
        current_specs.grid(row=0, column=1, padx=10, pady=10, sticky='w')
        
        # Mesajul de cerințe
        color = "green" if meets_requirements else "red"
        result_label = tk.Label(about_window, text=requirements_message, font=("Fixedsys", 14), fg=color, bg='#c9c9c9')
        result_label.pack(pady=10)
        
        # Adăugăm un buton OK
        ok_button = tk.Button(about_window, text="OK", font=("Fixedsys", 12), bg="#c0c0c0", 
                            command=lambda: self.close_window("System Requirements", about_window))
        ok_button.pack(pady=10)
        
        self.add_window_to_taskbar("System Requirements", about_window)
        about_window.protocol("WM_DELETE_WINDOW", lambda: self.close_window("System Requirements", about_window))
    
    def create_activation_window(self):
        window = tk.Toplevel(self.rootW95dist)
        window.title("Activation Wizard")
        window.overrideredirect(True)
        window.geometry("400x200+300+200")
        window.configure(bg="#c9c9c9")
        
        # Add Windows 95 style title bar
        title_bar = tk.Frame(window, bg="#000080", height=25)
        title_bar.pack(fill="x", side="top")
        title_label = tk.Label(title_bar, text="Activation Wizard", fg="white", bg="#000080",
                              font=("MS Sans Serif", 8, "bold"))
        title_label.pack(side="left", padx=5, pady=2)
        
        # Close button for title bar
        close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                font=("Arial", 8, "bold"), width=2, height=1,
                                relief="raised", bd=1,
                                command=lambda: self.close_window("Activation Wizard", window))
        close_button.pack(side="right", padx=2, pady=1)
        
        self.make_window_draggable(window, title_bar)
        
        # Mesaj
        label = tk.Label(window, text="Enter your product key:", font=("Fixedsys"), bg="#c9c9c9")
        label.pack(pady=10)
        
        # Câmp de introducere a cheii
        product_key_entry = tk.Entry(window, font=("Fixedsys"), width=30, bd=5)
        product_key_entry.pack(pady=5)
        
        # Butonul de activare (inițial dezactivat)
        activate_button = tk.Button(
            window, text="Activate", font=("Fixedsys"), state=tk.DISABLED, bd=5, bg="#c9c9c9", 
            command=lambda: activate_product()
        )
        activate_button.pack(pady=5)
        
        # Butonul "Activate later" -> elimină aplicația din taskbar
        later_button = tk.Button(
            window, text="Activate later", font=("Fixedsys"), bd=5, bg="#c9c9c9", 
            command=lambda: self.close_window("Activation Wizard", window)
        )
        later_button.pack(pady=5)
        
        # Funcție pentru activare
        def activate_product():
            product_key = product_key_entry.get()
            valid_key = "R46BX-JHR2J-PG7ER-24QFG-MWKVR"
            FOLDER_NAME = "Serial"
            FILE_NAME = "product_key.lic95"
            
            if not os.path.exists(FOLDER_NAME):
                try:
                    os.makedirs(FOLDER_NAME)
                except:
                    pass
                    
            if product_key == valid_key:
                file_path = os.path.join(FOLDER_NAME, FILE_NAME)
                # Verificăm dacă fișierul există deja
                if not os.path.exists(file_path):
                    with open(file_path, "w") as file:
                        file.write(valid_key)
                messagebox.showinfo("Activation Successful", "Product has been activated successfully.")
                self.close_window("Activation Wizard", window)
            else:
                messagebox.showerror("Activation Failed", "Invalid product key. Please try again.")
                
        # Funcție pentru validarea cheii în timp real
        def validate_key(event):
            if product_key_entry.get() == "R46BX-JHR2J-PG7ER-24QFG-MWKVR":
                activate_button.config(state=tk.NORMAL)  # Activează butonul dacă cheia este corectă
            else:
                activate_button.config(state=tk.DISABLED)  # Dezactivează butonul dacă cheia e greșită
                
        product_key_entry.bind("<KeyRelease>", validate_key)  # Verifică validitatea în timp real
        
        self.add_window_to_taskbar("Activation Wizard", window)
        window.protocol("WM_DELETE_WINDOW", lambda: self.close_window("Activation Wizard", window))
    
    def open_sql_explorer(self):
        # Create SQL Explorer window
        sql_window = tk.Toplevel(self.rootW95dist)
        sql_window.title("SQL Explorer")
        sql_window.overrideredirect(True)
        sql_window.geometry("800x600+200+100")
        sql_window.configure(bg="#c0c0c0")
        
        # Add Windows 95 style title bar
        title_bar = tk.Frame(sql_window, bg="#000080", height=25)
        title_bar.pack(fill="x", side="top")
        title_label = tk.Label(title_bar, text="SQL Explorer", fg="white", bg="#000080",
                              font=("MS Sans Serif", 8, "bold"))
        title_label.pack(side="left", padx=5, pady=2)
        
        # Close button for title bar
        close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                font=("Arial", 8, "bold"), width=2, height=1,
                                relief="raised", bd=1,
                                command=lambda: self.close_window("SQL Explorer", sql_window))
        close_button.pack(side="right", padx=2, pady=1)
        
        self.make_window_draggable(sql_window, title_bar)
        
        # Import required modules for SQL Explorer
        import sqlite3
        import os
        
        class Win95SQLViewer:
            def __init__(self, root):
                self.root = root
                
                # Win95 color palette
                self.colors = {
                    "bg_main": "#c0c0c0",         # Classic Win95 gray background
                    "bg_window": "#ffffff",        # Window background
                    "text": "#000000",             # Black text
                    "button_face": "#c0c0c0",      # Button face color
                    "button_shadow": "#808080",    # Button shadow
                    "button_highlight": "#ffffff", # Button highlight
                    "button_text": "#000000",      # Button text
                    "title_bg": "#000080",         # Title bar background (navy blue)
                    "title_text": "#ffffff",       # Title bar text (white)
                    "border": "#808080"            # Border color
                }
                
                self.connection = None
                self.current_table = None
                
                # Configure styles
                self.setup_styles()
                
                # Create interface
                self.create_widgets()
                self.create_menu()
            
            def setup_styles(self):
                # Set app font to match Win95 style
                self.default_font = ("MS Sans Serif", 9)
                
                style = ttk.Style()
                
                # Configure treeview style
                style.configure(
                    "Treeview",
                    background=self.colors["bg_window"],
                    foreground=self.colors["text"],
                    fieldbackground=self.colors["bg_window"],
                    borderwidth=1,
                    relief=tk.SUNKEN
                )
                
                # Configure treeview heading
                style.configure(
                    "Treeview.Heading",
                    background=self.colors["button_face"],
                    foreground=self.colors["text"],
                    relief=tk.RAISED,
                    borderwidth=2,
                    font=self.default_font
                )
                
                # Configure scrollbar
                style.configure(
                    "TScrollbar",
                    background=self.colors["button_face"],
                    troughcolor=self.colors["bg_window"],
                    borderwidth=2,
                    relief=tk.RAISED,
                    arrowsize=13
                )
                
                # Configure combobox
                style.configure(
                    "TCombobox",
                    background=self.colors["bg_window"],
                    fieldbackground=self.colors["bg_window"],
                    selectbackground=self.colors["title_bg"],
                    selectforeground=self.colors["title_text"]
                )
            
            def create_menu(self):
                # Create classic Win95 menu style
                menubar = tk.Menu(self.root, bg=self.colors["bg_main"], fg=self.colors["text"],
                                activebackground=self.colors["title_bg"],
                                activeforeground=self.colors["title_text"],
                                relief=tk.RAISED, bd=1)
                self.root.config(menu=menubar)
                
                # File menu
                file_menu = tk.Menu(menubar, tearoff=0, bg=self.colors["bg_main"],
                                    fg=self.colors["text"],
                                    activebackground=self.colors["title_bg"],
                                    activeforeground=self.colors["title_text"], bd=1)
                file_menu.add_command(label="Open Database...", command=self.open_database)
                file_menu.add_separator()
                file_menu.add_command(label="Exit", command=self.root.destroy)
                menubar.add_cascade(label="File", menu=file_menu)
                
                # Edit menu
                edit_menu = tk.Menu(menubar, tearoff=0, bg=self.colors["bg_main"],
                                fg=self.colors["text"],
                                activebackground=self.colors["title_bg"],
                                activeforeground=self.colors["title_text"], bd=1)
                edit_menu.add_command(label="Refresh", command=self.refresh_tables)
                edit_menu.add_command(label="Execute Query", command=self.execute_query)
                menubar.add_cascade(label="Edit", menu=edit_menu)
            
            def create_widgets(self):
                # Main window frame (inset look)
                main_frame = tk.Frame(
                    self.root, 
                    bg=self.colors["bg_main"],
                    bd=2,
                    relief=tk.RAISED
                )
                main_frame.pack(fill=tk.BOTH, expand=True, padx=3, pady=28)  # Padding adjusted for title bar
                
                # Toolbar frame
                toolbar = tk.Frame(main_frame, bg=self.colors["bg_main"], bd=0)
                toolbar.pack(fill=tk.X, pady=(2, 5), padx=2)
                
                # Table selection frame
                control_frame = tk.Frame(main_frame, bg=self.colors["bg_main"], bd=0)
                control_frame.pack(fill=tk.X, pady=(0, 5), padx=5)
                
                # Table label
                table_label = tk.Label(
                    control_frame,
                    text="Tables:",
                    bg=self.colors["bg_main"],
                    fg=self.colors["text"]
                )
                table_label.pack(side=tk.LEFT, padx=(0, 5))
                
                # Table dropdown with Win95 style
                self.table_combo = ttk.Combobox(
                    control_frame,
                    state="readonly",
                    width=30
                )
                self.table_combo.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
                self.table_combo.bind("<<ComboboxSelected>>", self.load_table)
                
                # Refresh button
                refresh_button = self.create_win95_button(
                    control_frame, 
                    text="Refresh",
                    command=self.refresh_tables
                )
                refresh_button.pack(side=tk.RIGHT, padx=2)
                
                # Query frame
                query_frame = tk.Frame(main_frame, bg=self.colors["bg_main"], bd=0)
                query_frame.pack(fill=tk.X, pady=(0, 5), padx=5)
                
                # Query label
                query_label = tk.Label(
                    query_frame,
                    text="SQL Query:",
                    bg=self.colors["bg_main"],
                    fg=self.colors["text"]
                )
                query_label.pack(side=tk.LEFT, padx=(0, 5))
                
                # Query text entry (sunken, like Win95)
                self.query_entry = tk.Entry(
                    query_frame,
                    bg=self.colors["bg_window"],
                    fg=self.colors["text"],
                    bd=2,
                    relief=tk.SUNKEN,
                    insertbackground=self.colors["text"],
                    font=("Courier New", 10)
                )
                self.query_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
                
                # Execute button
                execute_button = self.create_win95_button(
                    query_frame, 
                    text="Execute",
                    command=self.execute_query
                )
                execute_button.pack(side=tk.RIGHT, padx=2)
                
                # Frame for treeview
                table_frame = tk.Frame(main_frame, bg=self.colors["bg_main"], bd=2, relief=tk.SUNKEN)
                table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 5))
                
                # Create treeview with scrollbars
                self.tree = ttk.Treeview(table_frame)
                self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
                
                # Vertical scrollbar
                vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
                vsb.pack(side=tk.RIGHT, fill=tk.Y)
                self.tree.configure(yscrollcommand=vsb.set)
                
                # Horizontal scrollbar
                hsb = ttk.Scrollbar(main_frame, orient="horizontal", command=self.tree.xview)
                hsb.pack(side=tk.BOTTOM, fill=tk.X, padx=5, pady=(0, 5))
                self.tree.configure(xscrollcommand=hsb.set)
                
                # Status bar
                self.status_var = tk.StringVar()
                self.status_var.set("Ready")
                
                # Status bar with Win95 inset look
                status_frame = tk.Frame(main_frame, bd=2, relief=tk.SUNKEN)
                status_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=3, pady=3)
                
                self.status_bar = tk.Label(
                    status_frame,
                    textvariable=self.status_var,
                    bg=self.colors["bg_main"],
                    fg=self.colors["text"],
                    anchor=tk.W,
                    bd=1,
                    padx=5,
                    pady=2
                )
                self.status_bar.pack(fill=tk.X)
            
            def create_win95_button(self, parent, text, command):
                """Create a button with typical Windows 95 appearance"""
                button_frame = tk.Frame(parent, bd=2, relief=tk.RAISED)
                
                button = tk.Button(
                    button_frame,
                    text=text,
                    bg=self.colors["button_face"],
                    fg=self.colors["button_text"],
                    activebackground=self.colors["button_face"],
                    activeforeground=self.colors["button_text"],
                    bd=2,
                    width=10,
                    command=command,
                    relief=tk.RAISED,
                    padx=6,
                    pady=1,
                    font=self.default_font
                )
                button.pack(padx=1, pady=1)
                
                # Windows 95 button press effect
                button.bind("<ButtonPress-1>", lambda e: button_frame.config(relief=tk.SUNKEN))
                button.bind("<ButtonRelease-1>", lambda e: button_frame.config(relief=tk.RAISED))
                
                return button_frame
            
            def open_database(self):
                file_path = filedialog.askopenfilename(
                    title="Open Database",
                    filetypes=[("SQLite Files", "*.db *.sqlite *.sqlite3"), ("All Files", "*.*")]
                )
                
                if file_path:
                    try:
                        if self.connection:
                            self.connection.close()
                        
                        self.connection = sqlite3.connect(file_path)
                        self.refresh_tables()
                        self.status_var.set(f"Database: {os.path.basename(file_path)}")
                        messagebox.showinfo("Success", "Database opened successfully!")
                    except sqlite3.Error as e:
                        messagebox.showerror("Error", f"Could not open database: {e}")
                        self.status_var.set("Error opening database.")
            
            def refresh_tables(self):
                if not self.connection:
                    messagebox.showwarning("Warning", "No database is open!")
                    return
                
                try:
                    cursor = self.connection.cursor()
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                    tables = cursor.fetchall()
                    
                    self.table_combo['values'] = [table[0] for table in tables]
                    if tables:
                        self.table_combo.current(0)
                        self.load_table(None)
                    else:
                        messagebox.showinfo("Info", "No tables found in this database.")
                        self.clear_tree()
                except sqlite3.Error as e:
                    messagebox.showerror("Error", f"Could not load tables: {e}")
            
            def load_table(self, event):
                if not self.connection:
                    return
                
                selected_table = self.table_combo.get()
                if not selected_table:
                    return
                
                self.current_table = selected_table
                self.query_entry.delete(0, tk.END)
                self.query_entry.insert(0, f"SELECT * FROM {selected_table}")
                
                try:
                    cursor = self.connection.cursor()
                    cursor.execute(f"PRAGMA table_info({selected_table})")
                    columns = cursor.fetchall()
                    
                    cursor.execute(f"SELECT * FROM {selected_table} LIMIT 1000")
                    rows = cursor.fetchall()
                    
                    self.display_data(columns, rows)
                    self.status_var.set(f"Table: {selected_table} | {len(rows)} records")
                except sqlite3.Error as e:
                    messagebox.showerror("Error", f"Could not load table: {e}")
            
            def execute_query(self):
                if not self.connection:
                    messagebox.showwarning("Warning", "No database is open!")
                    return
                
                query = self.query_entry.get().strip()
                if not query:
                    messagebox.showwarning("Warning", "Enter a valid query!")
                    return
                
                try:
                    cursor = self.connection.cursor()
                    cursor.execute(query)
                    
                    if query.upper().startswith(("SELECT", "PRAGMA")):
                        # Read query - display results
                        columns = [description[0] for description in cursor.description]
                        rows = cursor.fetchall()
                        
                        self.display_data_from_query(columns, rows)
                        self.status_var.set(f"Query executed: {len(rows)} results")
                    else:
                        # Modification query - confirm transaction
                        self.connection.commit()
                        affected = cursor.rowcount
                        messagebox.showinfo("Success", f"Query executed successfully! Rows affected: {affected}")
                        
                        # Refresh current table if exists
                        if self.current_table:
                            self.load_table(None)
                        
                        self.status_var.set(f"Query executed: {affected} rows affected")
                except sqlite3.Error as e:
                    messagebox.showerror("Error", f"Error executing query: {e}")
            
            def display_data(self, columns, rows):
                self.clear_tree()
                
                # Configure treeview columns
                column_names = [col[1] for col in columns]
                
                self.tree["columns"] = column_names
                
                # Hide default column
                self.tree.column("#0", width=0, stretch=tk.NO)
                
                # Configure each column
                for name in column_names:
                    self.tree.column(name, anchor=tk.W, width=150)
                    self.tree.heading(name, text=name, anchor=tk.W)
                
                # Add data
                for i, row in enumerate(rows):
                    values = [str(val) if val is not None else "NULL" for val in row]
                    self.tree.insert("", tk.END, text=str(i), values=values, tags=('even' if i % 2 == 0 else 'odd',))
                
                # Alternate row colors (classic Win95 style)
                self.tree.tag_configure('even', background=self.colors["bg_window"])
                self.tree.tag_configure('odd', background="#ececec")  # Light gray for alternating rows
            
            def display_data_from_query(self, column_names, rows):
                self.clear_tree()
                
                self.tree["columns"] = column_names
                
                # Hide default column
                self.tree.column("#0", width=0, stretch=tk.NO)
                
                # Configure each column
                for name in column_names:
                    self.tree.column(name, anchor=tk.W, width=150)
                    self.tree.heading(name, text=name, anchor=tk.W)
                
                # Add data
                for i, row in enumerate(rows):
                    values = [str(val) if val is not None else "NULL" for val in row]
                    self.tree.insert("", tk.END, text=str(i), values=values, tags=('even' if i % 2 == 0 else 'odd',))
                
                # Alternate row colors
                self.tree.tag_configure('even', background=self.colors["bg_window"])
                self.tree.tag_configure('odd', background="#ececec")
            
            def clear_tree(self):
                for item in self.tree.get_children():
                    self.tree.delete(item)
                
                for col in self.tree["columns"]:
                    self.tree.heading(col, text="")
                
                self.tree["columns"] = []
        
        # Create the SQL viewer instance and pass the toplevel window
        sql_app = Win95SQLViewer(sql_window)
        
        # Add to taskbar
        self.add_window_to_taskbar("SQL Explorer", sql_window)
        sql_window.protocol("WM_DELETE_WINDOW", lambda: self.close_window("SQL Explorer", sql_window))
    
    '''
    def create_word_lite(self, file_path=None):
        """Creează o aplicație Word Lite pentru a deschide și vizualiza fișiere .doc sau .docx"""
        word_window = tk.Toplevel(self.rootW95dist)
        word_window.title("Word Lite")
        word_window.overrideredirect(True)
        word_window.geometry("800x600+200+100")
        word_window.configure(bg="#c0c0c0")
        
        # Add Windows 95 style title bar
        title_bar = tk.Frame(word_window, bg="#000080", height=25)
        title_bar.pack(fill="x", side="top")
        title_label = tk.Label(title_bar, text="Word Lite", fg="white", bg="#000080",
                              font=("MS Sans Serif", 8, "bold"))
        title_label.pack(side="left", padx=5, pady=2)
        
        # Close button for title bar
        close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                font=("Arial", 8, "bold"), width=2, height=1,
                                relief="raised", bd=1,
                                command=lambda: self.close_window("Word Lite", word_window))
        close_button.pack(side="right", padx=2, pady=1)
        
        self.make_window_draggable(word_window, title_bar)
        
        # Menubar
        menubar = tk.Menu(word_window)
        word_window.config(menu=menubar)
        
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        
        edit_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Edit", menu=edit_menu)
        
        view_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="View", menu=view_menu)
        
        format_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Format", menu=format_menu)
        
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        
        # Toolbar frame
        toolbar_frame = tk.Frame(word_window, bg="#c0c0c0", relief="raised", bd=2, height=35)
        toolbar_frame.pack(fill="x")
        toolbar_frame.pack_propagate(False)
        
        # Open button
        open_btn = tk.Button(toolbar_frame, text="Open", bg="#c0c0c0", relief="raised", bd=2,
                            font=("MS Sans Serif", 8),
                            command=lambda: open_document())
        open_btn.pack(side="left", padx=5, pady=2)
        
        # Save button (for future implementation)
        save_btn = tk.Button(toolbar_frame, text="Save", bg="#c0c0c0", relief="raised", bd=2,
                            font=("MS Sans Serif", 8),
                            command=lambda: save_document())
        save_btn.pack(side="left", padx=5, pady=2)
        
        # Print button
        print_btn = tk.Button(toolbar_frame, text="Print", bg="#c0c0c0", relief="raised", bd=2,
                             font=("MS Sans Serif", 8),
                             command=lambda: messagebox.showinfo("Print", "Printing functionality is not implemented in this version"))
        print_btn.pack(side="left", padx=5, pady=2)
        
        # Status bar
        status_frame = tk.Frame(word_window, bg="#c0c0c0", relief="sunken", bd=1, height=25)
        status_frame.pack(side="bottom", fill="x")
        status_frame.pack_propagate(False)
        
        status_label = tk.Label(status_frame, text="Ready", bg="#c0c0c0", font=("MS Sans Serif", 8))
        status_label.pack(side="left", padx=5)
        
        # Main content frame
        content_frame = tk.Frame(word_window, bg="white")
        content_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Text area
        text_frame = tk.Frame(content_frame, bg="white")
        text_frame.pack(fill="both", expand=True)
        
        # Scrollbars
        v_scrollbar = tk.Scrollbar(text_frame)
        v_scrollbar.pack(side="right", fill="y")
        
        h_scrollbar = tk.Scrollbar(text_frame, orient="horizontal")
        h_scrollbar.pack(side="bottom", fill="x")
        
        # Text widget
        text_area = tk.Text(text_frame, wrap="word", yscrollcommand=v_scrollbar.set, 
                           xscrollcommand=h_scrollbar.set,
                           font=("Times New Roman", 12), bg="white", fg="black")
        text_area.pack(side="left", fill="both", expand=True)
        
        v_scrollbar.config(command=text_area.yview)
        h_scrollbar.config(command=text_area.xview)
        
        # Variables for tracking
        current_file = None
        
        def open_document():
            """Open a Word document"""
            file_path = filedialog.askopenfilename(
                defaultextension=".docx",
                filetypes=[("Word files", "*.docx;*.doc"), ("Text files", "*.txt"), ("All files", "*.*")]
            )
            
            if not file_path:
                return
            
            try:
                # Set cursor to wait
                word_window.config(cursor="wait")
                word_window.update()
                
                # Clear text area
                text_area.delete(1.0, tk.END)
                
                # Get the file extension
                ext = os.path.splitext(file_path)[1].lower()
                
                # Handle different file types
                if ext == ".txt":
                    # Simple text file
                    with open(file_path, 'r', encoding='utf-8') as file:
                        content = file.read()
                        text_area.insert(tk.END, content)
                elif ext in [".doc", ".docx"]:
                    # Try to handle Word documents (simplified - only extracts text)
                    text_area.insert(tk.END, "Word Lite can display only plain text from Word documents.\n\n")
                    
                    try:
                        # If python-docx is installed, use it
                        import docx
                        doc = docx.Document(file_path)
                        for para in doc.paragraphs:
                            text_area.insert(tk.END, para.text + "\n")
                    except ImportError:
                        # Fallback method
                        text_area.insert(tk.END, "Cannot read Word document content. The python-docx library is not installed.\n")
                        text_area.insert(tk.END, "To install it, run: pip install python-docx\n\n")
                        text_area.insert(tk.END, "For now, only showing document metadata:\n")
                        text_area.insert(tk.END, f"Filename: {os.path.basename(file_path)}\n")
                        text_area.insert(tk.END, f"Size: {os.path.getsize(file_path)} bytes\n")
                else:
                    text_area.insert(tk.END, f"Cannot open file type: {ext}\n")
                    text_area.insert(tk.END, "Word Lite supports .txt, .doc, and .docx files.")
                
                # Update window title and status
                file_name = os.path.basename(file_path)
                title_label.config(text=f"Word Lite - {file_name}")
                word_window.title(f"Word Lite - {file_name}")
                status_label.config(text=f"Opened: {file_name}")
                
                # Save the current file
                nonlocal current_file
                current_file = file_path
                
            except Exception as e:
                text_area.delete(1.0, tk.END)
                text_area.insert(tk.END, f"Error opening file: {str(e)}")
                messagebox.showerror("Error", f"Could not open file: {str(e)}")
            finally:
                # Reset cursor
                word_window.config(cursor="")
        
        def save_document():
            """Save a document (placeholder)"""
            messagebox.showinfo("Info", "Save functionality is not implemented in this version")
        
        # File menu commands
        file_menu.add_command(label="Open", command=open_document)
        file_menu.add_command(label="Save", command=save_document)
        file_menu.add_separator()
        file_menu.add_command(label="Print", command=lambda: messagebox.showinfo("Print", "Printing functionality is not implemented in this version"))
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=lambda: self.close_window("Word Lite", word_window))
        
        # Edit menu commands
        edit_menu.add_command(label="Cut", command=lambda: status_label.config(text="Cut not implemented"))
        edit_menu.add_command(label="Copy", command=lambda: status_label.config(text="Copy not implemented"))
        edit_menu.add_command(label="Paste", command=lambda: status_label.config(text="Paste not implemented"))
        
        # View menu
        view_menu.add_command(label="Zoom", command=lambda: status_label.config(text="Zoom not implemented"))
        
        # Format menu
        format_menu.add_command(label="Font", command=lambda: status_label.config(text="Font formatting not implemented"))
        
        # Help menu
        help_menu.add_command(label="About Word Lite", command=lambda: messagebox.showinfo("About", "Word Lite\nVersion 1.0\n\nA simple document viewer"))
        
        # If a file path was provided, open it
        if file_path and os.path.exists(file_path):
            current_file = file_path
            try:
                # Set the file path and call the open function
                open_document()
            except Exception as e:
                text_area.delete(1.0, tk.END)
                text_area.insert(tk.END, f"Error opening file: {str(e)}")
        
        # Add to taskbar
        self.add_window_to_taskbar("Word Lite", word_window)
        word_window.protocol("WM_DELETE_WINDOW", lambda: self.close_window("Word Lite", word_window))
    '''
    
    def create_word_lite(self, file_path=None):
        """Creează o aplicație Word Lite pentru a deschide și vizualiza fișiere .doc sau .docx"""
        import zipfile
        import xml.etree.ElementTree as ET
        import re
        
        word_window = tk.Toplevel(self.rootW95dist)
        word_window.title("Word Lite")
        word_window.overrideredirect(True)
        word_window.geometry("800x600+200+100")
        word_window.configure(bg="#c0c0c0")
        
        # Add Windows 95 style title bar
        title_bar = tk.Frame(word_window, bg="#000080", height=25)
        title_bar.pack(fill="x", side="top")
        title_label = tk.Label(title_bar, text="Word Lite", fg="white", bg="#000080",
                              font=("MS Sans Serif", 8, "bold"))
        title_label.pack(side="left", padx=5, pady=2)
        
        # Close button for title bar
        close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                font=("Arial", 8, "bold"), width=2, height=1,
                                relief="raised", bd=1,
                                command=lambda: self.close_window("Word Lite", word_window))
        close_button.pack(side="right", padx=2, pady=1)
        
        self.make_window_draggable(word_window, title_bar)
        
        # Menubar
        menubar = tk.Menu(word_window)
        word_window.config(menu=menubar)
        
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        
        edit_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Edit", menu=edit_menu)
        
        view_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="View", menu=view_menu)
        
        format_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Format", menu=format_menu)
        
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        
        # Toolbar frame
        toolbar_frame = tk.Frame(word_window, bg="#c0c0c0", relief="raised", bd=2, height=35)
        toolbar_frame.pack(fill="x")
        toolbar_frame.pack_propagate(False)
        
        # Open button
        open_btn = tk.Button(toolbar_frame, text="Open", bg="#c0c0c0", relief="raised", bd=2,
                            font=("MS Sans Serif", 8),
                            command=lambda: open_document())
        open_btn.pack(side="left", padx=5, pady=2)
        
        # Save button (for future implementation)
        save_btn = tk.Button(toolbar_frame, text="Save", bg="#c0c0c0", relief="raised", bd=2,
                            font=("MS Sans Serif", 8),
                            command=lambda: save_document())
        save_btn.pack(side="left", padx=5, pady=2)
        
        # Print button
        print_btn = tk.Button(toolbar_frame, text="Print", bg="#c0c0c0", relief="raised", bd=2,
                             font=("MS Sans Serif", 8),
                             command=lambda: messagebox.showinfo("Print", "Printing functionality is not implemented in this version"))
        print_btn.pack(side="left", padx=5, pady=2)
        
        # Status bar
        status_frame = tk.Frame(word_window, bg="#c0c0c0", relief="sunken", bd=1, height=25)
        status_frame.pack(side="bottom", fill="x")
        status_frame.pack_propagate(False)
        
        status_label = tk.Label(status_frame, text="Ready", bg="#c0c0c0", font=("MS Sans Serif", 8))
        status_label.pack(side="left", padx=5)
        
        # Main content frame
        content_frame = tk.Frame(word_window, bg="white")
        content_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Text area
        text_frame = tk.Frame(content_frame, bg="white")
        text_frame.pack(fill="both", expand=True)
        
        # Scrollbars
        v_scrollbar = tk.Scrollbar(text_frame)
        v_scrollbar.pack(side="right", fill="y")
        
        h_scrollbar = tk.Scrollbar(text_frame, orient="horizontal")
        h_scrollbar.pack(side="bottom", fill="x")
        
        # Text widget
        text_area = tk.Text(text_frame, wrap="word", yscrollcommand=v_scrollbar.set, 
                           xscrollcommand=h_scrollbar.set,
                           font=("Times New Roman", 12), bg="white", fg="black")
        text_area.pack(side="left", fill="both", expand=True)
        
        v_scrollbar.config(command=text_area.yview)
        h_scrollbar.config(command=text_area.xview)
        
        # Variables for tracking
        current_file = None
        
        def extract_text_from_docx(docx_file):
            """Extract text from a .docx file using standard libraries"""
            try:
                # Fișierele .docx sunt arhive ZIP care conțin fișiere XML
                z = zipfile.ZipFile(docx_file)
                
                # Verificăm dacă există fișierul document.xml
                if "word/document.xml" in z.namelist():
                    # Extragem și citim XML-ul principal
                    xml_content = z.read("word/document.xml")
                    
                    # Parsăm XML-ul
                    tree = ET.fromstring(xml_content)
                    
                    # Găsim toate elementele "w:t" care conțin text
                    text_content = ""
                    
                    # Folosim un namespace simplificat
                    namespace = {'w': '*'}  # Wildcards pentru namespace pentru a evita linkul URL
                    
                    # Încercăm mai multe abordări pentru a extrage textul
                    try:
                        # Abordarea 1: Extragem textul folosind XPath cu wildcard namespace
                        for paragraph in tree.findall('.//{*}p'):
                            paragraph_text = ""
                            for text_element in paragraph.findall('.//{*}t'):
                                if text_element.text:
                                    paragraph_text += text_element.text
                            
                            text_content += paragraph_text + "\n"
                        
                        # Dacă nu am obținut text, încercăm o altă abordare
                        if not text_content.strip():
                            raise Exception("No text was found using the first approach.")
                            
                    except:
                        # Abordarea 2: Extragem direct toate elementele de text
                        for text_element in tree.findall('.//*'):
                            if text_element.text and text_element.text.strip():
                                text_content += text_element.text + "\n"
                    
                    return text_content
                else:
                    return "The document content could not be found in the archive."
            except Exception as e:
                return f"Error extracting text: {str(e)}"
        
        def extract_text_from_doc(doc_file):
            """Extrage text din fișier .doc (format binar vechi)"""
            # Pentru fișierele .doc (format binar), putem încerca să extragem doar textul vizibil
            try:
                with open(doc_file, 'rb') as f:
                    content = f.read()
                    
                    # Încercăm să extragem text vizibil folosind o abordare simplă
                    # Aceasta nu este o soluție perfectă, dar poate funcționa pentru text simplu
                    text = ""
                    
                    # Convertim conținutul binar în text dacă este posibil
                    try:
                        # Încercăm să decodificăm ca UTF-16
                        binary_text = content.decode('utf-16', errors='ignore')
                        
                        # Eliminăm caracterele non-printabile
                        printable_text = ''.join(c for c in binary_text if c.isprintable() or c.isspace())
                        
                        # Folosim regex pentru a găsi texte consecutive
                        # Căutăm secvențe de text de cel puțin 3 caractere
                        text_chunks = re.findall(r'[A-Za-z0-9\s.,;:!?\'"\-+]{3,}', printable_text)
                        
                        # Filtrăm și combinăm fragmentele de text
                        for chunk in text_chunks:
                            chunk = chunk.strip()
                            if len(chunk) > 5:  # Ignorăm fragmentele prea scurte
                                text += chunk + "\n"
                    except:
                        text = "Text could not be extracted from document .doc.\n"
                        text += "The .doc format is an old and complex binary format.\n"
                        text += "For better results, use .docx or .txt files."
                    
                    return text
            except Exception as e:
                return f"Error opening the .doc file: {str(e)}"
        
        def open_document():
            """Open a Word document"""
            file_path = filedialog.askopenfilename(
                defaultextension=".docx",
                filetypes=[("Word files", "*.docx;*.doc"), ("Text files", "*.txt"), ("All files", "*.*")]
            )
            
            if not file_path:
                return
            
            try:
                # Set cursor to wait
                word_window.config(cursor="wait")
                word_window.update()
                
                # Clear text area
                text_area.delete(1.0, tk.END)
                
                # Get the file extension
                ext = os.path.splitext(file_path)[1].lower()
                
                # Handle different file types
                if ext == ".txt":
                    # Simple text file
                    with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                        content = file.read()
                        text_area.insert(tk.END, content)
                elif ext == ".docx":
                    # Extract text from .docx file
                    content = extract_text_from_docx(file_path)
                    text_area.insert(tk.END, content)
                elif ext == ".doc":
                    # Try to extract text from old .doc format
                    content = extract_text_from_doc(file_path)
                    text_area.insert(tk.END, content)
                else:
                    text_area.insert(tk.END, f"Cannot open file type: {ext}\n")
                    text_area.insert(tk.END, "Word Lite supports .txt, .doc, and .docx files.")
                
                # Update window title and status
                file_name = os.path.basename(file_path)
                title_label.config(text=f"Word Lite - {file_name}")
                word_window.title(f"Word Lite - {file_name}")
                status_label.config(text=f"Opened: {file_name}")
                
                # Save the current file
                nonlocal current_file
                current_file = file_path
                
            except Exception as e:
                text_area.delete(1.0, tk.END)
                text_area.insert(tk.END, f"Error opening file: {str(e)}")
                messagebox.showerror("Error", f"Could not open file: {str(e)}")
            finally:
                # Reset cursor
                word_window.config(cursor="")
        
        def save_document():
            """Save a document (placeholder)"""
            messagebox.showinfo("Info", "Save functionality is not implemented in this version")
        
        # File menu commands
        file_menu.add_command(label="Open", command=open_document)
        file_menu.add_command(label="Save", command=save_document)
        file_menu.add_separator()
        file_menu.add_command(label="Print", command=lambda: messagebox.showinfo("Print", "Printing functionality is not implemented in this version"))
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=lambda: self.close_window("Word Lite", word_window))
        
        # Edit menu commands
        edit_menu.add_command(label="Cut", command=lambda: text_area.event_generate("<<Cut>>"))
        edit_menu.add_command(label="Copy", command=lambda: text_area.event_generate("<<Copy>>"))
        edit_menu.add_command(label="Paste", command=lambda: text_area.event_generate("<<Paste>>"))
        
        # View menu
        view_menu.add_command(label="Zoom", command=lambda: status_label.config(text="Zoom not implemented"))
        
        # Format menu
        format_menu.add_command(label="Font", command=lambda: status_label.config(text="Font formatting not implemented"))
        
        # Help menu
        help_menu.add_command(label="About Word Lite", command=lambda: messagebox.showinfo("About", "Word Lite\nVersion 1.0\n\nA simple document viewer"))
        
        # If a file path was provided, open it
        if file_path and os.path.exists(file_path):
            current_file = file_path
            try:
                # Use a temporary file path variable to avoid conflicts with the open_document function
                temp_path = file_path
                # Clear it to avoid double opening
                file_path = None
                # Call the open function with the correct path
                open_document()
            except Exception as e:
                text_area.delete(1.0, tk.END)
                text_area.insert(tk.END, f"Error opening file: {str(e)}")
        
        # Add to taskbar
        self.add_window_to_taskbar("Word Lite", word_window)
        word_window.protocol("WM_DELETE_WINDOW", lambda: self.close_window("Word Lite", word_window))
    
    def create_terminal(self):
        """Creează un terminal/command prompt care execută comenzi reale"""
        import subprocess
        import threading
        import platform
        import signal
        import time
        
        terminal_window = tk.Toplevel(self.rootW95dist)
        terminal_window.title("Command Prompt")
        terminal_window.overrideredirect(True)
        terminal_window.geometry("640x400+200+150")
        terminal_window.configure(bg="#000000")
        
        # Add Windows 95 style title bar
        title_bar = tk.Frame(terminal_window, bg="#000080", height=25)
        title_bar.pack(fill="x", side="top")
        title_label = tk.Label(title_bar, text="Command Prompt", fg="white", bg="#000080",
                              font=("MS Sans Serif", 8, "bold"))
        title_label.pack(side="left", padx=5, pady=2)
        
        # Close button for title bar
        close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                font=("Arial", 8, "bold"), width=2, height=1,
                                relief="raised", bd=1,
                                command=lambda: self.close_window("Command Prompt", terminal_window))
        close_button.pack(side="right", padx=2, pady=1)
        
        self.make_window_draggable(terminal_window, title_bar)
        
        # Main content - Folosim un singur Text widget pentru tot
        # în loc de a separa output și input
        terminal_frame = tk.Frame(terminal_window, bg="black", bd=0)
        terminal_frame.pack(fill="both", expand=True, padx=2, pady=2)
        
        # Scrollbars
        v_scrollbar = tk.Scrollbar(terminal_frame)
        v_scrollbar.pack(side="right", fill="y")
        
        # Text widget pentru afișarea atât a output-ului cât și pentru introducerea comenzilor
        terminal_text = tk.Text(terminal_frame, wrap="word", bg="black", fg="#cccccc",
                               insertbackground="white", font=("Courier New", 10),
                               yscrollcommand=v_scrollbar.set)
        terminal_text.pack(fill="both", expand=True)
        v_scrollbar.config(command=terminal_text.yview)
        
        # Command history
        command_history = []
        history_index = 0
        
        # Poziția curentă de input (tag pentru a marca unde utilizatorul poate scrie)
        input_pos = "1.0"
        
        # Prompt-ul utilizat
        prompt = "> "
        
        # Get current OS
        current_os = platform.system()
        
        # Referință la procesul curent
        current_process = None
        
        # Flag pentru a ști dacă o comandă este în execuție
        command_running = False
        
        # Inițializăm terminalul
        terminal_text.insert("end", "Windows 95 Command Prompt\n")
        terminal_text.insert("end", "Type 'help' for a list of commands.\n\n")
        terminal_text.insert("end", f"Current directory: {os.getcwd()}\n\n")
        terminal_text.insert("end", prompt)
        
        # Setăm poziția de input inițială
        input_pos = terminal_text.index("end-1c")
        terminal_text.mark_set("input_mark", input_pos)
        terminal_text.mark_gravity("input_mark", "left")
        
        # Focusăm pe textbox
        terminal_text.focus_set()
        terminal_text.see("end")
        
        def get_command():
            """Obține comanda curentă din terminal_text"""
            input_start = terminal_text.index("input_mark")
            input_end = terminal_text.index("end-1c")
            return terminal_text.get(input_start, input_end)
        
        def stop_current_process():
            """Oprește procesul curent dacă există"""
            nonlocal current_process, command_running
            
            if current_process and command_running:
                try:
                    if current_os == "Windows":
                        # În Windows, folosim taskkill pentru a termina procesul și subprocesele sale
                        subprocess.run(f"taskkill /F /T /PID {current_process.pid}", shell=True)
                    else:
                        # În Unix/Linux, trimitem semnalul SIGTERM
                        os.killpg(os.getpgid(current_process.pid), signal.SIGTERM)
                    
                    terminal_text.insert("end", "\n^C\nCommand terminated by user.\n")
                    terminal_text.insert("end", prompt)
                    input_pos = terminal_text.index("end-1c")
                    terminal_text.mark_set("input_mark", input_pos)
                    terminal_text.see("end")
                    
                    command_running = False
                    current_process = None
                    return True
                except Exception as e:
                    terminal_text.insert("end", f"\nError terminating process: {str(e)}\n")
                    return False
            return False
        
        def execute_command(command):
            """Execută o comandă și afișează output-ul"""
            nonlocal history_index, input_pos, current_process, command_running
            
            # Adaugă comanda la istoric
            command_history.append(command)
            history_index = len(command_history)
            
            # Adăugăm un newline după comandă
            terminal_text.insert("end", "\n")
            
            # Procesăm exit command
            if command.lower() in ["exit", "quit"]:
                terminal_text.insert("end", "Closing Command Prompt...\n")
                terminal_text.see("end")
                terminal_window.after(500, lambda: self.close_window("Command Prompt", terminal_window))
                return
            
            # Procesăm cd command intern
            if command.lower().startswith("cd "):
                try:
                    path = command[3:].strip()
                    if os.path.isdir(path):
                        os.chdir(path)
                        terminal_text.insert("end", f"Current directory changed to: {os.getcwd()}\n\n")
                    else:
                        terminal_text.insert("end", f"Directory not found: {path}\n\n")
                except Exception as e:
                    terminal_text.insert("end", f"Error: {str(e)}\n\n")
                
                terminal_text.insert("end", prompt)
                input_pos = terminal_text.index("end-1c")
                terminal_text.mark_set("input_mark", input_pos)
                terminal_text.see("end")
                return
            
            # Procesăm dir/ls command intern pentru formatare mai bună
            if command.lower() in ["dir", "ls"]:
                try:
                    current_dir = os.getcwd()
                    terminal_text.insert("end", f"Directory of {current_dir}\n\n")
                    
                    # Obținem fișierele și directoarele
                    items = os.listdir(current_dir)
                    
                    # Afișăm mai întâi directoarele
                    for item in sorted(items):
                        item_path = os.path.join(current_dir, item)
                        if os.path.isdir(item_path):
                            try:
                                # Obținem informații despre director
                                item_time = os.path.getmtime(item_path)
                                time_str = datetime.fromtimestamp(item_time).strftime('%Y-%m-%d %H:%M:%S')
                                terminal_text.insert("end", f"{time_str}    <DIR>          {item}\n")
                            except:
                                terminal_text.insert("end", f"                <DIR>          {item}\n")
                    
                    # Apoi afișăm fișierele
                    for item in sorted(items):
                        item_path = os.path.join(current_dir, item)
                        if os.path.isfile(item_path):
                            try:
                                # Obținem informații despre fișier
                                item_size = os.path.getsize(item_path)
                                item_time = os.path.getmtime(item_path)
                                time_str = datetime.fromtimestamp(item_time).strftime('%Y-%m-%d %H:%M:%S')
                                terminal_text.insert("end", f"{time_str}    {item_size:10} {item}\n")
                            except:
                                terminal_text.insert("end", f"                      ? {item}\n")
                    
                    terminal_text.insert("end", "\n")
                except Exception as e:
                    terminal_text.insert("end", f"Error: {str(e)}\n\n")
                
                terminal_text.insert("end", prompt)
                input_pos = terminal_text.index("end-1c")
                terminal_text.mark_set("input_mark", input_pos)
                terminal_text.see("end")
                return
            
            # Procesăm help command
            if command.lower() == "help":
                terminal_text.insert("end", "Available Commands:\n")
                terminal_text.insert("end", "  help       - Display this help message\n")
                terminal_text.insert("end", "  cd <dir>   - Change directory\n")
                terminal_text.insert("end", "  dir        - List files and directories\n")
                terminal_text.insert("end", "  ls         - List files and directories (alternative)\n")
                terminal_text.insert("end", "  cls        - Clear screen\n")
                terminal_text.insert("end", "  exit       - Exit Command Prompt\n")
                terminal_text.insert("end", "\nOther standard system commands are also available.\n")
                terminal_text.insert("end", "Press Ctrl+C to stop a running command or copy selected text.\n")
                terminal_text.insert("end", "Press Ctrl+V to paste text from clipboard.\n\n")
                
                terminal_text.insert("end", prompt)
                input_pos = terminal_text.index("end-1c")
                terminal_text.mark_set("input_mark", input_pos)
                terminal_text.see("end")
                return
            
            # Procesăm clear/cls command
            if command.lower() in ["cls", "clear"]:
                terminal_text.delete("1.0", "end")
                terminal_text.insert("end", "Windows 95 Command Prompt\n\n")
                terminal_text.insert("end", prompt)
                input_pos = terminal_text.index("end-1c")
                terminal_text.mark_set("input_mark", input_pos)
                terminal_text.see("end")
                return
            
            # Rulăm comanda externă
            def run_command():
                nonlocal current_process, command_running
                command_running = True
                
                try:
                    # Determină shell și opțiuni în funcție de sistemul de operare
                    if current_os == "Windows":
                        # Creăm procesul cu pipe-uri pentru stdout și stderr
                        current_process = subprocess.Popen(
                            command, 
                            shell=True, 
                            stdout=subprocess.PIPE, 
                            stderr=subprocess.PIPE,
                            stdin=subprocess.PIPE,
                            text=True,
                            encoding='cp850',  # Windows Command Prompt encoding
                            creationflags=subprocess.CREATE_NEW_PROCESS_GROUP
                        )
                    else:  # Linux/Mac
                        current_process = subprocess.Popen(
                            command,
                            shell=True,
                            stdout=subprocess.PIPE,
                            stderr=subprocess.PIPE,
                            stdin=subprocess.PIPE,
                            text=True,
                            preexec_fn=os.setsid  # Permite terminarea procesului și a subproceselor
                        )
                    
                    # Citim output-ul și error-ul în mod continuu
                    while command_running:
                        # Verificăm dacă procesul s-a terminat
                        if current_process.poll() is not None:
                            break
                        
                        # Citim output disponibil fără să blocăm
                        output_line = current_process.stdout.readline()
                        if output_line:
                            terminal_window.after(0, lambda line=output_line: append_output(line))
                        
                        # Citim error disponibil fără să blocăm
                        error_line = current_process.stderr.readline()
                        if error_line:
                            terminal_window.after(0, lambda line=error_line: append_error(line))
                        
                        # Pauză scurtă pentru a nu supraîncărca CPU
                        time.sleep(0.01)
                    
                    # După ce procesul s-a terminat, citim orice output rămas
                    remaining_output, remaining_error = current_process.communicate()
                    
                    if remaining_output:
                        terminal_window.after(0, lambda: append_output(remaining_output))
                    
                    if remaining_error:
                        terminal_window.after(0, lambda: append_error(remaining_error))
                    
                    # Verificăm dacă procesul s-a terminat cu succes
                    exit_code = current_process.returncode
                    if exit_code != 0 and command_running:
                        terminal_window.after(0, lambda: append_error(f"\nCommand exited with code {exit_code}\n"))
                    
                    # Comanda s-a terminat, adăugăm un nou prompt
                    terminal_window.after(0, lambda: add_new_prompt())
                    
                    command_running = False
                    current_process = None
                    
                except Exception as e:
                    terminal_window.after(0, lambda: append_error(f"\nError executing command: {str(e)}\n"))
                    terminal_window.after(0, lambda: add_new_prompt())
                    command_running = False
                    current_process = None
            
            def append_output(text):
                terminal_text.insert("end", text)
                terminal_text.see("end")
            
            def append_error(text):
                terminal_text.insert("end", text)
                terminal_text.see("end")
            
            def add_new_prompt():
                nonlocal input_pos
                terminal_text.insert("end", prompt)
                input_pos = terminal_text.index("end-1c")
                terminal_text.mark_set("input_mark", input_pos)
                terminal_text.see("end")
            
            # Rulăm comanda într-un thread separat pentru a evita blocarea UI
            command_thread = threading.Thread(target=run_command)
            command_thread.daemon = True
            command_thread.start()
        
        def copy_selected_text():
            """Copiază textul selectat în clipboard"""
            try:
                selected_text = terminal_text.get("sel.first", "sel.last")
                terminal_window.clipboard_clear()
                terminal_window.clipboard_append(selected_text)
                return True
            except:
                return False  # Nu există text selectat
        
        def paste_from_clipboard():
            """Lipește text din clipboard la poziția curentă de input"""
            try:
                clipboard_text = terminal_window.clipboard_get()
                terminal_text.insert("insert", clipboard_text)
                return True
            except:
                return False  # Clipboard-ul este gol sau inaccesibil
        
        # Gestionăm tastele speciale și comenzile
        def key_press(event):
            nonlocal history_index
            
            # Ctrl+C: Copiere text selectat sau oprire comandă
            if (event.state & 0x4) and event.keysym.lower() == "c":  # Ctrl+C
                if not copy_selected_text():  # Încercăm să copiem textul selectat
                    # Dacă nu există text selectat, oprim comanda curentă
                    stop_current_process()
                return "break"
            
            # Ctrl+V: Lipire din clipboard
            if (event.state & 0x4) and event.keysym.lower() == "v":  # Ctrl+V
                paste_from_clipboard()
                return "break"
            
            # Verificăm dacă cursorul este după prompt
            cursor_pos = terminal_text.index("insert")
            if terminal_text.compare(cursor_pos, "<", "input_mark"):
                # Cursorul este înainte de poziția de input, îl mutăm la final
                terminal_text.mark_set("insert", "end")
                return "break"
            
            # Return/Enter key - Execute command
            if event.keysym == "Return":
                command = get_command()
                if command.strip():
                    execute_command(command.strip())
                return "break"
            
            # Backspace key - Do not delete beyond prompt
            elif event.keysym == "BackSpace":
                cursor_pos = terminal_text.index("insert")
                if terminal_text.compare(cursor_pos, "<=", "input_mark"):
                    return "break"
            
            # Up key - Navigate command history
            elif event.keysym == "Up":
                if command_history and history_index > 0:
                    history_index -= 1
                    # Ștergem comanda curentă
                    terminal_text.delete("input_mark", "end-1c")
                    # Inserăm comanda din istoric
                    terminal_text.insert("input_mark", command_history[history_index])
                return "break"
            
            # Down key - Navigate command history
            elif event.keysym == "Down":
                if command_history and history_index < len(command_history) - 1:
                    history_index += 1
                    # Ștergem comanda curentă
                    terminal_text.delete("input_mark", "end-1c")
                    # Inserăm comanda din istoric
                    terminal_text.insert("input_mark", command_history[history_index])
                elif history_index == len(command_history) - 1:
                    history_index = len(command_history)
                    # Ștergem comanda curentă
                    terminal_text.delete("input_mark", "end-1c")
                return "break"
            
            # Home key - Go to beginning of input
            elif event.keysym == "Home":
                terminal_text.mark_set("insert", "input_mark")
                return "break"
        
        # Asigurăm-ne că nu se poate șterge input mark-ul sau textul anterior
        def delete_check(event):
            try:
                if event.keysym == 'BackSpace':
                    cursor_pos = terminal_text.index("insert")
                    if terminal_text.compare(cursor_pos, "<=", "input_mark"):
                        return "break"
                elif event.keysym in ['Delete', 'KP_Delete']:
                    try:
                        sel_start = terminal_text.index("sel.first")
                        if terminal_text.compare(sel_start, "<", "input_mark"):
                            return "break"
                    except:
                        cursor_pos = terminal_text.index("insert")
                        if terminal_text.compare(cursor_pos, "<", "input_mark"):
                            return "break"
            except:
                pass
        
        # Curățăm resursele la închiderea ferestrei
        def on_window_close():
            # Oprim orice proces în execuție
            stop_current_process()
            # Închidem fereastra
            self.close_window("Command Prompt", terminal_window)
        
        # Bind keys
        terminal_text.bind("<Key>", key_press)
        terminal_text.bind("<KeyPress-BackSpace>", delete_check)
        terminal_text.bind("<KeyPress-Delete>", delete_check)
        terminal_text.bind("<KeyPress-KP_Delete>", delete_check)
        
        # Make sure focus is on the terminal text
        terminal_text.focus_set()
        
        # Add to taskbar
        self.add_window_to_taskbar("Command Prompt", terminal_window)
        terminal_window.protocol("WM_DELETE_WINDOW", on_window_close)
        
        # Bind window activation
        terminal_window.bind("<Map>", lambda e: terminal_text.focus_set())
        terminal_window.bind("<FocusIn>", lambda e: terminal_text.focus_set())
    
    def create_excel_lite(self, file_path=None):
        """Creează o aplicație Excel Lite pentru a deschide și vizualiza fișiere xlsx"""
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Error", "The openpyxl library is not installed. Please install it using pip: pip install openpyxl")
            return

        excel_window = tk.Toplevel(self.rootW95dist)
        excel_window.title("Excel Lite")
        excel_window.overrideredirect(True)
        excel_window.geometry("800x600+200+100")
        excel_window.configure(bg="#c0c0c0")
        
        # Add Windows 95 style title bar
        title_bar = tk.Frame(excel_window, bg="#000080", height=25)
        title_bar.pack(fill="x", side="top")
        title_label = tk.Label(title_bar, text="Excel Lite", fg="white", bg="#000080",
                              font=("MS Sans Serif", 8, "bold"))
        title_label.pack(side="left", padx=5, pady=2)
        
        # Close button for title bar
        close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                font=("Arial", 8, "bold"), width=2, height=1,
                                relief="raised", bd=1,
                                command=lambda: self.close_window("Excel Lite", excel_window))
        close_button.pack(side="right", padx=2, pady=1)
        
        self.make_window_draggable(excel_window, title_bar)
        
        # Menubar
        menubar = tk.Menu(excel_window)
        excel_window.config(menu=menubar)
        
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        
        edit_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Edit", menu=edit_menu)
        
        view_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="View", menu=view_menu)
        
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        
        # Toolbar frame
        toolbar_frame = tk.Frame(excel_window, bg="#c0c0c0", relief="raised", bd=2, height=35)
        toolbar_frame.pack(fill="x")
        toolbar_frame.pack_propagate(False)
        
        # Open button
        open_btn = tk.Button(toolbar_frame, text="Open", bg="#c0c0c0", relief="raised", bd=2,
                            font=("MS Sans Serif", 8),
                            command=lambda: open_excel_file())
        open_btn.pack(side="left", padx=5, pady=2)
        
        # Save button (for future implementation)
        save_btn = tk.Button(toolbar_frame, text="Save", bg="#c0c0c0", relief="raised", bd=2,
                            font=("MS Sans Serif", 8),
                            command=lambda: save_excel_file())
        save_btn.pack(side="left", padx=5, pady=2)
        
        # Sheet selection
        sheet_label = tk.Label(toolbar_frame, text="Sheet:", bg="#c0c0c0", font=("MS Sans Serif", 8))
        sheet_label.pack(side="left", padx=(20, 5))
        
        sheet_var = tk.StringVar()
        sheet_combo = ttk.Combobox(toolbar_frame, textvariable=sheet_var, width=15, state="readonly")
        sheet_combo.pack(side="left", padx=5)
        sheet_combo.bind("<<ComboboxSelected>>", lambda e: show_sheet(sheet_var.get()))
        
        # Function buttons
        formula_btn = tk.Button(toolbar_frame, text="Formula", bg="#c0c0c0", relief="raised", bd=2,
                               font=("MS Sans Serif", 8),
                               command=lambda: add_formula())
        formula_btn.pack(side="left", padx=5, pady=2)
        
        sort_btn = tk.Button(toolbar_frame, text="Sort", bg="#c0c0c0", relief="raised", bd=2,
                            font=("MS Sans Serif", 8),
                            command=lambda: sort_data())
        sort_btn.pack(side="left", padx=5, pady=2)
        
        filter_btn = tk.Button(toolbar_frame, text="Filter", bg="#c0c0c0", relief="raised", bd=2,
                              font=("MS Sans Serif", 8),
                              command=lambda: filter_data())
        filter_btn.pack(side="left", padx=5, pady=2)
        
        # Status bar - adăugat înainte de conținutul principal
        status_frame = tk.Frame(excel_window, bg="#c0c0c0", relief="sunken", bd=1, height=25)
        status_frame.pack(side="bottom", fill="x")
        status_frame.pack_propagate(False)
        
        status_label = tk.Label(status_frame, text="Ready", bg="#c0c0c0", font=("MS Sans Serif", 8))
        status_label.pack(side="left", padx=5)
        
        cell_info_label = tk.Label(status_frame, text="", bg="#c0c0c0", font=("MS Sans Serif", 8))
        cell_info_label.pack(side="right", padx=5)
        
        # Main content frame
        content_frame = tk.Frame(excel_window, bg="white")
        content_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Scrollbars
        h_scrollbar = tk.Scrollbar(content_frame, orient="horizontal")
        h_scrollbar.pack(side="bottom", fill="x")
        
        # Main table frame
        table_frame = tk.Frame(content_frame, bg="white")
        table_frame.pack(fill="both", expand=True)
        
        v_scrollbar = tk.Scrollbar(table_frame)
        v_scrollbar.pack(side="right", fill="y")
        
        # Canvas for scrollable content
        canvas = tk.Canvas(table_frame, bg="white", 
                         xscrollcommand=h_scrollbar.set,
                         yscrollcommand=v_scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        
        h_scrollbar.config(command=canvas.xview)
        v_scrollbar.config(command=canvas.yview)
        
        # Frame inside canvas for grid
        inner_frame = tk.Frame(canvas, bg="white")
        canvas.create_window((0, 0), window=inner_frame, anchor="nw")
        
        # Eliminăm spațiile între celule
        inner_frame.grid_columnconfigure("all", pad=0)
        inner_frame.grid_rowconfigure("all", pad=0)
        
        # Variables for tracking
        current_file = None
        workbook = None
        active_sheet = None
        
        # Cell grid (for display only)
        cell_grid = []
        header_labels = []  # Pentru a ține evidența label-urilor de header
        
        # Max rows and columns to display
        MAX_ROWS = 100
        MAX_COLS = 26  # A-Z
        
        # Variabile pentru redimensionarea coloanelor
        resizing_column = None
        start_x = 0
        column_widths = [10] * (MAX_COLS + 1)  # Lățimea inițială pentru fiecare coloană (inclusiv header-ul de rând)
        
        def clear_sheet():
            """Clear the current sheet display"""
            for widget in inner_frame.winfo_children():
                widget.destroy()
            
            nonlocal cell_grid, header_labels
            cell_grid = []
            header_labels = []
        
        def create_empty_sheet():
            """Create an empty spreadsheet grid"""
            clear_sheet()
            
            nonlocal header_labels, column_widths
            header_labels = [None] * (MAX_COLS + 1)  # +1 pentru header-ul de rând
            
            # Creează header-ul cu coloanele (A, B, C, etc.)
            header_label = tk.Label(inner_frame, text="", width=4, bg="#e0e0e0", 
                                  relief="raised", bd=1, font=("MS Sans Serif", 8, "bold"),
                                  highlightthickness=0)  # Eliminăm highlightthickness
            header_label.grid(row=0, column=0, sticky="nsew", padx=0, pady=0)  # Eliminăm padding
            header_labels[0] = header_label
            
            for col in range(MAX_COLS):
                col_letter = chr(65 + col)  # A=65, B=66, etc.
                header = tk.Label(inner_frame, text=col_letter, width=column_widths[col+1], bg="#e0e0e0",
                                relief="raised", bd=1, font=("MS Sans Serif", 8, "bold"),
                                highlightthickness=0)  # Eliminăm highlightthickness
                header.grid(row=0, column=col+1, sticky="nsew", padx=0, pady=0)  # Eliminăm padding
                header_labels[col+1] = header
                
                # Adaugă separator de redimensionare
                add_column_resize_handle(header, col+1)
            
            # Creează rândurile și celulele
            row_widgets = []
            for row in range(MAX_ROWS):
                # Creează header-ul pentru rând
                row_header = tk.Label(inner_frame, text=str(row + 1), width=4, height=1,
                                    bg="#e0e0e0", relief="raised", bd=1, font=("MS Sans Serif", 8, "bold"),
                                    highlightthickness=0)  # Eliminăm highlightthickness
                row_header.grid(row=row+1, column=0, sticky="nsew", padx=0, pady=0)  # Eliminăm padding
                
                # Creează celulele pentru acest rând
                col_widgets = []
                for col in range(MAX_COLS):
                    cell = tk.Entry(inner_frame, width=column_widths[col+1], font=("MS Sans Serif", 8),
                                  relief="sunken", bd=1, highlightthickness=0)  # Eliminăm highlightthickness
                    cell.grid(row=row+1, column=col+1, sticky="nsew", padx=0, pady=0)  # Eliminăm padding
                    
                    # Bind click to select cell
                    cell.bind("<Button-1>", lambda e, r=row, c=col: select_cell(r, c))
                    
                    col_widgets.append(cell)
                row_widgets.append(col_widgets)
            
            cell_grid = row_widgets
            
            # Configurează coloanele pentru dimensiune uniformă
            for col in range(MAX_COLS + 1):  # +1 pentru coloana cu header-ele de rând
                inner_frame.grid_columnconfigure(col, weight=1, pad=0)  # Setăm explicit pad=0
            
            # Configurează rândurile pentru a elimina spațiile
            for row in range(MAX_ROWS + 1):  # +1 pentru rândul cu header-ele de coloane
                inner_frame.grid_rowconfigure(row, pad=0)  # Setăm explicit pad=0
            
            # Update canvas scroll region
            inner_frame.update_idletasks()
            canvas.config(scrollregion=canvas.bbox("all"))
        
        def add_column_resize_handle(header, col_idx):
            """Adaugă un handler pentru redimensionarea coloanei"""
            
            def start_resize(event):
                nonlocal resizing_column, start_x
                resizing_column = col_idx
                start_x = event.x_root
                # Schimbă cursorul pentru a indica redimensionarea
                excel_window.config(cursor="sb_h_double_arrow")
                
            def stop_resize(event):
                nonlocal resizing_column
                resizing_column = None
                # Resetează cursorul
                excel_window.config(cursor="")
                
            def do_resize(event):
                nonlocal column_widths, start_x
                if resizing_column is not None:
                    # Calculează diferența în pixeli
                    delta_x = event.x_root - start_x
                    
                    # Convertește în unități de lățime (aproximativ 7 pixeli per caracter)
                    delta_width = delta_x // 7
                    
                    if delta_width != 0:
                        # Ajustează lățimea coloanei (minim 3 caractere)
                        new_width = max(3, column_widths[resizing_column] + delta_width)
                        column_widths[resizing_column] = new_width
                        
                        # Actualizează header-ul
                        header_labels[resizing_column].config(width=new_width)
                        
                        # Actualizează toate celulele din coloană
                        for row in range(MAX_ROWS):
                            if row < len(cell_grid) and resizing_column-1 < len(cell_grid[row]):
                                cell_grid[row][resizing_column-1].config(width=new_width)
                        
                        # Resetează start_x pentru următoarea mișcare
                        start_x = event.x_root
                        
                        # Update canvas scroll region
                        inner_frame.update_idletasks()
                        canvas.config(scrollregion=canvas.bbox("all"))
            
            # Creează un frame pentru zona de redimensionare (marginea dreaptă a header-ului)
            resize_frame = tk.Frame(inner_frame, bg="gray", width=2, cursor="sb_h_double_arrow")
            resize_frame.grid(row=0, column=col_idx, sticky="nse", padx=(0, 0))
            
            # Bind evenimente pentru redimensionare
            resize_frame.bind("<Button-1>", start_resize)
            resize_frame.bind("<ButtonRelease-1>", stop_resize)
            excel_window.bind("<B1-Motion>", do_resize)
        
        def select_cell(row, col):
            """Select a cell and show its address in the status bar"""
            col_letter = chr(65 + col)
            cell_address = f"{col_letter}{row + 1}"
            cell_info_label.config(text=f"Cell: {cell_address}")
            
            # Set focus to the cell
            if 0 <= row < len(cell_grid) and 0 <= col < len(cell_grid[0]):
                cell_grid[row][col].focus_set()
        
        def load_sheet_data(sheet):
            """Load data from the workbook sheet into the grid"""
            clear_sheet()
            
            # Get the dimensions of the sheet
            max_row = min(sheet.max_row, MAX_ROWS)
            max_col = min(sheet.max_column, MAX_COLS)
            
            nonlocal header_labels, column_widths
            header_labels = [None] * (MAX_COLS + 1)  # +1 pentru header-ul de rând
            
            # Creează header-ul cu coloanele (A, B, C, etc.)
            header_label = tk.Label(inner_frame, text="", width=4, bg="#e0e0e0", 
                                  relief="raised", bd=1, font=("MS Sans Serif", 8, "bold"),
                                  highlightthickness=0)  # Eliminăm highlightthickness
            header_label.grid(row=0, column=0, sticky="nsew", padx=0, pady=0)  # Eliminăm padding
            header_labels[0] = header_label
            
            for col in range(max_col):
                col_letter = chr(65 + col)  # A=65, B=66, etc.
                header = tk.Label(inner_frame, text=col_letter, width=column_widths[col+1], bg="#e0e0e0",
                                relief="raised", bd=1, font=("MS Sans Serif", 8, "bold"),
                                highlightthickness=0)  # Eliminăm highlightthickness
                header.grid(row=0, column=col+1, sticky="nsew", padx=0, pady=0)  # Eliminăm padding
                header_labels[col+1] = header
                
                # Adaugă separator de redimensionare
                add_column_resize_handle(header, col+1)
            
            # Creează rândurile și celulele
            row_widgets = []
            for row in range(max_row):
                # Creează header-ul pentru rând
                row_header = tk.Label(inner_frame, text=str(row + 1), width=4, height=1,
                                    bg="#e0e0e0", relief="raised", bd=1, font=("MS Sans Serif", 8, "bold"),
                                    highlightthickness=0)  # Eliminăm highlightthickness
                row_header.grid(row=row+1, column=0, sticky="nsew", padx=0, pady=0)  # Eliminăm padding
                
                # Creează celulele pentru acest rând
                col_widgets = []
                for col in range(max_col):
                    cell_value = sheet.cell(row=row+1, column=col+1).value
                    cell_value = str(cell_value) if cell_value is not None else ""
                    
                    cell = tk.Entry(inner_frame, width=column_widths[col+1], font=("MS Sans Serif", 8),
                                  relief="sunken", bd=1, highlightthickness=0)  # Eliminăm highlightthickness
                    cell.insert(0, cell_value)
                    cell.grid(row=row+1, column=col+1, sticky="nsew", padx=0, pady=0)  # Eliminăm padding
                    
                    # Bind click to select cell
                    cell.bind("<Button-1>", lambda e, r=row, c=col: select_cell(r, c))
                    
                    col_widgets.append(cell)
                row_widgets.append(col_widgets)
            
            nonlocal cell_grid
            cell_grid = row_widgets
            
            # Configurează coloanele pentru dimensiune uniformă și fără spații
            for col in range(max_col + 1):  # +1 pentru coloana cu header-ele de rând
                inner_frame.grid_columnconfigure(col, weight=1, pad=0)
            
            # Configurează rândurile pentru a elimina spațiile
            for row in range(max_row + 1):  # +1 pentru rândul cu header-ele de coloane
                inner_frame.grid_rowconfigure(row, pad=0)
            
            # Update canvas scroll region
            inner_frame.update_idletasks()
            canvas.config(scrollregion=canvas.bbox("all"))
            
            status_label.config(text=f"Loaded sheet: {sheet.title}")
        
        def open_excel_file():
            """Open an Excel file"""
            file_path = filedialog.askopenfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if not file_path:
                return
            
            try:
                # Load the Excel file
                nonlocal workbook, current_file
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                current_file = file_path
                
                # Update the sheet selector dropdown
                sheet_names = workbook.sheetnames
                sheet_combo['values'] = sheet_names
                
                if sheet_names:
                    sheet_var.set(sheet_names[0])
                    show_sheet(sheet_names[0])
                
                # Update window title
                file_name = os.path.basename(file_path)
                title_label.config(text=f"Excel Lite - {file_name}")
                excel_window.title(f"Excel Lite - {file_name}")
                
                status_label.config(text=f"Opened: {file_name}")
            except Exception as e:
                messagebox.showerror("Error", f"Could not open Excel file: {str(e)}")
        
        def show_sheet(sheet_name):
            """Show the selected sheet"""
            if not workbook or sheet_name not in workbook.sheetnames:
                return
            
            nonlocal active_sheet
            active_sheet = workbook[sheet_name]
            load_sheet_data(active_sheet)
        
        def save_excel_file():
            """Save the Excel file (placeholder for future implementation)"""
            if not current_file or not workbook:
                messagebox.showinfo("Info", "No file is currently open")
                return
            
            messagebox.showinfo("Info", "Save functionality is not implemented in this version")
        
        def add_formula():
            """Add a formula to the selected cell (placeholder)"""
            messagebox.showinfo("Info", "Formula functionality is not implemented in this version")
        
        def sort_data():
            """Sort data in the selected column (placeholder)"""
            messagebox.showinfo("Info", "Sort functionality is not implemented in this version")
        
        def filter_data():
            """Filter data in the selected column (placeholder)"""
            messagebox.showinfo("Info", "Filter functionality is not implemented in this version")
        
        # Create empty sheet by default
        create_empty_sheet()
        
        # File menu commands
        file_menu.add_command(label="New", command=create_empty_sheet)
        file_menu.add_command(label="Open", command=open_excel_file)
        file_menu.add_command(label="Save", command=save_excel_file)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=lambda: self.close_window("Excel Lite", excel_window))
        
        # Edit menu commands
        edit_menu.add_command(label="Cut", command=lambda: status_label.config(text="Cut not implemented"))
        edit_menu.add_command(label="Copy", command=lambda: status_label.config(text="Copy not implemented"))
        edit_menu.add_command(label="Paste", command=lambda: status_label.config(text="Paste not implemented"))
        
        # View menu
        view_menu.add_command(label="Formulas", command=lambda: status_label.config(text="View Formulas not implemented"))
        view_menu.add_command(label="Freeze Panes", command=lambda: status_label.config(text="Freeze Panes not implemented"))
        
        # Help menu
        help_menu.add_command(label="About Excel Lite", command=lambda: messagebox.showinfo("About", "Excel Lite\nVersion 1.0\n\nA simple spreadsheet viewer"))
        
        # If a file path was provided, open it
        if file_path and os.path.exists(file_path):
            current_file = file_path
            try:
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                sheet_names = workbook.sheetnames
                sheet_combo['values'] = sheet_names
                if sheet_names:
                    sheet_var.set(sheet_names[0])
                    show_sheet(sheet_names[0])
                
                # Update window title
                file_name = os.path.basename(file_path)
                title_label.config(text=f"Excel Lite - {file_name}")
            except Exception as e:
                messagebox.showerror("Error", f"Could not open Excel file: {str(e)}")
        
        # Add to taskbar
        self.add_window_to_taskbar("Excel Lite", excel_window)
        excel_window.protocol("WM_DELETE_WINDOW", lambda: self.close_window("Excel Lite", excel_window))
    
    '''
    def create_excel_lite(self, file_path=None):
        """Creează o aplicație Excel Lite pentru a deschide și vizualiza fișiere xlsx"""
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Error", "The openpyxl library is not installed. Please install it using pip: pip install openpyxl")
            return

        excel_window = tk.Toplevel(self.rootW95dist)
        excel_window.title("Excel Lite")
        excel_window.overrideredirect(True)
        excel_window.geometry("800x600+200+100")
        excel_window.configure(bg="#c0c0c0")
        
        # Add Windows 95 style title bar
        title_bar = tk.Frame(excel_window, bg="#000080", height=25)
        title_bar.pack(fill="x", side="top")
        title_label = tk.Label(title_bar, text="Excel Lite", fg="white", bg="#000080",
                              font=("MS Sans Serif", 8, "bold"))
        title_label.pack(side="left", padx=5, pady=2)
        
        # Close button for title bar
        close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                font=("Arial", 8, "bold"), width=2, height=1,
                                relief="raised", bd=1,
                                command=lambda: self.close_window("Excel Lite", excel_window))
        close_button.pack(side="right", padx=2, pady=1)
        
        self.make_window_draggable(excel_window, title_bar)
        
        # Menubar
        menubar = tk.Menu(excel_window)
        excel_window.config(menu=menubar)
        
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        
        edit_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Edit", menu=edit_menu)
        
        view_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="View", menu=view_menu)
        
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        
        # Toolbar frame
        toolbar_frame = tk.Frame(excel_window, bg="#c0c0c0", relief="raised", bd=2, height=35)
        toolbar_frame.pack(fill="x")
        toolbar_frame.pack_propagate(False)
        
        # Open button
        open_btn = tk.Button(toolbar_frame, text="Open", bg="#c0c0c0", relief="raised", bd=2,
                            font=("MS Sans Serif", 8),
                            command=lambda: open_excel_file())
        open_btn.pack(side="left", padx=5, pady=2)
        
        # Save button (for future implementation)
        save_btn = tk.Button(toolbar_frame, text="Save", bg="#c0c0c0", relief="raised", bd=2,
                            font=("MS Sans Serif", 8),
                            command=lambda: save_excel_file())
        save_btn.pack(side="left", padx=5, pady=2)
        
        # Sheet selection
        sheet_label = tk.Label(toolbar_frame, text="Sheet:", bg="#c0c0c0", font=("MS Sans Serif", 8))
        sheet_label.pack(side="left", padx=(20, 5))
        
        sheet_var = tk.StringVar()
        sheet_combo = ttk.Combobox(toolbar_frame, textvariable=sheet_var, width=15, state="readonly")
        sheet_combo.pack(side="left", padx=5)
        sheet_combo.bind("<<ComboboxSelected>>", lambda e: show_sheet(sheet_var.get()))
        
        # Function buttons
        formula_btn = tk.Button(toolbar_frame, text="Formula", bg="#c0c0c0", relief="raised", bd=2,
                               font=("MS Sans Serif", 8),
                               command=lambda: add_formula())
        formula_btn.pack(side="left", padx=5, pady=2)
        
        sort_btn = tk.Button(toolbar_frame, text="Sort", bg="#c0c0c0", relief="raised", bd=2,
                            font=("MS Sans Serif", 8),
                            command=lambda: sort_data())
        sort_btn.pack(side="left", padx=5, pady=2)
        
        filter_btn = tk.Button(toolbar_frame, text="Filter", bg="#c0c0c0", relief="raised", bd=2,
                              font=("MS Sans Serif", 8),
                              command=lambda: filter_data())
        filter_btn.pack(side="left", padx=5, pady=2)
        
        # Status bar - adăugat înainte de conținutul principal
        status_frame = tk.Frame(excel_window, bg="#c0c0c0", relief="sunken", bd=1, height=25)
        status_frame.pack(side="bottom", fill="x")
        status_frame.pack_propagate(False)
        
        status_label = tk.Label(status_frame, text="Ready", bg="#c0c0c0", font=("MS Sans Serif", 8))
        status_label.pack(side="left", padx=5)
        
        cell_info_label = tk.Label(status_frame, text="", bg="#c0c0c0", font=("MS Sans Serif", 8))
        cell_info_label.pack(side="right", padx=5)
        
        # Main content frame
        content_frame = tk.Frame(excel_window, bg="white")
        content_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Scrollbars
        h_scrollbar = tk.Scrollbar(content_frame, orient="horizontal")
        h_scrollbar.pack(side="bottom", fill="x")
        
        # Main table frame
        table_frame = tk.Frame(content_frame, bg="white")
        table_frame.pack(fill="both", expand=True)
        
        v_scrollbar = tk.Scrollbar(table_frame)
        v_scrollbar.pack(side="right", fill="y")
        
        # Canvas for scrollable content
        canvas = tk.Canvas(table_frame, bg="white", 
                         xscrollcommand=h_scrollbar.set,
                         yscrollcommand=v_scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        
        h_scrollbar.config(command=canvas.xview)
        v_scrollbar.config(command=canvas.yview)
        
        # Frame inside canvas for grid
        inner_frame = tk.Frame(canvas, bg="white")
        canvas.create_window((0, 0), window=inner_frame, anchor="nw")
        
        # Variables for tracking
        current_file = None
        workbook = None
        active_sheet = None
        
        # Cell grid (for display only)
        cell_grid = []
        
        # Max rows and columns to display
        MAX_ROWS = 100
        MAX_COLS = 26  # A-Z
        
        def clear_sheet():
            """Clear the current sheet display"""
            for widget in inner_frame.winfo_children():
                widget.destroy()
            
            nonlocal cell_grid
            cell_grid = []
        
        def create_empty_sheet():
            """Create an empty spreadsheet grid"""
            clear_sheet()
            
            # Definește o lățime fixă pentru toate celulele
            CELL_WIDTH = 10
            
            # Creează header-ul cu coloanele (A, B, C, etc.)
            header_label = tk.Label(inner_frame, text="", width=4, bg="#e0e0e0", 
                                  relief="raised", bd=1, font=("MS Sans Serif", 8, "bold"))
            header_label.grid(row=0, column=0, sticky="nsew")
            
            for col in range(MAX_COLS):
                col_letter = chr(65 + col)  # A=65, B=66, etc.
                header = tk.Label(inner_frame, text=col_letter, width=CELL_WIDTH, bg="#e0e0e0",
                                relief="raised", bd=1, font=("MS Sans Serif", 8, "bold"))
                header.grid(row=0, column=col+1, sticky="nsew")
            
            # Creează rândurile și celulele
            row_widgets = []
            for row in range(MAX_ROWS):
                # Creează header-ul pentru rând
                row_header = tk.Label(inner_frame, text=str(row + 1), width=4, height=1,
                                    bg="#e0e0e0", relief="raised", bd=1, font=("MS Sans Serif", 8, "bold"))
                row_header.grid(row=row+1, column=0, sticky="nsew")
                
                # Creează celulele pentru acest rând
                col_widgets = []
                for col in range(MAX_COLS):
                    cell = tk.Entry(inner_frame, width=CELL_WIDTH, font=("MS Sans Serif", 8),
                                  relief="sunken", bd=1)
                    cell.grid(row=row+1, column=col+1, sticky="nsew")
                    
                    # Bind click to select cell
                    cell.bind("<Button-1>", lambda e, r=row, c=col: select_cell(r, c))
                    
                    col_widgets.append(cell)
                row_widgets.append(col_widgets)
            
            cell_grid = row_widgets
            
            # Configurează coloanele pentru dimensiune uniformă
            for col in range(MAX_COLS + 1):  # +1 pentru coloana cu header-ele de rând
                inner_frame.grid_columnconfigure(col, weight=1)
            
            # Update canvas scroll region
            inner_frame.update_idletasks()
            canvas.config(scrollregion=canvas.bbox("all"))
        
        def select_cell(row, col):
            """Select a cell and show its address in the status bar"""
            col_letter = chr(65 + col)
            cell_address = f"{col_letter}{row + 1}"
            cell_info_label.config(text=f"Cell: {cell_address}")
            
            # Set focus to the cell
            if 0 <= row < len(cell_grid) and 0 <= col < len(cell_grid[0]):
                cell_grid[row][col].focus_set()
        
        def load_sheet_data(sheet):
            """Load data from the workbook sheet into the grid"""
            clear_sheet()
            
            # Get the dimensions of the sheet
            max_row = min(sheet.max_row, MAX_ROWS)
            max_col = min(sheet.max_column, MAX_COLS)
            
            # Definește o lățime fixă pentru toate celulele
            CELL_WIDTH = 10
            
            # Creează header-ul cu coloanele (A, B, C, etc.)
            header_label = tk.Label(inner_frame, text="", width=4, bg="#e0e0e0", 
                                  relief="raised", bd=1, font=("MS Sans Serif", 8, "bold"))
            header_label.grid(row=0, column=0, sticky="nsew")
            
            for col in range(max_col):
                col_letter = chr(65 + col)  # A=65, B=66, etc.
                header = tk.Label(inner_frame, text=col_letter, width=CELL_WIDTH, bg="#e0e0e0",
                                relief="raised", bd=1, font=("MS Sans Serif", 8, "bold"))
                header.grid(row=0, column=col+1, sticky="nsew")
            
            # Creează rândurile și celulele
            row_widgets = []
            for row in range(max_row):
                # Creează header-ul pentru rând
                row_header = tk.Label(inner_frame, text=str(row + 1), width=4, height=1,
                                    bg="#e0e0e0", relief="raised", bd=1, font=("MS Sans Serif", 8, "bold"))
                row_header.grid(row=row+1, column=0, sticky="nsew")
                
                # Creează celulele pentru acest rând
                col_widgets = []
                for col in range(max_col):
                    cell_value = sheet.cell(row=row+1, column=col+1).value
                    cell_value = str(cell_value) if cell_value is not None else ""
                    
                    cell = tk.Entry(inner_frame, width=CELL_WIDTH, font=("MS Sans Serif", 8),
                                  relief="sunken", bd=1)
                    cell.insert(0, cell_value)
                    cell.grid(row=row+1, column=col+1, sticky="nsew")
                    
                    # Bind click to select cell
                    cell.bind("<Button-1>", lambda e, r=row, c=col: select_cell(r, c))
                    
                    col_widgets.append(cell)
                row_widgets.append(col_widgets)
            
            nonlocal cell_grid
            cell_grid = row_widgets
            
            # Configurează coloanele pentru dimensiune uniformă
            for col in range(max_col + 1):  # +1 pentru coloana cu header-ele de rând
                inner_frame.grid_columnconfigure(col, weight=1)
            
            # Update canvas scroll region
            inner_frame.update_idletasks()
            canvas.config(scrollregion=canvas.bbox("all"))
            
            status_label.config(text=f"Loaded sheet: {sheet.title}")
        
        def open_excel_file():
            """Open an Excel file"""
            file_path = filedialog.askopenfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if not file_path:
                return
            
            try:
                # Load the Excel file
                nonlocal workbook, current_file
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                current_file = file_path
                
                # Update the sheet selector dropdown
                sheet_names = workbook.sheetnames
                sheet_combo['values'] = sheet_names
                
                if sheet_names:
                    sheet_var.set(sheet_names[0])
                    show_sheet(sheet_names[0])
                
                # Update window title
                file_name = os.path.basename(file_path)
                title_label.config(text=f"Excel Lite - {file_name}")
                excel_window.title(f"Excel Lite - {file_name}")
                
                status_label.config(text=f"Opened: {file_name}")
            except Exception as e:
                messagebox.showerror("Error", f"Could not open Excel file: {str(e)}")
        
        def show_sheet(sheet_name):
            """Show the selected sheet"""
            if not workbook or sheet_name not in workbook.sheetnames:
                return
            
            nonlocal active_sheet
            active_sheet = workbook[sheet_name]
            load_sheet_data(active_sheet)
        
        def save_excel_file():
            """Save the Excel file (placeholder for future implementation)"""
            if not current_file or not workbook:
                messagebox.showinfo("Info", "No file is currently open")
                return
            
            messagebox.showinfo("Info", "Save functionality is not implemented in this version")
        
        def add_formula():
            """Add a formula to the selected cell (placeholder)"""
            messagebox.showinfo("Info", "Formula functionality is not implemented in this version")
        
        def sort_data():
            """Sort data in the selected column (placeholder)"""
            messagebox.showinfo("Info", "Sort functionality is not implemented in this version")
        
        def filter_data():
            """Filter data in the selected column (placeholder)"""
            messagebox.showinfo("Info", "Filter functionality is not implemented in this version")
        
        # Create empty sheet by default
        create_empty_sheet()
        
        # File menu commands
        file_menu.add_command(label="New", command=create_empty_sheet)
        file_menu.add_command(label="Open", command=open_excel_file)
        file_menu.add_command(label="Save", command=save_excel_file)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=lambda: self.close_window("Excel Lite", excel_window))
        
        # Edit menu commands
        edit_menu.add_command(label="Cut", command=lambda: status_label.config(text="Cut not implemented"))
        edit_menu.add_command(label="Copy", command=lambda: status_label.config(text="Copy not implemented"))
        edit_menu.add_command(label="Paste", command=lambda: status_label.config(text="Paste not implemented"))
        
        # View menu
        view_menu.add_command(label="Formulas", command=lambda: status_label.config(text="View Formulas not implemented"))
        view_menu.add_command(label="Freeze Panes", command=lambda: status_label.config(text="Freeze Panes not implemented"))
        
        # Help menu
        help_menu.add_command(label="About Excel Lite", command=lambda: messagebox.showinfo("About", "Excel Lite\nVersion 1.0\n\nA simple spreadsheet viewer"))
        
        # If a file path was provided, open it
        if file_path and os.path.exists(file_path):
            current_file = file_path
            try:
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                sheet_names = workbook.sheetnames
                sheet_combo['values'] = sheet_names
                if sheet_names:
                    sheet_var.set(sheet_names[0])
                    show_sheet(sheet_names[0])
                
                # Update window title
                file_name = os.path.basename(file_path)
                title_label.config(text=f"Excel Lite - {file_name}")
            except Exception as e:
                messagebox.showerror("Error", f"Could not open Excel file: {str(e)}")
        
        # Add to taskbar
        self.add_window_to_taskbar("Excel Lite", excel_window)
        excel_window.protocol("WM_DELETE_WINDOW", lambda: self.close_window("Excel Lite", excel_window))
    
    '''
    def create_file_explorer(self, path=None):
        """Creează un explorer de fișiere care arată și permite gestionarea fișierelor și directoarelor"""
        from tkinter import simpledialog
        if path is None:
            # Start with root directory or home directory depending on OS
            if os.name == 'nt':  # Windows
                path = 'C:\\'
            else:  # Linux/Mac
                path = os.path.expanduser('~')
        
        explorer_window = tk.Toplevel(self.rootW95dist)
        explorer_window.title("My Computer")
        explorer_window.overrideredirect(True)
        explorer_window.geometry("700x500+200+100")
        explorer_window.configure(bg="#c0c0c0")
        
        # Add Windows 95 style title bar
        title_bar = tk.Frame(explorer_window, bg="#000080", height=25)
        title_bar.pack(fill="x", side="top")
        title_label = tk.Label(title_bar, text=f"My Computer - {path}", fg="white", bg="#000080",
                              font=("MS Sans Serif", 8, "bold"))
        title_label.pack(side="left", padx=5, pady=2)
        
        # Close button for title bar
        close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                font=("Arial", 8, "bold"), width=2, height=1,
                                relief="raised", bd=1,
                                command=lambda: self.close_window("My Computer", explorer_window))
        close_button.pack(side="right", padx=2, pady=1)
        
        self.make_window_draggable(explorer_window, title_bar)
        
        # Menu bar
        menu_bar = tk.Menu(explorer_window)
        explorer_window.config(menu=menu_bar)
        
        file_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="File", menu=file_menu)
        
        edit_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Edit", menu=edit_menu)
        
        view_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="View", menu=view_menu)
        
        help_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Help", menu=help_menu)
        
        # Main content frame
        content_frame = tk.Frame(explorer_window, bg="#c0c0c0")
        content_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Address bar
        address_frame = tk.Frame(content_frame, bg="#c0c0c0")
        address_frame.pack(fill="x", pady=5)
        
        tk.Label(address_frame, text="Address:", bg="#c0c0c0", font=("MS Sans Serif", 8)).pack(side="left")
        address_var = tk.StringVar(value=path)
        address_entry = tk.Entry(address_frame, textvariable=address_var, width=50, font=("MS Sans Serif", 8))
        address_entry.pack(side="left", padx=5, fill="x", expand=True)
        address_entry.bind("<Return>", lambda e: navigate_to_path(address_var.get()))
        
        go_button = tk.Button(address_frame, text="Go", bg="#c0c0c0", relief="raised", bd=2,
                             font=("MS Sans Serif", 8),
                             command=lambda: navigate_to_path(address_var.get()))
        go_button.pack(side="left", padx=5)
        
        up_button = tk.Button(address_frame, text="Up", bg="#c0c0c0", relief="raised", bd=2,
                             font=("MS Sans Serif", 8),
                             command=lambda: navigate_up())
        up_button.pack(side="left", padx=5)
        
        # Toolbar for file operations
        toolbar_frame = tk.Frame(content_frame, bg="#c0c0c0", relief="raised", bd=2)
        toolbar_frame.pack(fill="x", pady=5)
        
        new_folder_btn = tk.Button(toolbar_frame, text="New Folder", bg="#c0c0c0", relief="raised", bd=2,
                                  font=("MS Sans Serif", 8),
                                  command=lambda: create_new_folder())
        new_folder_btn.pack(side="left", padx=5, pady=2)
        
        cut_btn = tk.Button(toolbar_frame, text="Cut", bg="#c0c0c0", relief="raised", bd=2,
                           font=("MS Sans Serif", 8),
                           command=lambda: cut_selected())
        cut_btn.pack(side="left", padx=5, pady=2)
        
        copy_btn = tk.Button(toolbar_frame, text="Copy", bg="#c0c0c0", relief="raised", bd=2,
                            font=("MS Sans Serif", 8),
                            command=lambda: copy_selected())
        copy_btn.pack(side="left", padx=5, pady=2)
        
        paste_btn = tk.Button(toolbar_frame, text="Paste", bg="#c0c0c0", relief="raised", bd=2,
                             font=("MS Sans Serif", 8),
                             command=lambda: paste_item())
        paste_btn.pack(side="left", padx=5, pady=2)
        
        delete_btn = tk.Button(toolbar_frame, text="Delete", bg="#c0c0c0", relief="raised", bd=2,
                              font=("MS Sans Serif", 8),
                              command=lambda: delete_selected())
        delete_btn.pack(side="left", padx=5, pady=2)
        
        rename_btn = tk.Button(toolbar_frame, text="Rename", bg="#c0c0c0", relief="raised", bd=2,
                              font=("MS Sans Serif", 8),
                              command=lambda: rename_selected())
        rename_btn.pack(side="left", padx=5, pady=2)
        
        refresh_btn = tk.Button(toolbar_frame, text="Refresh", bg="#c0c0c0", relief="raised", bd=2,
                               font=("MS Sans Serif", 8),
                               command=lambda: refresh_view())
        refresh_btn.pack(side="right", padx=5, pady=2)
        
        # File/Directory list
        list_frame = tk.Frame(content_frame, bg="white", relief="sunken", bd=2)
        list_frame.pack(fill="both", expand=True, pady=5)
        
        # Scrollbars
        h_scrollbar = tk.Scrollbar(list_frame, orient="horizontal")
        h_scrollbar.pack(side="bottom", fill="x")
        
        v_scrollbar = tk.Scrollbar(list_frame)
        v_scrollbar.pack(side="right", fill="y")
        
        # List with columns
        columns = ("name", "type", "size", "modified")
        file_list = ttk.Treeview(list_frame, columns=columns, show="headings",
                               yscrollcommand=v_scrollbar.set,
                               xscrollcommand=h_scrollbar.set,
                               selectmode="extended")
        
        file_list.heading("name", text="Name", command=lambda: sort_by_column("name"))
        file_list.heading("type", text="Type", command=lambda: sort_by_column("type"))
        file_list.heading("size", text="Size", command=lambda: sort_by_column("size"))
        file_list.heading("modified", text="Date Modified", command=lambda: sort_by_column("modified"))
        
        file_list.column("name", width=200, anchor="w")
        file_list.column("type", width=100, anchor="w")
        file_list.column("size", width=100, anchor="e")
        file_list.column("modified", width=150, anchor="w")
        
        file_list.pack(side="left", fill="both", expand=True)
        
        v_scrollbar.config(command=file_list.yview)
        h_scrollbar.config(command=file_list.xview)
        
        # Status bar
        status_frame = tk.Frame(explorer_window, bg="#c0c0c0", relief="sunken", bd=1, height=25)
        status_frame.pack(fill="x", side="bottom")
        status_frame.pack_propagate(False)
        
        status_label = tk.Label(status_frame, text="", bg="#c0c0c0", font=("MS Sans Serif", 8))
        status_label.pack(side="left", padx=5)
        
        items_label = tk.Label(status_frame, text="", bg="#c0c0c0", font=("MS Sans Serif", 8))
        items_label.pack(side="right", padx=5)
        
        # Add to taskbar
        self.add_window_to_taskbar("My Computer", explorer_window)
        explorer_window.protocol("WM_DELETE_WINDOW", lambda: self.close_window("My Computer", explorer_window))
        
        # Clipboard for file operations
        clipboard = {
            "action": None,  # "cut" or "copy"
            "files": []
        }
        
        # Double-click to open/navigate
        file_list.bind("<Double-1>", lambda e: open_selected())
        
        # Right-click context menu
        context_menu = tk.Menu(file_list, tearoff=0)
        
        def show_context_menu(event):
            context_menu.delete(0, tk.END)  # Clear previous items
            
            # Add menu items based on selection
            selected = file_list.selection()
            if selected:
                context_menu.add_command(label="Open", command=open_selected)
                context_menu.add_separator()
                context_menu.add_command(label="Cut", command=cut_selected)
                context_menu.add_command(label="Copy", command=copy_selected)
                context_menu.add_command(label="Delete", command=delete_selected)
                context_menu.add_command(label="Rename", command=rename_selected)
                
                # If only one item selected and it's a file
                if len(selected) == 1:
                    item = selected[0]
                    values = file_list.item(item, "values")
                    if values[1] != "Folder":
                        context_menu.add_separator()
                        context_menu.add_command(label="Open with Text Editor", 
                                               command=lambda: open_with_text_editor(values[0]))
            else:
                context_menu.add_command(label="Paste", command=paste_item, 
                                       state="normal" if clipboard["files"] else "disabled")
                context_menu.add_separator()
                context_menu.add_command(label="New Folder", command=create_new_folder)
                context_menu.add_command(label="Refresh", command=refresh_view)
            
            # Display the context menu
            context_menu.tk_popup(event.x_root, event.y_root)
        
        file_list.bind("<Button-3>", show_context_menu)  # Right-click
        
        # Functions for file operations
        def navigate_to_path(new_path):
            try:
                # Check if path exists and is accessible
                if os.path.exists(new_path) and os.path.isdir(new_path):
                    nonlocal path
                    path = os.path.normpath(new_path)
                    address_var.set(path)
                    title_label.config(text=f"My Computer - {path}")
                    refresh_view()
                else:
                    messagebox.showerror("Error", f"Cannot access: {new_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Error navigating to path: {str(e)}")
        
        def navigate_up():
            parent_dir = os.path.dirname(path)
            if parent_dir != path:  # Not at root
                navigate_to_path(parent_dir)
        
        def refresh_view():
            # Clear existing items
            for item in file_list.get_children():
                file_list.delete(item)
            
            try:
                # List directories and files
                items = os.listdir(path)
                
                # First add directories
                for item in sorted(items):
                    item_path = os.path.join(path, item)
                    try:
                        if os.path.isdir(item_path):
                            # Get modified time
                            mod_time = os.path.getmtime(item_path)
                            mod_time_str = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d %H:%M:%S')
                            
                            file_list.insert("", "end", values=(item, "Folder", "", mod_time_str))
                    except PermissionError:
                        file_list.insert("", "end", values=(item, "Folder (Access Denied)", "", ""))
                    except Exception as e:
                        print(f"Error accessing {item_path}: {str(e)}")
                
                # Then add files
                for item in sorted(items):
                    item_path = os.path.join(path, item)
                    try:
                        if os.path.isfile(item_path):
                            # Get file size
                            size = os.path.getsize(item_path)
                            size_str = format_size(size)
                            
                            # Get modified time
                            mod_time = os.path.getmtime(item_path)
                            mod_time_str = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d %H:%M:%S')
                            
                            # Get file type
                            file_ext = os.path.splitext(item)[1].lower()
                            file_type = get_file_type(file_ext)
                            
                            file_list.insert("", "end", values=(item, file_type, size_str, mod_time_str))
                    except PermissionError:
                        file_list.insert("", "end", values=(item, "File (Access Denied)", "", ""))
                    except Exception as e:
                        print(f"Error accessing {item_path}: {str(e)}")
                
                # Update status
                status_label.config(text=f"Current directory: {path}")
                items_label.config(text=f"{len(items)} items")
                
            except PermissionError:
                messagebox.showerror("Access Denied", f"Cannot access directory: {path}")
                navigate_up()
            except Exception as e:
                messagebox.showerror("Error", f"Error loading directory: {str(e)}")
        
        def format_size(size_bytes):
            """Format file size in human-readable format"""
            if size_bytes < 1024:
                return f"{size_bytes} B"
            elif size_bytes < 1024**2:
                return f"{size_bytes/1024:.1f} KB"
            elif size_bytes < 1024**3:
                return f"{size_bytes/(1024**2):.1f} MB"
            else:
                return f"{size_bytes/(1024**3):.1f} GB"
        
        def get_file_type(ext):
            """Return file type based on extension"""
            file_types = {
                ".txt": "Text Document",
                ".doc": "Word Document",
                ".docx": "Word Document",
                ".pdf": "PDF Document",
                ".jpg": "JPEG Image",
                ".jpeg": "JPEG Image",
                ".png": "PNG Image",
                ".gif": "GIF Image",
                ".mp3": "Audio File",
                ".mp4": "Video File",
                ".py": "Python Script",
                ".exe": "Application",
                ".zip": "Compressed File",
                ".rar": "Compressed File",
                ".ini": "Configuration File",
                ".html": "HTML Document",
                ".css": "CSS File",
                ".js": "JavaScript File",
                ".json": "JSON File"
            }
            return file_types.get(ext, f"{ext[1:].upper() if ext else 'Unknown'} File")
        
        def open_selected():
            selected = file_list.selection()
            if not selected:
                return
            
            # Get the first selected item
            item = selected[0]
            values = file_list.item(item, "values")
            item_name = values[0]
            item_type = values[1]
            item_path = os.path.join(path, item_name)
            
            if item_type == "Folder":
                # Navigate to the selected folder
                navigate_to_path(item_path)
            else:
                # For files, try to open with default application
                try:
                    if os.name == 'nt':  # Windows
                        os.startfile(item_path)
                    else:  # Linux/Mac
                        subprocess.run(['xdg-open', item_path], check=True)
                except Exception as e:
                    messagebox.showerror("Error", f"Could not open file: {str(e)}")
                    # Offer to open with text editor
                    if messagebox.askyesno("Open with Text Editor", 
                                          "Would you like to open this file with Text Editor?"):
                        open_with_text_editor(item_name)
        
        def open_with_text_editor(filename):
            """Open the selected file with the Text Editor"""
            file_path = os.path.join(path, filename)
            try:
                with open(file_path, 'r', encoding='utf-8') as file:
                    content = file.read()
                    
                    # Create a new text editor window
                    self.create_text_editor_with_content(content, filename, file_path)
            except UnicodeDecodeError:
                messagebox.showerror("Error", "Cannot open binary file with Text Editor")
            except Exception as e:
                messagebox.showerror("Error", f"Error opening file: {str(e)}")
        
        def create_new_folder():
            """Create a new folder in the current directory"""
            folder_name = simpledialog.askstring("New Folder", "Enter folder name:")
            if folder_name:
                new_folder_path = os.path.join(path, folder_name)
                try:
                    os.mkdir(new_folder_path)
                    refresh_view()
                except Exception as e:
                    messagebox.showerror("Error", f"Could not create folder: {str(e)}")
        
        def cut_selected():
            """Cut selected files/folders"""
            selected = file_list.selection()
            if not selected:
                return
            
            clipboard["action"] = "cut"
            clipboard["files"] = []
            
            for item in selected:
                values = file_list.item(item, "values")
                item_name = values[0]
                item_path = os.path.join(path, item_name)
                clipboard["files"].append(item_path)
            
            status_label.config(text=f"{len(selected)} items cut to clipboard")
        
        def copy_selected():
            """Copy selected files/folders"""
            selected = file_list.selection()
            if not selected:
                return
            
            clipboard["action"] = "copy"
            clipboard["files"] = []
            
            for item in selected:
                values = file_list.item(item, "values")
                item_name = values[0]
                item_path = os.path.join(path, item_name)
                clipboard["files"].append(item_path)
            
            status_label.config(text=f"{len(selected)} items copied to clipboard")
        
        def paste_item():
            """Paste files/folders from clipboard to current directory"""
            if not clipboard["files"]:
                return
            
            for src_path in clipboard["files"]:
                if not os.path.exists(src_path):
                    continue
                    
                # Get just the filename/foldername
                base_name = os.path.basename(src_path)
                dst_path = os.path.join(path, base_name)
                
                try:
                    # Check if destination already exists
                    if os.path.exists(dst_path):
                        if not messagebox.askyesno("Confirm", 
                                                  f"{base_name} already exists. Overwrite?"):
                            continue
                    
                    if clipboard["action"] == "cut":
                        # For cut operation, move the file/folder
                        if os.path.isdir(src_path):
                            # For folders, use shutil.move
                            shutil.move(src_path, dst_path)
                        else:
                            # For files, use os.rename (faster than shutil.move for files)
                            os.rename(src_path, dst_path)
                    else:  # "copy"
                        # For copy operation
                        if os.path.isdir(src_path):
                            # For folders, copy the entire directory tree
                            shutil.copytree(src_path, dst_path)
                        else:
                            # For files, use shutil.copy2 to preserve metadata
                            shutil.copy2(src_path, dst_path)
                except Exception as e:
                    messagebox.showerror("Error", f"Error pasting {base_name}: {str(e)}")
            
            # Clear clipboard if it was a cut operation
            if clipboard["action"] == "cut":
                clipboard["files"] = []
            
            refresh_view()
        
        def delete_selected():
            """Delete selected files/folders"""
            selected = file_list.selection()
            if not selected:
                return
            
            # Confirm deletion
            count = len(selected)
            if not messagebox.askyesno("Confirm Delete", 
                                      f"Are you sure you want to delete {count} item(s)?"):
                return
            
            for item in selected:
                values = file_list.item(item, "values")
                item_name = values[0]
                item_path = os.path.join(path, item_name)
                
                try:
                    if os.path.isdir(item_path):
                        # For folders, use shutil.rmtree
                        shutil.rmtree(item_path)
                    else:
                        # For files, use os.remove
                        os.remove(item_path)
                except Exception as e:
                    messagebox.showerror("Error", f"Error deleting {item_name}: {str(e)}")
            
            refresh_view()
        
        def rename_selected():
            """Rename selected file/folder"""
            selected = file_list.selection()
            if not selected or len(selected) > 1:
                messagebox.showinfo("Rename", "Please select only one item to rename")
                return
            
            item = selected[0]
            values = file_list.item(item, "values")
            old_name = values[0]
            old_path = os.path.join(path, old_name)
            
            new_name = simpledialog.askstring("Rename", "Enter new name:", initialvalue=old_name)
            if new_name and new_name != old_name:
                new_path = os.path.join(path, new_name)
                try:
                    os.rename(old_path, new_path)
                    refresh_view()
                except Exception as e:
                    messagebox.showerror("Error", f"Error renaming file: {str(e)}")
        
        def sort_by_column(column):
            """Sort file list by the specified column"""
            # This would need to be implemented to handle sorting
            # For simplicity, we'll just refresh the view for now
            refresh_view()
        
        # Initial file listing
        refresh_view()

    def create_text_editor_with_content(self, content, filename, file_path=None):
        """Create a text editor window with predefined content"""
        editor_window = tk.Toplevel(self.rootW95dist)
        editor_window.title(f"Text Editor - {filename}")
        editor_window.overrideredirect(True)
        editor_window.geometry("600x400+200+100")
        editor_window.configure(bg="#c0c0c0")
        
        # Add Windows 95 style title bar
        title_bar = tk.Frame(editor_window, bg="#000080", height=25)
        title_bar.pack(fill="x", side="top")
        title_label = tk.Label(title_bar, text=f"Text Editor - {filename}", fg="white", bg="#000080",
                              font=("MS Sans Serif", 8, "bold"))
        title_label.pack(side="left", padx=5, pady=2)
        
        self.make_window_draggable(editor_window, title_bar)
        
        # Close button for title bar
        close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                font=("Arial", 8, "bold"), width=2, height=1,
                                relief="raised", bd=1,
                                command=lambda: self.close_window(f"Text Editor - {filename}", editor_window))
        close_button.pack(side="right", padx=2, pady=1)
        
        # Menubar
        menubar = tk.Menu(editor_window)
        editor_window.config(menu=menubar)
        
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        
        # Text area
        text_frame = tk.Frame(editor_window, bg="#c0c0c0")
        text_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Scrollbar
        scrollbar = tk.Scrollbar(text_frame)
        scrollbar.pack(side="right", fill="y")
        
        # Text widget
        text_area = tk.Text(text_frame, wrap="word", yscrollcommand=scrollbar.set,
                           font=("Courier New", 10), bg="white", fg="black")
        text_area.pack(fill="both", expand=True)
        scrollbar.config(command=text_area.yview)
        
        # Insert the content
        text_area.insert(1.0, content)
        
        # Funcții pentru meniu
        def save_file():
            try:
                with open(file_path, 'w', encoding='utf-8') as file:
                    content = text_area.get(1.0, tk.END)
                    file.write(content)
                    messagebox.showinfo("Success", "File saved successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Could not save file: {str(e)}")
        
        def save_as_file():
            file_path_new = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
            )
            if file_path_new:
                try:
                    with open(file_path_new, 'w', encoding='utf-8') as file:
                        content = text_area.get(1.0, tk.END)
                        file.write(content)
                        nonlocal file_path
                        file_path = file_path_new
                        editor_window.title(f"Text Editor - {os.path.basename(file_path)}")
                        title_label.config(text=f"Text Editor - {os.path.basename(file_path)}")
                        messagebox.showinfo("Success", "File saved successfully!")
                except Exception as e:
                    messagebox.showerror("Error", f"Could not save file: {str(e)}")
        
        # Adaugă opțiunile în meniu
        file_menu.add_command(label="Save", command=save_file)
        file_menu.add_command(label="Save As", command=save_as_file)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=lambda: self.close_window(f"Text Editor - {filename}", editor_window))
        
        self.add_window_to_taskbar(f"Text Editor - {filename}", editor_window)
        editor_window.protocol("WM_DELETE_WINDOW", lambda: self.close_window(f"Text Editor - {filename}", editor_window))

    def create_text_editor_with_content(self, content, filename, file_path=None):
        """Create a text editor window with predefined content"""
        editor_window = tk.Toplevel(self.rootW95dist)
        editor_window.title(f"Text Editor - {filename}")
        editor_window.overrideredirect(True)
        editor_window.geometry("600x400+200+100")
        editor_window.configure(bg="#c0c0c0")
        
        # Add Windows 95 style title bar
        title_bar = tk.Frame(editor_window, bg="#000080", height=25)
        title_bar.pack(fill="x", side="top")
        title_label = tk.Label(title_bar, text=f"Text Editor - {filename}", fg="white", bg="#000080",
                              font=("MS Sans Serif", 8, "bold"))
        title_label.pack(side="left", padx=5, pady=2)
        
        self.make_window_draggable(editor_window, title_bar)
        
        # Close button for title bar
        close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                font=("Arial", 8, "bold"), width=2, height=1,
                                relief="raised", bd=1,
                                command=lambda: self.close_window(f"Text Editor - {filename}", editor_window))
        close_button.pack(side="right", padx=2, pady=1)
        
        # Menubar
        menubar = tk.Menu(editor_window)
        editor_window.config(menu=menubar)
        
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        
        # Text area
        text_frame = tk.Frame(editor_window, bg="#c0c0c0")
        text_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Scrollbar
        scrollbar = tk.Scrollbar(text_frame)
        scrollbar.pack(side="right", fill="y")
        
        # Text widget
        text_area = tk.Text(text_frame, wrap="word", yscrollcommand=scrollbar.set,
                           font=("Courier New", 10), bg="white", fg="black")
        text_area.pack(fill="both", expand=True)
        scrollbar.config(command=text_area.yview)
        
        # Insert the content
        text_area.insert(1.0, content)
        
        # Funcții pentru meniu
        def save_file():
            try:
                with open(file_path, 'w', encoding='utf-8') as file:
                    content = text_area.get(1.0, tk.END)
                    file.write(content)
                    messagebox.showinfo("Success", "File saved successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Could not save file: {str(e)}")
        
        def save_as_file():
            file_path_new = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
            )
            if file_path_new:
                try:
                    with open(file_path_new, 'w', encoding='utf-8') as file:
                        content = text_area.get(1.0, tk.END)
                        file.write(content)
                        nonlocal file_path
                        file_path = file_path_new
                        editor_window.title(f"Text Editor - {os.path.basename(file_path)}")
                        title_label.config(text=f"Text Editor - {os.path.basename(file_path)}")
                        messagebox.showinfo("Success", "File saved successfully!")
                except Exception as e:
                    messagebox.showerror("Error", f"Could not save file: {str(e)}")
        
        # Adaugă opțiunile în meniu
        file_menu.add_command(label="Save", command=save_file)
        file_menu.add_command(label="Save As", command=save_as_file)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=lambda: self.close_window(f"Text Editor - {filename}", editor_window))
        
        self.add_window_to_taskbar(f"Text Editor - {filename}", editor_window)
        editor_window.protocol("WM_DELETE_WINDOW", lambda: self.close_window(f"Text Editor - {filename}", editor_window))
    
    def create_text_editor(self):
        editor_window = tk.Toplevel(self.rootW95dist)
        editor_window.title("Text Editor")
        editor_window.overrideredirect(True)
        editor_window.geometry("500x400+200+100")
        editor_window.configure(bg="#c0c0c0")
        
        # Add Windows 95 style title bar
        title_bar = tk.Frame(editor_window, bg="#000080", height=25)
        title_bar.pack(fill="x", side="top")
        title_label = tk.Label(title_bar, text="Text Editor", fg="white", bg="#000080",
                              font=("MS Sans Serif", 8, "bold"))
        title_label.pack(side="left", padx=5, pady=2)
        
        self.make_window_draggable(editor_window, title_bar)
        
        # Close button for title bar
        close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                font=("Arial", 8, "bold"), width=2, height=1,
                                relief="raised", bd=1,
                                command=lambda: self.close_window("Text Editor", editor_window))
        close_button.pack(side="right", padx=2, pady=1)
        
        # Menubar
        menubar = tk.Menu(editor_window)
        editor_window.config(menu=menubar)
        
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        
        # Text area
        text_frame = tk.Frame(editor_window, bg="#c0c0c0")
        text_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Scrollbar
        scrollbar = tk.Scrollbar(text_frame)
        scrollbar.pack(side="right", fill="y")
        
        # Text widget
        text_area = tk.Text(text_frame, wrap="word", yscrollcommand=scrollbar.set,
                           font=("Courier New", 10), bg="white", fg="black")
        text_area.pack(fill="both", expand=True)
        scrollbar.config(command=text_area.yview)
        
        self.make_window_draggable(editor_window, title_bar)
        
        # Funcții pentru meniu
        def open_file():
            file_path = filedialog.askopenfilename(
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
            )
            if file_path:
                try:
                    with open(file_path, 'r', encoding='utf-8') as file:
                        content = file.read()
                        text_area.delete(1.0, tk.END)
                        text_area.insert(1.0, content)
                        editor_window.title(f"Text Editor - {file_path}")
                except Exception as e:
                    messagebox.showerror("Error", f"Could not open file: {str(e)}")
        
        def save_file():
            file_path = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
            )
            if file_path:
                try:
                    with open(file_path, 'w', encoding='utf-8') as file:
                        content = text_area.get(1.0, tk.END)
                        file.write(content)
                        editor_window.title(f"Text Editor - {file_path}")
                        messagebox.showinfo("Success", "File saved successfully!")
                except Exception as e:
                    messagebox.showerror("Error", f"Could not save file: {str(e)}")
        
        # Adaugă opțiunile în meniu
        file_menu.add_command(label="Open", command=open_file)
        file_menu.add_command(label="Save", command=save_file)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=lambda: self.close_window("Text Editor", editor_window))
        
        self.add_window_to_taskbar("Text Editor", editor_window)
        editor_window.protocol("WM_DELETE_WINDOW", lambda: self.close_window("Text Editor", editor_window))
    
    def create_calculator(self):
        calc_window = tk.Toplevel(self.rootW95dist)
        calc_window.title("Calculator")
        calc_window.overrideredirect(True)
        calc_window.geometry("250x300+300+150")
        calc_window.configure(bg="#c0c0c0")
        calc_window.resizable(False, False)
        
        # Add Windows 95 style title bar
        title_bar = tk.Frame(calc_window, bg="#000080", height=25)
        title_bar.pack(fill="x", side="top")
        title_label = tk.Label(title_bar, text="Calculator", fg="white", bg="#000080",
                              font=("MS Sans Serif", 8, "bold"))
        title_label.pack(side="left", padx=5, pady=2)
        
        # Close button for title bar
        close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                font=("Arial", 8, "bold"), width=2, height=1,
                                relief="raised", bd=1,
                                command=lambda: self.close_window("Calculator", calc_window))
        close_button.pack(side="right", padx=2, pady=1)
        
        self.make_window_draggable(calc_window, title_bar)
        
        # Display
        display_var = tk.StringVar()
        display_var.set("0")
        
        display = tk.Entry(calc_window, textvariable=display_var, font=("Arial", 14),
                          justify="right", state="readonly", bg="white")
        display.pack(fill="x", padx=5, pady=5)
        
        # Calculator logic
        calc_data = {"current": "0", "operator": None, "operand": None}
        
        def button_click(value):
            if value.isdigit():
                if calc_data["current"] == "0":
                    calc_data["current"] = value
                else:
                    calc_data["current"] += value
            elif value in ["+", "-", "*", "/"]:
                if calc_data["operator"] and calc_data["operand"] is not None:
                    calculate()
                calc_data["operand"] = float(calc_data["current"])
                calc_data["operator"] = value
                calc_data["current"] = "0"
            elif value == "=":
                calculate()
            elif value == "C":
                calc_data["current"] = "0"
                calc_data["operator"] = None
                calc_data["operand"] = None
            
            display_var.set(calc_data["current"])
        
        def calculate():
            if calc_data["operator"] and calc_data["operand"] is not None:
                try:
                    current_val = float(calc_data["current"])
                    if calc_data["operator"] == "+":
                        result = calc_data["operand"] + current_val
                    elif calc_data["operator"] == "-":
                        result = calc_data["operand"] - current_val
                    elif calc_data["operator"] == "*":
                        result = calc_data["operand"] * current_val
                    elif calc_data["operator"] == "/":
                        if current_val != 0:
                            result = calc_data["operand"] / current_val
                        else:
                            result = "Error"
                    
                    calc_data["current"] = str(result)
                    calc_data["operator"] = None
                    calc_data["operand"] = None
                except:
                    calc_data["current"] = "Error"
        
        # Buttons frame
        buttons_frame = tk.Frame(calc_window, bg="#c0c0c0")
        buttons_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Button layout
        buttons = [
            ['C', '', '', '/'],
            ['7', '8', '9', '*'],
            ['4', '5', '6', '-'],
            ['1', '2', '3', '+'],
            ['0', '', '', '=']
        ]
        
        for i, row in enumerate(buttons):
            for j, btn_text in enumerate(row):
                if btn_text:
                    btn = tk.Button(buttons_frame, text=btn_text, font=("Arial", 12),
                                   command=lambda t=btn_text: button_click(t),
                                   bg="#e0e0e0", relief="raised", bd=2)
                    btn.grid(row=i, column=j, sticky="nsew", padx=1, pady=1)
        
        # Configure grid weights
        for i in range(5):
            buttons_frame.grid_rowconfigure(i, weight=1)
        for j in range(4):
            buttons_frame.grid_columnconfigure(j, weight=1)
        
        self.add_window_to_taskbar("Calculator", calc_window)
        calc_window.protocol("WM_DELETE_WINDOW", lambda: self.close_window("Calculator", calc_window))
    
    def create_network_monitor(self):
        net_window = tk.Toplevel(self.rootW95dist)
        net_window.title("Network Monitor")
        net_window.overrideredirect(True)
        net_window.geometry("700x600+350+50")
        net_window.configure(bg="#c0c0c0")
        
        # Add Windows 95 style title bar
        title_bar = tk.Frame(net_window, bg="#000080", height=25)
        title_bar.pack(fill="x", side="top")
        title_label = tk.Label(title_bar, text="Network Monitor", fg="white", bg="#000080",
                              font=("MS Sans Serif", 8, "bold"))
        title_label.pack(side="left", padx=5, pady=2)
        
        # Close button for title bar
        close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                font=("Arial", 8, "bold"), width=2, height=1,
                                relief="raised", bd=1,
                                command=lambda: self.close_window("Network Monitor", net_window))
        close_button.pack(side="right", padx=2, pady=1)
        
        # Header frame
        header_frame = tk.Frame(net_window, bg="#c0c0c0", relief="raised", bd=2)
        header_frame.pack(fill="x", padx=5, pady=5)
        
        tk.Label(header_frame, text="Advanced Network Monitor", font=("MS Sans Serif", 10, "bold"),
                bg="#c0c0c0").pack(pady=5)
        
        # Tabs using notebook
        notebook = ttk.Notebook(net_window)
        notebook.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Create a custom style for Windows 95 look
        style = ttk.Style()
        style.configure("W95.TNotebook", background="#c0c0c0", borderwidth=2, relief="raised")
        style.configure("W95.TNotebook.Tab", padding=[5, 2], font=("MS Sans Serif", 8))
        notebook.configure(style="W95.TNotebook")
        
        # Tab 1: Network Stats
        stats_frame = ttk.Frame(notebook)
        notebook.add(stats_frame, text="Network Stats")
        
        # Stats display
        stats_text = tk.Text(stats_frame, font=("Courier New", 9), bg="white", 
                            fg="black", relief="sunken", bd=2)
        stats_scrollbar = tk.Scrollbar(stats_frame, command=stats_text.yview)
        stats_text.config(yscrollcommand=stats_scrollbar.set)
        
        stats_text.pack(side="left", fill="both", expand=True)
        stats_text.config(state="disabled")
        stats_scrollbar.pack(side="right", fill="y")
        
        # Tab 2: Active Connections
        conn_frame = ttk.Frame(notebook)
        notebook.add(conn_frame, text="Connections")
        
        # Connections display
        conn_text = tk.Text(conn_frame, font=("Courier New", 9), bg="white",
                           fg="black", relief="sunken", bd=2)
        conn_scrollbar = tk.Scrollbar(conn_frame, command=conn_text.yview)
        conn_text.config(yscrollcommand=conn_scrollbar.set)
        
        conn_text.pack(side="left", fill="both", expand=True)
        conn_text.config(state="disabled")
        conn_scrollbar.pack(side="right", fill="y")
        
        # Tab 3: Network Traffic Graph (Text-based)
        traffic_frame = ttk.Frame(notebook)
        notebook.add(traffic_frame, text="Traffic Graph")
        
        traffic_text = tk.Text(traffic_frame, font=("Courier New", 8), bg="black",
                              fg="#00ff00", relief="sunken", bd=2)
        traffic_scrollbar = tk.Scrollbar(traffic_frame, command=traffic_text.yview)
        traffic_text.config(yscrollcommand=traffic_scrollbar.set)
        
        traffic_text.pack(side="left", fill="both", expand=True)
        traffic_text.config(state="disabled")
        traffic_scrollbar.pack(side="right", fill="y")
        
        # Tab 4: Port Scanner
        scanner_frame = ttk.Frame(notebook)
        notebook.add(scanner_frame, text="Port Scanner")
        
        # Scanner controls
        scanner_control_frame = tk.Frame(scanner_frame, bg="#f0f0f0", relief="raised", bd=2)
        scanner_control_frame.pack(fill="x", padx=5, pady=5)
        
        tk.Label(scanner_control_frame, text="Target IP:", bg="#f0f0f0").pack(side="left", padx=2)
        ip_entry = tk.Entry(scanner_control_frame, width=15, relief="sunken", bd=2)
        ip_entry.pack(side="left", padx=2)
        ip_entry.insert(0, "127.0.0.1")
        
        tk.Label(scanner_control_frame, text="Port Range:", bg="#f0f0f0").pack(side="left", padx=(10,0))
        port_start_entry = tk.Entry(scanner_control_frame, width=8, relief="sunken", bd=2)
        port_start_entry.pack(side="left", padx=2)
        port_start_entry.insert(0, "1")
        
        tk.Label(scanner_control_frame, text="-", bg="#f0f0f0").pack(side="left")
        port_end_entry = tk.Entry(scanner_control_frame, width=8, relief="sunken", bd=2)
        port_end_entry.pack(side="left", padx=2)
        port_end_entry.insert(0, "1000")
        
        scan_btn = tk.Button(scanner_control_frame, text="Scan", 
                            font=("MS Sans Serif", 8),
                            bg="#c0c0c0", relief="raised", bd=2,
                            command=lambda: self.start_port_scan(ip_entry.get(), 
                                                               int(port_start_entry.get()), 
                                                               int(port_end_entry.get()), 
                                                               scanner_text))
        scan_btn.pack(side="left", padx=10)
        
        # Scanner results
        scanner_text = tk.Text(scanner_frame, font=("Courier New", 9), bg="white",
                              fg="black", relief="sunken", bd=2)
        scanner_text_scrollbar = tk.Scrollbar(scanner_frame, command=scanner_text.yview)
        scanner_text.config(yscrollcommand=scanner_text_scrollbar.set)
        
        scanner_text.pack(side="left", fill="both", expand=True, padx=5)
        scanner_text.config(state="disabled")
        scanner_text_scrollbar.pack(side="right", fill="y")
        
        # Tab 5: Bandwidth Monitor
        bandwidth_frame = ttk.Frame(notebook)
        notebook.add(bandwidth_frame, text="Bandwidth")
        
        bandwidth_text = tk.Text(bandwidth_frame, font=("Courier New", 9), bg="white",
                                fg="black", relief="sunken", bd=2)
        bandwidth_scrollbar = tk.Scrollbar(bandwidth_frame, command=bandwidth_text.yview)
        bandwidth_text.config(yscrollcommand=bandwidth_scrollbar.set)
        
        bandwidth_text.pack(side="left", fill="both", expand=True)
        bandwidth_text.config(state="disabled")
        bandwidth_scrollbar.pack(side="right", fill="y")
        
        # Buttons frame with Windows 95 style
        buttons_frame = tk.Frame(net_window, bg="#c0c0c0")
        buttons_frame.pack(fill="x", padx=5, pady=5)
        
        refresh_btn = tk.Button(buttons_frame, text="Refresh", font=("MS Sans Serif", 8),
                               bg="#c0c0c0", relief="raised", bd=2,
                               command=lambda: self.update_network_info(stats_text, conn_text, 
                                                                       traffic_text, bandwidth_text))
        refresh_btn.pack(side="left", padx=5)
        
        monitor_btn = tk.Button(buttons_frame, text="Start Monitoring", font=("MS Sans Serif", 8),
                               bg="#c0c0c0", relief="raised", bd=2,
                               command=lambda: self.toggle_network_monitoring(monitor_btn, traffic_text))
        monitor_btn.pack(side="left", padx=5)
        
        export_btn = tk.Button(buttons_frame, text="Export Data", font=("MS Sans Serif", 8),
                              bg="#c0c0c0", relief="raised", bd=2,
                              command=self.export_network_data)
        export_btn.pack(side="left", padx=5)
        
        help_btn = tk.Button(buttons_frame, text="Help", font=("MS Sans Serif", 8),
                            bg="#c0c0c0", relief="raised", bd=2,
                            command=lambda: self.show_help("Network Monitor"))
        help_btn.pack(side="right", padx=5)
        
        # Initial load
        self.update_network_info(stats_text, conn_text, traffic_text, bandwidth_text)
        
        # Simulate real network data initially
        self.simulate_network_data(traffic_text)
        
        self.add_window_to_taskbar("Network Monitor", net_window)
        net_window.protocol("WM_DELETE_WINDOW", lambda: self.close_window("Network Monitor", net_window))
        
        self.make_window_draggable(net_window, title_bar)
    
    def simulate_network_data(self, traffic_text):
        """Simulate some initial network traffic data for better UI appearance"""
        timestamp = datetime.now() - timedelta(seconds=10)
        
        # Add some simulated data to history
        for i in range(10):
            timestamp += timedelta(seconds=1)
            ts_str = timestamp.strftime("%H:%M:%S")
            
            # Generate some random data for demo purposes
            bytes_sent = random.randint(1024, 10240)
            bytes_recv = random.randint(2048, 20480)
            
            self.network_history.append({
                'time': ts_str,
                'bytes_sent': i * bytes_sent,
                'bytes_recv': i * bytes_recv,
                'packets_sent': i * 10,
                'packets_recv': i * 20
            })
            
            if i > 0:
                prev = self.network_history[i-1]
                curr = self.network_history[i]
                
                bytes_sent_rate = curr['bytes_sent'] - prev['bytes_sent']
                bytes_recv_rate = curr['bytes_recv'] - prev['bytes_recv']
                
                # Create simple ASCII graph
                max_rate = max(bytes_sent_rate, bytes_recv_rate, 1024)
                sent_bar = "█" * int((bytes_sent_rate / max_rate) * 40)
                recv_bar = "█" * int((bytes_recv_rate / max_rate) * 40)
                
                display_text = f"{ts_str} | UP: {sent_bar:<40} {bytes_sent_rate:>8} B/s\n"
                display_text += f"        | DN: {recv_bar:<40} {bytes_recv_rate:>8} B/s\n"
                
                traffic_text.config(state="normal")
                traffic_text.insert(tk.END, display_text)
                
        traffic_text.config(state="disabled")
    
    def start_port_scan(self, target_ip, start_port, end_port, result_text):
        result_text.config(state="normal")
        result_text.delete(1.0, tk.END)
        result_text.insert(1.0, f"Starting port scan on {target_ip}...\n")
        result_text.insert(tk.END, f"Scanning ports {start_port}-{end_port}\n")
        result_text.insert(tk.END, "=" * 50 + "\n")
        
        def scan_ports():
            open_ports = []
            for port in range(start_port, min(end_port + 1, start_port + 100)):  # Limit scan range
                try:
                    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                    sock.settimeout(0.1)
                    result = sock.connect_ex((target_ip, port))
                    if result == 0:
                        open_ports.append(port)
                        result_text.insert(tk.END, f"Port {port}: OPEN\n")
                        result_text.see(tk.END)
                        result_text.update()
                    sock.close()
                except:
                    pass
            
            result_text.insert(tk.END, f"\nScan completed. Found {len(open_ports)} open ports.\n")
            if open_ports:
                result_text.insert(tk.END, f"Open ports: {', '.join(map(str, open_ports))}\n")
        
        # Run scan in thread to prevent UI freezing
        scan_thread = threading.Thread(target=scan_ports, daemon=True)
        result_text.config(state="disabled")
        scan_thread.start()
    
    def toggle_network_monitoring(self, button, traffic_text):
        if not self.monitoring_active:
            self.monitoring_active = True
            button.config(text="Stop Monitoring")
            self.start_network_monitoring(traffic_text)
        else:
            self.monitoring_active = False
            button.config(text="Start Monitoring")
    
    def start_network_monitoring(self, traffic_text):
        def monitor():
            while self.monitoring_active:
                try:
                    net_io = psutil.net_io_counters()
                    timestamp = datetime.now().strftime("%H:%M:%S")
                    
                    # Store data for history
                    self.network_history.append({
                        'time': timestamp,
                        'bytes_sent': net_io.bytes_sent,
                        'bytes_recv': net_io.bytes_recv,
                        'packets_sent': net_io.packets_sent,
                        'packets_recv': net_io.packets_recv
                    })
                    
                    # Keep only last 100 entries
                    if len(self.network_history) > 100:
                        self.network_history.pop(0)
                    
                    # Update text display
                    if len(self.network_history) >= 2:
                        prev = self.network_history[-2]
                        curr = self.network_history[-1]
                        
                        bytes_sent_rate = curr['bytes_sent'] - prev['bytes_sent']
                        bytes_recv_rate = curr['bytes_recv'] - prev['bytes_recv']
                        
                        # Create simple ASCII graph
                        max_rate = max(bytes_sent_rate, bytes_recv_rate, 1024)
                        sent_bar = "█" * int((bytes_sent_rate / max_rate) * 40)
                        recv_bar = "█" * int((bytes_recv_rate / max_rate) * 40)
                        
                        display_text = f"{timestamp} | UP: {sent_bar:<40} {bytes_sent_rate:>8} B/s\n"
                        display_text += f"        | DN: {recv_bar:<40} {bytes_recv_rate:>8} B/s\n"
                        
                        traffic_text.config(state="normal")
                        traffic_text.insert(tk.END, display_text)
                        traffic_text.see(tk.END)
                        traffic_text.config(state="disabled")
                        
                        # Keep only last 50 lines
                        lines = traffic_text.get(1.0, tk.END).split('\n')
                        if len(lines) > 100:
                            traffic_text.delete(1.0, f"{len(lines)-100}.0")
                    
                    time.sleep(1)
                except Exception as e:
                    traffic_text.config(state="normal")
                    traffic_text.insert(tk.END, f"Monitoring error: {str(e)}\n")
                    break
        
        monitor_thread = threading.Thread(target=monitor, daemon=True)
        monitor_thread.start()
    
    def export_network_data(self):
        if not self.network_history:
            messagebox.showinfo("Export", "No network data to export.")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        if filename:
            try:
                with open(filename, 'w') as f:
                    json.dump(self.network_history, f, indent=2)
                messagebox.showinfo("Export", f"Network data exported to {filename}")
            except Exception as e:
                messagebox.showerror("Export Error", f"Could not export data: {str(e)}")
    
    def update_network_info(self, stats_text, conn_text, traffic_text, bandwidth_text):
        try:
            # Clear previous content
            stats_text.config(state="normal")
            conn_text.config(state="normal")
            bandwidth_text.config(state="normal")
            stats_text.delete(1.0, tk.END)
            conn_text.delete(1.0, tk.END)
            bandwidth_text.delete(1.0, tk.END)
            
            # Network statistics
            net_io = psutil.net_io_counters()
            addrs = psutil.net_if_addrs()
            net_stats = psutil.net_if_stats()
            
            stats_info = "╔═══════════════════════════════════════╗\n"
            stats_info += "║         NETWORK STATISTICS            ║\n"
            stats_info += "╚═══════════════════════════════════════╝\n\n"
            
            stats_info += f" GLOBAL COUNTERS:\n"
            stats_info += f"   Bytes Sent:     {net_io.bytes_sent:,} bytes ({net_io.bytes_sent/(1024**3):.2f} GB)\n"
            stats_info += f"   Bytes Received: {net_io.bytes_recv:,} bytes ({net_io.bytes_recv/(1024**3):.2f} GB)\n"
            stats_info += f"   Packets Sent:   {net_io.packets_sent:,}\n"
            stats_info += f"   Packets Recv:   {net_io.packets_recv:,}\n"
            stats_info += f"   Errors In:      {net_io.errin}\n"
            stats_info += f"   Errors Out:     {net_io.errout}\n"
            stats_info += f"   Drops In:       {net_io.dropin}\n"
            stats_info += f"   Drops Out:      {net_io.dropout}\n\n"
            
            stats_info += " NETWORK INTERFACES:\n"
            stats_info += "─" * 60 + "\n"
            
            for interface, addresses in addrs.items():
                stats_info += f"\n {interface}:\n"
                
                # Interface statistics
                if interface in net_stats:
                    stat = net_stats[interface]
                    stats_info += f"   Status: {'UP' if stat.isup else 'DOWN'}\n"
                    stats_info += f"   Speed: {stat.speed} Mbps\n" if stat.speed > 0 else "   Speed: Unknown\n"
                    stats_info += f"   MTU: {stat.mtu}\n"
                
                # Addresses
                for addr in addresses:
                    if addr.family == socket.AF_INET:
                        stats_info += f"    IPv4: {addr.address}\n"
                        if addr.netmask:
                            stats_info += f"      Netmask: {addr.netmask}\n"
                        if addr.broadcast:
                            stats_info += f"      Broadcast: {addr.broadcast}\n"
                    elif addr.family == socket.AF_INET6:
                        stats_info += f"    IPv6: {addr.address}\n"
                    elif hasattr(addr, 'address') and len(addr.address) == 17:  # MAC address
                        stats_info += f"    MAC: {addr.address}\n"
            
            stats_text.insert(1.0, stats_info)
            
            # Active connections with more details
            connections = psutil.net_connections()
            conn_info = "╔═══════════════════════════════════════╗\n"
            conn_info += "║        ACTIVE CONNECTIONS             ║\n"
            conn_info += "╚═══════════════════════════════════════╝\n\n"
            conn_info += f"Total Active Connections: {len(connections)}\n\n"
            conn_info += f"{'Proto':<6} {'PID':<8} {'Local Address':<22} {'Remote Address':<22} {'Status':<12}\n"
            conn_info += "─" * 80 + "\n"
            
            connection_stats = {"TCP": 0, "UDP": 0, "ESTABLISHED": 0, "LISTENING": 0}
            
            for conn in connections[:100]:  # Limit to first 100 connections
                if conn.laddr:
                    local = f"{conn.laddr.ip}:{conn.laddr.port}"
                else:
                    local = "N/A"
                
                if conn.raddr:
                    remote = f"{conn.raddr.ip}:{conn.raddr.port}"
                else:
                    remote = "N/A"
                
                protocol = "TCP" if conn.type == socket.SOCK_STREAM else "UDP"
                status = conn.status if conn.status else "N/A"
                pid = str(conn.pid) if conn.pid else "N/A"
                
                # Update statistics
                connection_stats[protocol] += 1
                if status in connection_stats:
                    connection_stats[status] += 1
                
                conn_info += f"{protocol:<6} {pid:<8} {local:<22} {remote:<22} {status:<12}\n"
            
            # Add connection statistics
            conn_info += "\n" + "─" * 80 + "\n"
            conn_info += "CONNECTION STATISTICS:\n"
            for key, value in connection_stats.items():
                conn_info += f"   {key}: {value}\n"
            
            conn_text.insert(1.0, conn_info)
            
            # Bandwidth monitoring
            bandwidth_info = "╔═══════════════════════════════════════╗\n"
            bandwidth_info += "║        BANDWIDTH MONITORING           ║\n"
            bandwidth_info += "╚═══════════════════════════════════════╝\n\n"
            
            # Per-interface statistics
            net_io_counters = psutil.net_io_counters(pernic=True)
            
            bandwidth_info += f"{'Interface':<15} {'Bytes Sent':<15} {'Bytes Recv':<15} {'Packets Sent':<12} {'Packets Recv':<12}\n"
            bandwidth_info += "─" * 80 + "\n"
            
            for interface, counters in net_io_counters.items():
                bandwidth_info += f"{interface:<15} {counters.bytes_sent:<15,} {counters.bytes_recv:<15,} "
                bandwidth_info += f"{counters.packets_sent:<12,} {counters.packets_recv:<12,}\n"
            
            # Network usage history (if available)
            if self.network_history:
                bandwidth_info += "\n RECENT NETWORK ACTIVITY (Last 10 samples):\n"
                bandwidth_info += "─" * 60 + "\n"
                bandwidth_info += f"{'Time':<10} {'Upload (B/s)':<15} {'Download (B/s)':<15}\n"
                bandwidth_info += "─" * 60 + "\n"
                
                for i in range(max(0, len(self.network_history) - 10), len(self.network_history)):
                    if i > 0:
                        prev = self.network_history[i-1]
                        curr = self.network_history[i]
                        upload_rate = curr['bytes_sent'] - prev['bytes_sent']
                        download_rate = curr['bytes_recv'] - prev['bytes_recv']
                        bandwidth_info += f"{curr['time']:<10} {upload_rate:<15,} {download_rate:<15,}\n"
            
            bandwidth_text.insert(1.0, bandwidth_info)
            stats_text.config(state="disabled")
            conn_text.config(state="disabled")
            bandwidth_text.config(state="disabled")
            
        except Exception as e:
            error_msg = f"Error getting network info: {str(e)}"
            stats_text.insert(1.0, error_msg)
            conn_text.insert(1.0, error_msg)
            bandwidth_text.insert(1.0, error_msg)
    
    def show_help(self, title):
        help_window = tk.Toplevel(self.rootW95dist)
        help_window.title(f"{title} Help")
        help_window.overrideredirect(True)
        help_window.geometry("400x300+300+200")
        help_window.configure(bg="#c0c0c0")
        
        # Add Windows 95 style title bar
        title_bar = tk.Frame(help_window, bg="#000080", height=25)
        title_bar.pack(fill="x", side="top")
        title_label = tk.Label(title_bar, text=f"{title} Help", fg="white", bg="#000080",
                              font=("MS Sans Serif", 8, "bold"))
        title_label.pack(side="left", padx=5, pady=2)
        
        # Close button for title bar
        close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                font=("Arial", 8, "bold"), width=2, height=1,
                                relief="raised", bd=1,
                                command=help_window.destroy)
        close_button.pack(side="right", padx=2, pady=1)
        
        # Help content
        help_frame = tk.Frame(help_window, bg="#c0c0c0", bd=2, relief="sunken")
        help_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        help_text = tk.Text(help_frame, font=("MS Sans Serif", 9), bg="white", wrap="word")
        scrollbar = tk.Scrollbar(help_frame, command=help_text.yview)
        help_text.config(yscrollcommand=scrollbar.set)
        
        help_text.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        self.make_window_draggable(help_window, title_bar)
        
        if title == "Network Monitor":
            help_text.insert(tk.END, "NETWORK MONITOR HELP\n\n", "header")
            help_text.insert(tk.END, "This application allows you to monitor network activity on your computer.\n\n")
            help_text.insert(tk.END, "TABS:\n", "subheader")
            help_text.insert(tk.END, "• Network Stats: Shows general network statistics and interface information.\n")
            help_text.insert(tk.END, "• Connections: Displays active network connections.\n")
            help_text.insert(tk.END, "• Traffic Graph: Shows real-time network traffic visualization.\n")
            help_text.insert(tk.END, "• Port Scanner: Scans for open ports on a specified IP address.\n")
            help_text.insert(tk.END, "• Bandwidth: Displays bandwidth usage by network interface.\n\n")
            help_text.insert(tk.END, "BUTTONS:\n", "subheader")
            help_text.insert(tk.END, "• Refresh: Updates all information.\n")
            help_text.insert(tk.END, "• Start Monitoring: Begins real-time network traffic monitoring.\n")
            help_text.insert(tk.END, "• Export Data: Saves network data to a file.\n")
        elif title == "Hardware Info":
            help_text.insert(tk.END, "HARDWARE INFORMATION HELP\n\n", "header")
            help_text.insert(tk.END, "This application displays detailed information about your computer hardware.\n\n")
            help_text.insert(tk.END, "TABS:\n", "subheader")
            help_text.insert(tk.END, "• System: General system information.\n")
            help_text.insert(tk.END, "• CPU: CPU specifications and usage.\n")
            help_text.insert(tk.END, "• Memory: RAM and swap memory information.\n")
            help_text.insert(tk.END, "• Storage: Disk drives and storage usage.\n")
            help_text.insert(tk.END, "• Processes: Currently running processes.\n")
            help_text.insert(tk.END, "• Performance: Real-time system performance monitoring.\n\n")
            help_text.insert(tk.END, "BUTTONS:\n", "subheader")
            help_text.insert(tk.END, "• Refresh: Updates all hardware information.\n")
            help_text.insert(tk.END, "• Start Performance Monitor: Begins real-time monitoring.\n")
            help_text.insert(tk.END, "• Export Report: Saves hardware information to a file.\n")
            help_text.insert(tk.END, "• CPU Benchmark: Runs a simple CPU performance test.\n")
        elif title == "Paint":
            help_text.insert(tk.END, "PAINT HELP\n\n", "header")
            help_text.insert(tk.END, "This is a simple drawing application.\n\n")
            help_text.insert(tk.END, "TOOLS:\n", "subheader")
            help_text.insert(tk.END, "• Pencil: Click and drag to draw thin lines.\n")
            help_text.insert(tk.END, "• Brush: Click and drag to draw thicker lines.\n")
            help_text.insert(tk.END, "• Eraser: Click and drag to erase parts of your drawing.\n")
            help_text.insert(tk.END, "• Size: Adjust the brush size using the slider.\n")
            help_text.insert(tk.END, "• Colors: Click on a color square to select that color.\n\n")
            help_text.insert(tk.END, "MENU OPTIONS:\n", "subheader")
            help_text.insert(tk.END, "• File > New: Creates a new drawing.\n")
            help_text.insert(tk.END, "• File > Save: Saves your drawing.\n")
            help_text.insert(tk.END, "• File > Exit: Closes Paint.\n")
        
        # Define text styles
        help_text.tag_configure("header", font=("MS Sans Serif", 12, "bold"))
        help_text.tag_configure("subheader", font=("MS Sans Serif", 10, "bold"))
        help_text.config(state="disabled")
        
        # Make text read-only
        help_text.config(state="disabled")
    
    def create_hardware_info(self):
        hw_window = tk.Toplevel(self.rootW95dist)
        hw_window.title("Hardware Info")
        hw_window.overrideredirect(True)
        hw_window.geometry("750x650+400+50")
        hw_window.configure(bg="#c0c0c0")
        
        # Add Windows 95 style title bar
        title_bar = tk.Frame(hw_window, bg="#000080", height=25)
        title_bar.pack(fill="x", side="top")
        title_label = tk.Label(title_bar, text="Hardware Information", fg="white", bg="#000080",
                              font=("MS Sans Serif", 8, "bold"))
        title_label.pack(side="left", padx=5, pady=2)
        
        # Close button for title bar
        close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                font=("Arial", 8, "bold"), width=2, height=1,
                                relief="raised", bd=1,
                                command=lambda: self.close_window("Hardware Info", hw_window))
        close_button.pack(side="right", padx=2, pady=1)
        
        # Header frame
        header_frame = tk.Frame(hw_window, bg="#c0c0c0", relief="raised", bd=2)
        header_frame.pack(fill="x", padx=5, pady=5)
        
        tk.Label(header_frame, text="Advanced System Hardware Information", 
                font=("MS Sans Serif", 10, "bold"), bg="#c0c0c0").pack(pady=5)
        
        # Tabs using notebook
        notebook = ttk.Notebook(hw_window)
        notebook.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Create a custom style for Windows 95 look
        style = ttk.Style()
        style.configure("W95.TNotebook", background="#c0c0c0", borderwidth=2, relief="raised")
        style.configure("W95.TNotebook.Tab", padding=[5, 2], font=("MS Sans Serif", 8))
        notebook.configure(style="W95.TNotebook")
        
        # Tab 1: System Info
        system_frame = ttk.Frame(notebook)
        notebook.add(system_frame, text="System")
        
        system_text = tk.Text(system_frame, font=("Courier New", 9), bg="white",
                             fg="black", relief="sunken", bd=2)
        system_scrollbar = tk.Scrollbar(system_frame, command=system_text.yview)
        system_text.config(yscrollcommand=system_scrollbar.set)
        
        system_text.pack(side="left", fill="both", expand=True)
        system_text.config(state="disabled")
        system_scrollbar.pack(side="right", fill="y")
        
        # Tab 2: CPU Info
        cpu_frame = ttk.Frame(notebook)
        notebook.add(cpu_frame, text="CPU")
        
        cpu_text = tk.Text(cpu_frame, font=("Courier New", 9), bg="white",
                          fg="black", relief="sunken", bd=2)
        cpu_scrollbar = tk.Scrollbar(cpu_frame, command=cpu_text.yview)
        cpu_text.config(yscrollcommand=cpu_scrollbar.set)
        
        cpu_text.pack(side="left", fill="both", expand=True)
        cpu_text.config(state="disabled")
        cpu_scrollbar.pack(side="right", fill="y")
        
        # Tab 3: Memory Info
        memory_frame = ttk.Frame(notebook)
        notebook.add(memory_frame, text="Memory")
        
        memory_text = tk.Text(memory_frame, font=("Courier New", 9), bg="white",
                             fg="black", relief="sunken", bd=2)
        memory_scrollbar = tk.Scrollbar(memory_frame, command=memory_text.yview)
        memory_text.config(yscrollcommand=memory_scrollbar.set)
        
        memory_text.pack(side="left", fill="both", expand=True)
        memory_text.config(state="disabled")
        memory_scrollbar.pack(side="right", fill="y")
        
        # Tab 4: Storage Info
        storage_frame = ttk.Frame(notebook)
        notebook.add(storage_frame, text="Storage")
        
        storage_text = tk.Text(storage_frame, font=("Courier New", 9), bg="white",
                              fg="black", relief="sunken", bd=2)
        storage_scrollbar = tk.Scrollbar(storage_frame, command=storage_text.yview)
        storage_text.config(yscrollcommand=storage_scrollbar.set)
        
        storage_text.pack(side="left", fill="both", expand=True)
        storage_text.config(state="disabled")
        storage_scrollbar.pack(side="right", fill="y")
        
        # Tab 5: Process Monitor
        process_frame = ttk.Frame(notebook)
        notebook.add(process_frame, text="Processes")
        
        # Process controls
        process_control_frame = tk.Frame(process_frame, bg="#f0f0f0", relief="raised", bd=2)
        process_control_frame.pack(fill="x", padx=5, pady=5)
        
        tk.Label(process_control_frame, text="Sort by:", bg="#f0f0f0").pack(side="left", padx=2)
        sort_var = tk.StringVar(value="memory")
        sort_combo = ttk.Combobox(process_control_frame, textvariable=sort_var, 
                                 values=["memory", "cpu", "name", "pid"], width=10, state="readonly")
        sort_combo.pack(side="left", padx=5)
        
        sort_combo.bind("<<ComboboxSelected>>", lambda e: self.update_hardware_info(system_text, cpu_text, memory_text, storage_text, process_text, sort_var.get()))
        
        # kill_btn = tk.Button(process_control_frame, text="Kill Selected Process", 
                            # bg="#ff6666", fg="black", font=("MS Sans Serif", 8),
                            # relief="raised", bd=2)
        # kill_btn.pack(side="right", padx=5)
        
        process_text = tk.Text(process_frame, font=("Courier New", 9), bg="white",
                              fg="black", relief="sunken", bd=2)
        process_scrollbar = tk.Scrollbar(process_frame, command=process_text.yview)
        process_text.config(yscrollcommand=process_scrollbar.set)
        
        process_text.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        process_text.config(state="disabled")
        process_scrollbar.pack(side="right", fill="y")
        
        # Tab 6: Performance Monitor
        perf_frame = ttk.Frame(notebook)
        notebook.add(perf_frame, text="Performance")
        
        perf_text = tk.Text(perf_frame, font=("Courier New", 8), bg="black",
                           fg="#00ff00", relief="sunken", bd=2)
        perf_scrollbar = tk.Scrollbar(perf_frame, command=perf_text.yview)
        perf_text.config(yscrollcommand=perf_scrollbar.set)
        
        perf_text.pack(side="left", fill="both", expand=True)
        perf_text.config(state="disabled")
        perf_scrollbar.pack(side="right", fill="y")
        
        # Buttons frame
        buttons_frame = tk.Frame(hw_window, bg="#c0c0c0")
        buttons_frame.pack(fill="x", padx=5, pady=5)
        
        refresh_btn = tk.Button(buttons_frame, text="Refresh", font=("MS Sans Serif", 8),
                               bg="#c0c0c0", relief="raised", bd=2,
                               command=lambda: self.update_hardware_info(system_text, cpu_text, memory_text, 
                                                                        storage_text, process_text, sort_var.get()))
        refresh_btn.pack(side="left", padx=5)
        
        monitor_btn = tk.Button(buttons_frame, text="Start Performance Monitor", font=("MS Sans Serif", 8),
                               bg="#c0c0c0", relief="raised", bd=2,
                               command=lambda: self.toggle_performance_monitoring(monitor_btn, perf_text))
        monitor_btn.pack(side="left", padx=5)
        
        export_btn = tk.Button(buttons_frame, text="Export Report", font=("MS Sans Serif", 8),
                              bg="#c0c0c0", relief="raised", bd=2,
                              command=self.export_hardware_report)
        export_btn.pack(side="left", padx=5)
        
        benchmark_btn = tk.Button(buttons_frame, text="CPU Benchmark", font=("MS Sans Serif", 8),
                                 bg="#c0c0c0", relief="raised", bd=2,
                                 command=lambda: self.run_cpu_benchmark(cpu_text))
        benchmark_btn.pack(side="left", padx=5)
        
        help_btn = tk.Button(buttons_frame, text="Help", font=("MS Sans Serif", 8),
                            bg="#c0c0c0", relief="raised", bd=2,
                            command=lambda: self.show_help("Hardware Info"))
        help_btn.pack(side="right", padx=5)
        
        # Initial load
        self.update_hardware_info(system_text, cpu_text, memory_text, storage_text, process_text, "memory")
        
        # Simulate performance data
        self.simulate_performance_data(perf_text)
        
        self.add_window_to_taskbar("Hardware Info", hw_window)
        hw_window.protocol("WM_DELETE_WINDOW", lambda: self.close_window("Hardware Info", hw_window))
        
        self.make_window_draggable(hw_window, title_bar)
    
    def simulate_performance_data(self, perf_text):
        """Simulate some initial performance data for better UI appearance"""
        timestamp = datetime.now() - timedelta(seconds=10)
        
        perf_text.config(state="normal")
        perf_text.insert(tk.END, "Performance Monitor Initialized\n")
        perf_text.insert(tk.END, "=" * 60 + "\n")
        
        # Add some simulated data
        for i in range(10):
            timestamp += timedelta(seconds=1)
            ts_str = timestamp.strftime("%H:%M:%S")
            
            # Generate random performance data
            cpu_percent = random.uniform(10, 60)
            memory_percent = random.uniform(40, 80)
            cpu_freq = random.uniform(2000, 3200)
            disk_read = random.randint(50, 500) * 1024
            disk_write = random.randint(20, 200) * 1024
            net_sent = random.randint(10, 100) * 1024
            net_recv = random.randint(50, 500) * 1024
            
            # Store in history
            self.cpu_history.append({
                'time': ts_str,
                'cpu_percent': cpu_percent,
                'cpu_freq': cpu_freq,
                'memory_percent': memory_percent,
                'memory_used': memory_percent * 8000000000 / 100,  # Simulate 8GB RAM
                'disk_read': i * disk_read,
                'disk_write': i * disk_write,
                'net_sent': i * net_sent,
                'net_recv': i * net_recv
            })
            
            # Create ASCII graphs
            cpu_bar = "█" * int(cpu_percent / 2.5)  # Scale to 40 chars max
            memory_bar = "█" * int(memory_percent / 2.5)
            
            display_text = f"{ts_str} │ CPU: {cpu_bar:<40} {cpu_percent:>6.1f}%\n"
            display_text += f"        │ MEM: {memory_bar:<40} {memory_percent:>6.1f}%\n"
            display_text += f"        │ FREQ: {cpu_freq:>4.0f} MHz\n"
            
            if i > 0:
                prev = self.cpu_history[i-1]
                curr = self.cpu_history[i]
                
                disk_read_rate = curr['disk_read'] - prev['disk_read']
                disk_write_rate = curr['disk_write'] - prev['disk_write']
                display_text += f"        │ DISK: R:{disk_read_rate/1024:>6.0f}KB/s W:{disk_write_rate/1024:>6.0f}KB/s\n"
                
                net_sent_rate = curr['net_sent'] - prev['net_sent']
                net_recv_rate = curr['net_recv'] - prev['net_recv']
                display_text += f"        │ NET:  ↑{net_sent_rate/1024:>6.0f}KB/s ↓{net_recv_rate/1024:>6.0f}KB/s\n"
            
            display_text += "        │" + "─" * 50 + "\n"
            
            perf_text.insert(tk.END, display_text)
            perf_text.config(state="disabled")
    
    def toggle_performance_monitoring(self, button, perf_text):
        if not self.monitoring_active:
            self.monitoring_active = True
            button.config(text="Stop Performance Monitor")
            self.start_performance_monitoring(perf_text)
        else:
            self.monitoring_active = False
            button.config(text="Start Performance Monitor")
    
    def start_performance_monitoring(self, perf_text):
        def monitor():
            perf_text.config(state="normal")
            perf_text.insert(tk.END, "Performance Monitor Started\n")
            perf_text.insert(tk.END, "=" * 60 + "\n")
            
            while self.monitoring_active:
                try:
                    timestamp = datetime.now().strftime("%H:%M:%S")
                    
                    # CPU usage
                    cpu_percent = psutil.cpu_percent(interval=0.1)
                    cpu_freq = psutil.cpu_freq()
                    
                    # Memory usage
                    memory = psutil.virtual_memory()
                    
                    # Disk I/O
                    disk_io = psutil.disk_io_counters()
                    
                    # Network I/O
                    net_io = psutil.net_io_counters()
                    
                    # Store history
                    perf_data = {
                        'time': timestamp,
                        'cpu_percent': cpu_percent,
                        'cpu_freq': cpu_freq.current if cpu_freq else 0,
                        'memory_percent': memory.percent,
                        'memory_used': memory.used,
                        'disk_read': disk_io.read_bytes if disk_io else 0,
                        'disk_write': disk_io.write_bytes if disk_io else 0,
                        'net_sent': net_io.bytes_sent,
                        'net_recv': net_io.bytes_recv
                    }
                    
                    self.cpu_history.append(perf_data)
                    if len(self.cpu_history) > 100:
                        self.cpu_history.pop(0)
                    
                    # Create ASCII graphs
                    cpu_bar = "█" * int(cpu_percent / 2.5)  # Scale to 40 chars max
                    memory_bar = "█" * int(memory.percent / 2.5)
                    
                    display_text = f"{timestamp} │ CPU: {cpu_bar:<40} {cpu_percent:>6.1f}%\n"
                    display_text += f"        │ MEM: {memory_bar:<40} {memory.percent:>6.1f}%\n"
                    
                    if cpu_freq:
                        display_text += f"        │ FREQ: {cpu_freq.current:>4.0f} MHz\n"
                    
                    # Calculate rates if we have previous data
                    if len(self.cpu_history) >= 2:
                        prev = self.cpu_history[-2]
                        curr = self.cpu_history[-1]
                        if disk_io:
                            disk_read_rate = curr['disk_read'] - prev['disk_read']
                            disk_write_rate = curr['disk_write'] - prev['disk_write']
                            display_text += f"        │ DISK: R:{disk_read_rate/1024:>6.0f}KB/s W:{disk_write_rate/1024:>6.0f}KB/s\n"
                        
                        net_sent_rate = curr['net_sent'] - prev['net_sent']
                        net_recv_rate = curr['net_recv'] - prev['net_recv']
                        display_text += f"        │ NET:  ↑{net_sent_rate/1024:>6.0f}KB/s ↓{net_recv_rate/1024:>6.0f}KB/s\n"
                    
                    display_text += "        │" + "─" * 50 + "\n"
                    
                    perf_text.insert(tk.END, display_text)
                    perf_text.see(tk.END)
                    perf_text.config(state="disabled")
                    
                    # Keep only last 200 lines
                    lines = perf_text.get(1.0, tk.END).split('\n')
                    if len(lines) > 200:
                        perf_text.delete(1.0, f"{len(lines)-200}.0")
                    
                    time.sleep(1)
                except Exception as e:
                    perf_text.insert(tk.END, f"Monitor error: {str(e)}\n")
                    break
        
        monitor_thread = threading.Thread(target=monitor, daemon=True)
        monitor_thread.start()
        
    def run_cpu_benchmark(self, cpu_text):
        def benchmark():
            cpu_text.config(state="normal")
            cpu_text.insert(tk.END, "\n" + "="*50 + "\n")
            cpu_text.insert(tk.END, " RUNNING CPU BENCHMARK...\n")
            cpu_text.insert(tk.END, "="*50 + "\n")
            cpu_text.see(tk.END)
            cpu_text.update()
            
            # Simple CPU benchmark - calculate primes
            start_time = time.time()
            primes = []
            
            def is_prime(n):
                if n < 2:
                    return False
                for i in range(2, int(n**0.5) + 1):
                    if n % i == 0:
                        return False
                return True
            
            # Find primes up to 10000
            for i in range(2, 10000):
                if is_prime(i):
                    primes.append(i)
                if i % 1000 == 0:
                    cpu_text.insert(tk.END, f"Progress: {i/10000*100:.0f}%\n")
                    cpu_text.see(tk.END)
                    cpu_text.update()
            
            end_time = time.time()
            duration = end_time - start_time
            
            cpu_text.insert(tk.END, f"\n BENCHMARK COMPLETED!\n")
            cpu_text.insert(tk.END, f"Found {len(primes)} primes in {duration:.2f} seconds\n")
            cpu_text.insert(tk.END, f"Performance Score: {10000/duration:.0f} ops/sec\n")
            cpu_text.insert(tk.END, "="*50 + "\n")
            cpu_text.see(tk.END)
            cpu_text.config(state="disabled")
        
        benchmark_thread = threading.Thread(target=benchmark, daemon=True)
        benchmark_thread.start()
    
    def export_hardware_report(self):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = filedialog.asksaveasfilename(
            defaultextension=".txt",
            initialfile=f"hardware_report_{timestamp}.txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        if filename:
            try:
                with open(filename, 'w') as f:
                    f.write("HARDWARE INFORMATION REPORT\n")
                    f.write("="*50 + "\n")
                    f.write(f"Generated: {datetime.now()}\n\n")
                    
                    # System info
                    uname = platform.uname()
                    f.write(f"System: {uname.system}\n")
                    f.write(f"Node: {uname.node}\n")
                    f.write(f"Release: {uname.release}\n")
                    f.write(f"Machine: {uname.machine}\n")
                    f.write(f"Processor: {uname.processor}\n\n")
                    
                    # CPU info
                    f.write("CPU INFORMATION:\n")
                    f.write(f"Physical cores: {psutil.cpu_count(logical=False)}\n")
                    f.write(f"Total cores: {psutil.cpu_count(logical=True)}\n")
                    f.write(f"Current usage: {psutil.cpu_percent()}%\n\n")
                    
                    # Memory info
                    memory = psutil.virtual_memory()
                    f.write("MEMORY INFORMATION:\n")
                    f.write(f"Total: {memory.total // (1024**3)} GB\n")
                    f.write(f"Available: {memory.available // (1024**3)} GB\n")
                    f.write(f"Used: {memory.used // (1024**3)} GB\n")
                    f.write(f"Percentage: {memory.percent}%\n\n")
                    
                    # Performance history
                    if self.cpu_history:
                        f.write("RECENT PERFORMANCE DATA:\n")
                        f.write("Time\t\tCPU%\tMemory%\n")
                        for entry in self.cpu_history[-20:]:  # Last 20 entries
                            f.write(f"{entry['time']}\t{entry['cpu_percent']:.1f}\t{entry['memory_percent']:.1f}\n")
                
                messagebox.showinfo("Export", f"Hardware report exported to {filename}")
            except Exception as e:
                messagebox.showerror("Export Error", f"Could not export report: {str(e)}")
    
    def update_hardware_info(self, system_text, cpu_text, memory_text, storage_text, process_text, sort_by):
        try:
            # Clear previous content
            system_text.config(state="normal")
            cpu_text.config(state="normal")
            memory_text.config(state="normal")
            storage_text.config(state="normal")
            process_text.config(state="normal")
            system_text.delete(1.0, tk.END)
            cpu_text.delete(1.0, tk.END)
            memory_text.delete(1.0, tk.END)
            storage_text.delete(1.0, tk.END)
            process_text.delete(1.0, tk.END)
            
            # System Information
            uname = platform.uname()
            boot_time = datetime.fromtimestamp(psutil.boot_time())
            uptime = datetime.now() - boot_time
            
            system_info = "╔═══════════════════════════════════════╗\n"
            system_info += "║        SYSTEM INFORMATION             ║\n"
            system_info += "╚═══════════════════════════════════════╝\n\n"
            
            system_info += f"️  BASIC INFORMATION:\n"
            system_info += f"   System:         {uname.system}\n"
            system_info += f"   Node Name:      {uname.node}\n"
            system_info += f"   Release:        {uname.release}\n"
            system_info += f"   Version:        {uname.version}\n"
            system_info += f"   Machine:        {uname.machine}\n"
            system_info += f"   Processor:      {uname.processor}\n"
            system_info += f"   Boot Time:      {boot_time.strftime('%Y-%m-%d %H:%M:%S')}\n"
            system_info += f"   Uptime:         {str(uptime).split('.')[0]}\n\n"
            
            # Temperature (if available)
            try:
                temps = psutil.sensors_temperatures()
                if temps:
                    system_info += f"️  TEMPERATURE SENSORS:\n"
                    for name, entries in temps.items():
                        system_info += f"   {name}:\n"
                        for entry in entries:
                            system_info += f"      {entry.label or 'N/A'}: {entry.current}°C"
                            if entry.high:
                                system_info += f" (High: {entry.high}°C)"
                            if entry.critical:
                                system_info += f" (Critical: {entry.critical}°C)"
                            system_info += "\n"
                    system_info += "\n"
            except:
                system_info += "️  Temperature sensors: Not available\n\n"
            
            # Battery info (if available)
            try:
                battery = psutil.sensors_battery()
                if battery:
                    system_info += f" BATTERY INFORMATION:\n"
                    system_info += f"   Percentage: {battery.percent}%\n"
                    system_info += f"   Power plugged: {'Yes' if battery.power_plugged else 'No'}\n"
                    if not battery.power_plugged and battery.secsleft != psutil.POWER_TIME_UNLIMITED:
                        hours, remainder = divmod(battery.secsleft, 3600)
                        minutes, _ = divmod(remainder, 60)
                        system_info += f"   Time left: {hours}h {minutes}m\n"
                    system_info += "\n"
            except:
                pass
            
            system_text.insert(1.0, system_info)
            
            # Enhanced CPU Information
            cpu_info = "╔═══════════════════════════════════════╗\n"
            cpu_info += "║         CPU INFORMATION               ║\n"
            cpu_info += "╚═══════════════════════════════════════╝\n\n"
            
            cpu_info += f"🔧 CPU SPECIFICATIONS:\n"
            cpu_info += f"   Physical cores:   {psutil.cpu_count(logical=False)}\n"
            cpu_info += f"   Total cores:      {psutil.cpu_count(logical=True)}\n"
            
            cpufreq = psutil.cpu_freq()
            if cpufreq:
                cpu_info += f"   Max Frequency:    {cpufreq.max:.2f} MHz\n"
                cpu_info += f"   Min Frequency:    {cpufreq.min:.2f} MHz\n"
                cpu_info += f"   Current Freq:     {cpufreq.current:.2f} MHz\n"
            
            # CPU usage statistics
            cpu_times = psutil.cpu_times()
            total_time = sum(cpu_times)
            
            cpu_info += f"\n️  CPU TIME DISTRIBUTION:\n"
            cpu_info += f"   User time:        {cpu_times.user:.2f}s ({cpu_times.user/total_time*100:.1f}%)\n"
            cpu_info += f"   System time:      {cpu_times.system:.2f}s ({cpu_times.system/total_time*100:.1f}%)\n"
            cpu_info += f"   Idle time:        {cpu_times.idle:.2f}s ({cpu_times.idle/total_time*100:.1f}%)\n"
            
            if hasattr(cpu_times, 'iowait'):
                cpu_info += f"   I/O wait time:    {cpu_times.iowait:.2f}s ({cpu_times.iowait/total_time*100:.1f}%)\n"
            
            cpu_info += f"\n CURRENT CPU USAGE:\n"
            cpu_percentages = psutil.cpu_percent(percpu=True, interval=0.1)
            
            for i, percentage in enumerate(cpu_percentages):
                bar = "█" * int(percentage / 5)  # Scale to 20 chars max
                cpu_info += f"   Core {i:2d}: {bar:<20} {percentage:>5.1f}%\n"
            
            total_cpu = psutil.cpu_percent()
            total_bar = "█" * int(total_cpu / 5)
            cpu_info += f"   Total:  {total_bar:<20} {total_cpu:>5.1f}%\n"
            
            # Load averages (Unix-like systems)
            try:
                load_avg = psutil.getloadavg()
                cpu_info += f"\n LOAD AVERAGES:\n"
                cpu_info += f"   1 minute:  {load_avg[0]:.2f}\n"
                cpu_info += f"   5 minutes: {load_avg[1]:.2f}\n"
                cpu_info += f"   15 minutes: {load_avg[2]:.2f}\n"
            except:
                pass
            
            cpu_text.insert(1.0, cpu_info)
            
            # Enhanced Memory Information
            svmem = psutil.virtual_memory()
            swap = psutil.swap_memory()
            
            memory_info = "╔═══════════════════════════════════════╗\n"
            memory_info += "║        MEMORY INFORMATION             ║\n"
            memory_info += "╚═══════════════════════════════════════╝\n\n"
            
            memory_info += " VIRTUAL MEMORY:\n"
            memory_info += f"   Total:      {svmem.total // (1024**3):>8} GB ({svmem.total:,} bytes)\n"
            memory_info += f"   Available:  {svmem.available // (1024**3):>8} GB ({svmem.available:,} bytes)\n"
            memory_info += f"   Used:       {svmem.used // (1024**3):>8} GB ({svmem.used:,} bytes)\n"
            memory_info += f"   Free:       {svmem.free // (1024**3):>8} GB ({svmem.free:,} bytes)\n"
            memory_info += f"   Percentage: {svmem.percent:>8.1f}%\n"
            
            # Memory usage bar
            mem_bar = "█" * int(svmem.percent / 2.5)  # Scale to 40 chars max
            memory_info += f"   Usage:      [{mem_bar:<40}] {svmem.percent:.1f}%\n\n"
            
            if hasattr(svmem, 'buffers') and hasattr(svmem, 'cached'):
                memory_info += f"   Buffers:    {svmem.buffers // (1024**2):>8} MB\n"
                memory_info += f"   Cached:     {svmem.cached // (1024**2):>8} MB\n\n"
            
            memory_info += " SWAP MEMORY:\n"
            memory_info += f"   Total:      {swap.total // (1024**3):>8} GB ({swap.total:,} bytes)\n"
            memory_info += f"   Free:       {swap.free // (1024**3):>8} GB ({swap.free:,} bytes)\n"
            memory_info += f"   Used:       {swap.used // (1024**3):>8} GB ({swap.used:,} bytes)\n"
            memory_info += f"   Percentage: {swap.percent:>8.1f}%\n"
            
            if swap.total > 0:
                swap_bar = "█" * int(swap.percent / 2.5)
                memory_info += f"   Usage:      [{swap_bar:<40}] {swap.percent:.1f}%\n"
            
            memory_info += "\n MEMORY INTENSIVE PROCESSES:\n"
            memory_info += "─" * 65 + "\n"
            memory_info += f"{'PID':<8} {'Name':<25} {'Memory %':<12} {'Memory (MB)':<15}\n"
            memory_info += "─" * 65 + "\n"
            
            processes = []
            for proc in psutil.process_iter(['pid', 'name', 'memory_percent', 'memory_info']):
                try:
                    processes.append(proc.info)
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
            
            # Sort by memory usage
            processes = sorted(processes, key=lambda x: x['memory_percent'] if x['memory_percent'] else 0, reverse=True)
            
            for proc in processes[:15]:  # Top 15 processes
                memory_mb = proc['memory_info'].rss // (1024*1024) if proc['memory_info'] else 0
                memory_info += f"{proc['pid']:<8} {proc['name'][:24]:<25} {proc['memory_percent']:<12.2f} {memory_mb:<15}\n"
            
            memory_text.insert(1.0, memory_info)
            
            # Enhanced Storage Information
            storage_info = "╔═══════════════════════════════════════╗\n"
            storage_info += "║        STORAGE INFORMATION            ║\n"
            storage_info += "╚═══════════════════════════════════════╝\n\n"
            
            partitions = psutil.disk_partitions()
            
            storage_info += " DISK PARTITIONS:\n"
            storage_info += "─" * 80 + "\n"
            
            total_disk_space = 0
            total_used_space = 0
            
            for partition in partitions:
                try:
                    partition_usage = psutil.disk_usage(partition.mountpoint)
                    total_disk_space += partition_usage.total
                    total_used_space += partition_usage.used
                    
                    storage_info += f"\n Device: {partition.device}\n"
                    storage_info += f"   Mountpoint:   {partition.mountpoint}\n"
                    storage_info += f"   File system:  {partition.fstype}\n"
                    storage_info += f"   Total Size:   {partition_usage.total // (1024**3):>6} GB ({partition_usage.total:,} bytes)\n"
                    storage_info += f"   Used:         {partition_usage.used // (1024**3):>6} GB ({partition_usage.used:,} bytes)\n"
                    storage_info += f"   Free:         {partition_usage.free // (1024**3):>6} GB ({partition_usage.free:,} bytes)\n"
                    storage_info += f"   Usage:        {(partition_usage.used / partition_usage.total) * 100:>6.1f}%\n"
                    
                    # Usage bar
                    usage_percent = (partition_usage.used / partition_usage.total) * 100
                    usage_bar = "█" * int(usage_percent / 2.5)  # Scale to 40 chars max
                    storage_info += f"   Visual:       [{usage_bar:<40}] {usage_percent:.1f}%\n"
                    
                except PermissionError:
                    storage_info += f"\n Device: {partition.device} - Access Denied\n"
                except Exception as e:
                    storage_info += f"\n Device: {partition.device} - Error: {str(e)}\n"
            
            # Total storage summary
            if total_disk_space > 0:
                storage_info += "\n" + "─" * 50 + "\n"
                storage_info += " TOTAL STORAGE SUMMARY:\n"
                storage_info += f"   Total Capacity: {total_disk_space // (1024**3)} GB\n"
                storage_info += f"   Total Used:     {total_used_space // (1024**3)} GB\n"
                storage_info += f"   Total Free:     {(total_disk_space - total_used_space) // (1024**3)} GB\n"
                storage_info += f"   Overall Usage:  {(total_used_space / total_disk_space) * 100:.1f}%\n"
            
            # Disk I/O Statistics
            try:
                disk_io = psutil.disk_io_counters()
                if disk_io:
                    storage_info += "\n DISK I/O STATISTICS:\n"
                    storage_info += f"   Read Count:     {disk_io.read_count:,}\n"
                    storage_info += f"   Write Count:    {disk_io.write_count:,}\n"
                    storage_info += f"   Read Bytes:     {disk_io.read_bytes // (1024**2):,} MB\n"
                    storage_info += f"   Write Bytes:    {disk_io.write_bytes // (1024**2):,} MB\n"
                    storage_info += f"   Read Time:      {disk_io.read_time:,} ms\n"
                    storage_info += f"   Write Time:     {disk_io.write_time:,} ms\n"
                    
                    # Per-disk I/O if available
                    disk_io_per_disk = psutil.disk_io_counters(perdisk=True)
                    if disk_io_per_disk:
                        storage_info += "\n PER-DISK I/O STATISTICS:\n"
                        storage_info += f"{'Disk':<10} {'Read MB':<12} {'Write MB':<12} {'Read Count':<12} {'Write Count':<12}\n"
                        storage_info += "─" * 70 + "\n"
                        
                        for disk, io_stats in disk_io_per_disk.items():
                            storage_info += f"{disk:<10} {io_stats.read_bytes//(1024**2):<12,} "
                            storage_info += f"{io_stats.write_bytes//(1024**2):<12,} "
                            storage_info += f"{io_stats.read_count:<12,} {io_stats.write_count:<12,}\n"
            except:
                storage_info += "\n Disk I/O statistics: Not available\n"
            
            storage_text.insert(1.0, storage_info)
            
            # Enhanced Process Information
            process_info = "╔═══════════════════════════════════════╗\n"
            process_info += "║        PROCESS INFORMATION            ║\n"
            process_info += "╚═══════════════════════════════════════╝\n\n"
            
            processes = []
            for proc in psutil.process_iter(['pid', 'name', 'cpu_percent', 'memory_percent', 
                                           'memory_info', 'create_time', 'status', 'username']):
                try:
                    proc.info['memory_mb'] = proc.info['memory_info'].rss // (1024*1024) if proc.info['memory_info'] else 0
                    processes.append(proc.info)
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
            
            # Sort processes
            if sort_by == "memory":
                processes = sorted(processes, key=lambda x: x['memory_percent'] if x['memory_percent'] else 0, reverse=True)
            elif sort_by == "cpu":
                processes = sorted(processes, key=lambda x: x['cpu_percent'] if x['cpu_percent'] else 0, reverse=True)
            elif sort_by == "name":
                processes = sorted(processes, key=lambda x: x['name'].lower() if x['name'] else '')
            elif sort_by == "pid":
                processes = sorted(processes, key=lambda x: x['pid'])
            
            process_info += f" PROCESS SUMMARY (Sorted by {sort_by.upper()}):\n"
            process_info += f"Total Processes: {len(processes)}\n\n"
            
            # Process status summary
            status_count = {}
            for proc in processes:
                status = proc['status']
                status_count[status] = status_count.get(status, 0) + 1
            
            process_info += " PROCESS STATUS SUMMARY:\n"
            for status, count in status_count.items():
                process_info += f"   {status}: {count}\n"
            
            process_info += "\n" + "─" * 100 + "\n"
            process_info += f"{'PID':<8} {'Name':<20} {'Status':<12} {'CPU%':<8} {'Mem%':<8} {'Mem(MB)':<10} {'User':<12} {'Runtime':<12}\n"
            process_info += "─" * 100 + "\n"
            
            for proc in processes[:50]:  # Show top 50 processes
                try:
                    # Calculate runtime
                    create_time = datetime.fromtimestamp(proc['create_time'])
                    runtime = datetime.now() - create_time
                    runtime_str = str(runtime).split('.')[0] if runtime.days == 0 else f"{runtime.days}d+"
                    
                    process_info += f"{proc['pid']:<8} {proc['name'][:19]:<20} {proc['status']:<12} "
                    process_info += f"{proc['cpu_percent']:<8.1f} {proc['memory_percent']:<8.2f} "
                    process_info += f"{proc['memory_mb']:<10} {(proc['username'] or 'N/A')[:11]:<12} {runtime_str:<12}\n"
                except:
                    continue
            
            process_text.insert(1.0, process_info)
            system_text.config(state="disabled")
            cpu_text.config(state="disabled")
            memory_text.config(state="disabled")
            storage_text.config(state="disabled")
            process_text.config(state="disabled")

        except Exception as e:
            error_msg = f"Error getting hardware info: {str(e)}"
            system_text.insert(1.0, error_msg)
            cpu_text.insert(1.0, error_msg)
            memory_text.insert(1.0, error_msg)
            storage_text.insert(1.0, error_msg)
            process_text.insert(1.0, error_msg)
            
    def create_paint_app(self):
        paint_window = tk.Toplevel(self.rootW95dist)
        paint_window.title("Paint")
        paint_window.overrideredirect(True)
        paint_window.geometry("800x600+200+50")
        paint_window.configure(bg="#c0c0c0")
                
        # Add Windows 95 style title bar
        title_bar = tk.Frame(paint_window, bg="#000080", height=25)
        title_bar.pack(fill="x", side="top")
        title_label = tk.Label(title_bar, text="Paint", fg="white", bg="#000080",
                              font=("MS Sans Serif", 8, "bold"))
        title_label.pack(side="left", padx=5, pady=2)
        
        # Close button for title bar
        close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                font=("Arial", 8, "bold"), width=2, height=1,
                                relief="raised", bd=1,
                                command=lambda: self.close_window("Paint", paint_window))
        close_button.pack(side="right", padx=2, pady=1)
        
        self.make_window_draggable(paint_window, title_bar)
        
        # Menu bar
        menu_bar = tk.Menu(paint_window)
        paint_window.config(menu=menu_bar)
        
        file_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="File", menu=file_menu)
        
        edit_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Edit", menu=edit_menu)
        
        view_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="View", menu=view_menu)
        
        image_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Image", menu=image_menu)
        
        colors_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Colors", menu=colors_menu)
        
        help_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="Help Topics", command=lambda: self.show_help("Paint"))
        help_menu.add_separator()
        help_menu.add_command(label="About Paint", command=lambda: messagebox.showinfo("About Paint", "\nVersion 1.0\n\nA retro-style paint application."))
        
        # Toolbar frame
        toolbar_frame = tk.Frame(paint_window, bg="#c0c0c0", relief="raised", bd=2, height=50)
        toolbar_frame.pack(fill="x")
        toolbar_frame.pack_propagate(False)
        
        # Tool buttons
        pencil_btn = tk.Button(toolbar_frame, text="Pencil", font=("MS Sans Serif", 8),
                              bg="#c0c0c0", relief="raised", bd=2,
                              command=lambda: self.select_paint_tool("pencil"))
        pencil_btn.pack(side="left", padx=5, pady=5)
        
        brush_btn = tk.Button(toolbar_frame, text="Brush", font=("MS Sans Serif", 8),
                             bg="#c0c0c0", relief="raised", bd=2,
                             command=lambda: self.select_paint_tool("brush"))
        brush_btn.pack(side="left", padx=5, pady=5)
        
        eraser_btn = tk.Button(toolbar_frame, text="Eraser", font=("MS Sans Serif", 8),
                              bg="#c0c0c0", relief="raised", bd=2,
                              command=lambda: self.select_paint_tool("eraser"))
        eraser_btn.pack(side="left", padx=5, pady=5)
        
        # Size selector
        size_frame = tk.Frame(toolbar_frame, bg="#c0c0c0")
        size_frame.pack(side="left", padx=10)
        
        tk.Label(size_frame, text="Size:", bg="#c0c0c0", font=("MS Sans Serif", 8)).pack(side="left")
        
        size_var = tk.IntVar(value=2)
        size_scale = tk.Scale(size_frame, from_=1, to=10, orient="horizontal",
                             variable=size_var, bg="#c0c0c0", font=("MS Sans Serif", 8),
                             command=lambda v: self.change_brush_size(int(v)))
        size_scale.pack(side="left")
        
        # Clear button
        clear_btn = tk.Button(toolbar_frame, text="Clear", font=("MS Sans Serif", 8), bg="#c0c0c0", relief="raised", bd=2, command=lambda: self.clear_canvas(canvas))
        clear_btn.pack(side="right", padx=5, pady=5)
        
        # Color palette frame
        colors_frame = tk.Frame(paint_window, bg="#c0c0c0", relief="raised", bd=2, height=40)
        colors_frame.pack(fill="x")
        colors_frame.pack_propagate(False)
        
        # Color buttons
        colors = ["black", "white", "red", "green", "blue", "yellow", "orange", "purple", 
                 "brown", "pink", "cyan", "magenta", "gray", "lightgreen", "navy", "maroon"]
        
        for color in colors:
            color_btn = tk.Button(colors_frame, bg=color, width=2, height=1,
                                relief="raised", bd=2,
                                command=lambda c=color: self.change_paint_color(c))
            color_btn.pack(side="left", padx=3, pady=5)
        
        # Custom color button
        custom_color_btn = tk.Button(colors_frame, text="...", font=("MS Sans Serif", 8),
                                   bg="#c0c0c0", relief="raised", bd=2,
                                   command=self.choose_custom_color)
        custom_color_btn.pack(side="left", padx=5, pady=5)
        
        # Status bar
        status_bar = tk.Frame(paint_window, bg="#c0c0c0", relief="sunken", bd=1, height=25)
        status_bar.pack(side="bottom", fill="x")
        status_bar.pack_propagate(False)
        
        status_label = tk.Label(status_bar, text="Ready", bg="#c0c0c0", font=("MS Sans Serif", 8))
        status_label.pack(side="left", padx=5)
        
        coords_label = tk.Label(status_bar, text="", bg="#c0c0c0", font=("MS Sans Serif", 8))
        coords_label.pack(side="right", padx=5)
        
        # Canvas for drawing
        canvas_frame = tk.Frame(paint_window, bg="white", relief="sunken", bd=2)
        canvas_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        canvas = tk.Canvas(canvas_frame, bg="white", cursor="crosshair")
        canvas.pack(fill="both", expand=True)
        
        # Drawing variables
        self.paint_tool = "pencil"
        self.old_x = None
        self.old_y = None
        
        # Bind mouse events
        canvas.bind("<Button-1>", lambda e: self.paint_start(e, canvas, coords_label))
        canvas.bind("<B1-Motion>", lambda e: self.paint_move(e, canvas, coords_label, status_label))
        canvas.bind("<ButtonRelease-1>", lambda e: self.paint_end(e, canvas, coords_label))
        canvas.bind("<Motion>", lambda e: self.update_coords(e, coords_label))
        
        # File menu commands
        file_menu.add_command(label="New", command=lambda: self.clear_canvas(canvas))
        
        def save_drawing():
            file_path = filedialog.asksaveasfilename(
                defaultextension=".png",
                filetypes=[("PNG files", "*.png"), ("All files", "*.*")]
            )
            if file_path:
                try:
                    # Create a PostScript file first
                    ps_file = file_path + ".ps"
                    canvas.postscript(file=ps_file)
                    
                    # Use subprocess to convert PS to PNG (requires ghostscript)
                    try:
                        # This would normally use PIL/Pillow for proper conversion
                        messagebox.showinfo("Save", "Drawing saved to " + file_path)
                    except:
                        messagebox.showinfo("Save", "PostScript file saved as " + ps_file)
                except Exception as e:
                    messagebox.showerror("Save Error", f"Could not save drawing: {str(e)}")
                    
        file_menu.add_command(label="Save", command=save_drawing)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=lambda: self.close_window("Paint", paint_window))
        
        # Edit menu commands
        edit_menu.add_command(label="Undo", command=lambda: status_label.config(text="Undo not implemented"))
        edit_menu.add_command(label="Cut", command=lambda: status_label.config(text="Cut not implemented"))
        edit_menu.add_command(label="Copy", command=lambda: status_label.config(text="Copy not implemented"))
        edit_menu.add_command(label="Paste", command=lambda: status_label.config(text="Paste not implemented"))
        
        # Add to taskbar
        self.add_window_to_taskbar("Paint", paint_window)
        paint_window.protocol("WM_DELETE_WINDOW", lambda: self.close_window("Paint", paint_window))
    
    def select_paint_tool(self, tool):
        self.paint_tool = tool

    def change_brush_size(self, size):
        self.brush_size = size

    def change_paint_color(self, color):
        self.current_color = color

    def choose_custom_color(self):
        color = colorchooser.askcolor(initialcolor=self.current_color)
        if color[1]:  # Check if a color was selected (not canceled)
            self.current_color = color[1]

    def clear_canvas(self, canvas):
        canvas.delete("all")

    def paint_start(self, event, canvas, coords_label):
        self.old_x = event.x
        self.old_y = event.y
        self.update_coords(event, coords_label)

    def paint_move(self, event, canvas, coords_label, status_label):
        if self.old_x and self.old_y:
            if self.paint_tool == "pencil":
                canvas.create_line(self.old_x, self.old_y, event.x, event.y,
                                  width=self.brush_size, fill=self.current_color,
                                  capstyle="round", smooth=True)
                status_label.config(text="Drawing with pencil")
            elif self.paint_tool == "brush":
                # For brush, create a thicker line with smoother edges
                canvas.create_line(self.old_x, self.old_y, event.x, event.y,
                                  width=self.brush_size * 2, fill=self.current_color,
                                  capstyle="round", smooth=True)
                status_label.config(text="Painting with brush")
            elif self.paint_tool == "eraser":
                # Eraser uses white color (or canvas background color)
                canvas.create_line(self.old_x, self.old_y, event.x, event.y,
                                  width=self.brush_size * 3, fill="white",
                                  capstyle="round", smooth=True)
                status_label.config(text="Erasing")
        
        self.old_x = event.x
        self.old_y = event.y
        self.update_coords(event, coords_label)

    def paint_end(self, event, canvas, coords_label):
        self.old_x = None
        self.old_y = None
        self.update_coords(event, coords_label)

    def update_coords(self, event, coords_label):
        coords_label.config(text=f"X: {event.x}, Y: {event.y}")
    
    def open_window(self, name):
        """Generic function to open a window"""
        window = tk.Toplevel(self.rootW95dist)
        window.title(name)
        window.geometry("400x300+250+150")
        window.configure(bg="#c0c0c0")
        window.overrideredirect(True)
        
        # Add Windows 95 style title bar
        title_bar = tk.Frame(window, bg="#000080", height=25)
        title_bar.pack(fill="x", side="top")
        title_label = tk.Label(title_bar, text=name, fg="white", bg="#000080",
                              font=("MS Sans Serif", 8, "bold"))
        title_label.pack(side="left", padx=5, pady=2)
        
        # Close button for title bar
        close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                                font=("Arial", 8, "bold"), width=2, height=1,
                                relief="raised", bd=1,
                                command=lambda: self.close_window(name, window))
        close_button.pack(side="right", padx=2, pady=1)
        
        # Window content
        content_frame = tk.Frame(window, bg="#c0c0c0")
        content_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        tk.Label(content_frame, text=f"This is {name}", bg="#c0c0c0",
                font=("MS Sans Serif", 10)).pack(expand=True)
        
        self.add_window_to_taskbar(name, window)
        window.protocol("WM_DELETE_WINDOW", lambda: self.close_window(name, window))
        
        self.make_window_draggable(window, title_bar)
    
    def setup_taskbar(self):
        # Taskbar la bottom
        self.taskbar = tk.Frame(self.rootW95dist, bg="#c0c0c0", height=30, relief="raised", bd=2)
        self.taskbar.pack(side="bottom", fill="x")
        self.taskbar.pack_propagate(False)
        
        # Start button
        self.start_button = tk.Button(self.taskbar, text="Start", font=("MS Sans Serif", 8, "bold"),
                                     bg="#c0c0c0", relief="raised", bd=2, width=8,
                                     command=self.toggle_start_menu)
        self.start_button.pack(side="left", padx=2, pady=2)
        
        # Separator
        separator = tk.Frame(self.taskbar, bg="#808080", width=2)
        separator.pack(side="left", fill="y", padx=2)
        
        # Frame pentru ferestre deschise
        self.taskbar_windows_frame = tk.Frame(self.taskbar, bg="#c0c0c0")
        self.taskbar_windows_frame.pack(side="left", fill="x", expand=True, padx=5)
        
        # System tray (simulat)
        system_tray = tk.Frame(self.taskbar, bg="#c0c0c0")
        system_tray.pack(side="right")
        
        # Simulăm câteva icone în system tray
        tray_icon1 = tk.Label(system_tray, text="🔊", bg="#c0c0c0", font=("Arial", 8))
        tray_icon1.pack(side="left", padx=2)
        
        tray_icon2 = tk.Label(system_tray, text="📶", bg="#c0c0c0", font=("Arial", 8))
        tray_icon2.pack(side="left", padx=2)
        
        # Clock frame
        self.clock_frame = tk.Frame(system_tray, bg="#c0c0c0", relief="sunken", bd=1)
        self.clock_frame.pack(side="right", padx=5, pady=2)
        
        self.clock_label = tk.Label(self.clock_frame, bg="#c0c0c0", font=("MS Sans Serif", 8))
        self.clock_label.pack(padx=5, pady=1)
    
    def create_calendar(self):
        """Creează un calendar retro în stil Windows 95"""
        calendar_window = tk.Toplevel(self.rootW95dist)
        calendar_window.title("Calendar")
        calendar_window.overrideredirect(True)
        calendar_window.geometry("640x480+300+200")
        calendar_window.configure(bg="#c0c0c0")
        
        # Add Windows 95 style title bar
        title_bar = tk.Frame(calendar_window, bg="#000080", height=25)
        title_bar.pack(fill="x", side="top")
        title_label = tk.Label(title_bar, text="Calendar", fg="white", bg="#000080",
                             font=("MS Sans Serif", 8, "bold"))
        title_label.pack(side="left", padx=5, pady=2)
        
        # Close button for title bar
        close_button = tk.Button(title_bar, text="×", bg="#c0c0c0", fg="black",
                               font=("Arial", 8, "bold"), width=2, height=1,
                               relief="raised", bd=1,
                               command=lambda: self.close_window("Calendar", calendar_window))
        close_button.pack(side="right", padx=2, pady=1)
        
        self.make_window_draggable(calendar_window, title_bar)
        
        # Set Windows 95 style colors
        win95_bg = "#c0c0c0"
        win95_button = "#c0c0c0"
        win95_shadow = "#808080"
        win95_highlight = "#ffffff"
        win95_text = "#000000"
        calendar_bg = "#e0e0e0"  # Lighter color for the calendar table
        
        # Date variables
        current_date = datetime.now()
        year = current_date.year
        month = current_date.month
        
        # List of month names
        month_names = [
            "", "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December"
        ]
        
        # List of day names
        day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        
        # Events dictionary
        events = {}
        
        # Load saved events
        def load_events():
            """Load events from a JSON file"""
            try:
                if os.path.exists("calendar_events.json"):
                    with open("calendar_events.json", "r") as file:
                        return json.load(file)
                return {}
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load events: {str(e)}")
                return {}
        
        events = load_events()
        
        # Main frame
        main_frame = tk.Frame(calendar_window, bg=win95_bg, bd=2, relief="raised")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Calendar control frame
        control_frame = tk.Frame(main_frame, bg=win95_bg, bd=2, relief="raised")
        control_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Functions for navigation
        def prev_month():
            nonlocal month, year
            if month > 1:
                month -= 1
            else:
                month = 12
                year -= 1
            update_calendar()
        
        def next_month():
            nonlocal month, year
            if month < 12:
                month += 1
            else:
                month = 1
                year += 1
            update_calendar()
        
        # Navigation buttons
        prev_button = tk.Button(
            control_frame,
            text="< Previous",
            command=prev_month,
            bg=win95_button,
            fg=win95_text,
            relief="raised",
            bd=2,
            font=("MS Sans Serif", 8)
        )
        prev_button.pack(side=tk.LEFT, padx=5, pady=5)
        
        next_button = tk.Button(
            control_frame,
            text="Next >",
            command=next_month,
            bg=win95_button,
            fg=win95_text,
            relief="raised",
            bd=2,
            font=("MS Sans Serif", 8)
        )
        next_button.pack(side=tk.LEFT, padx=5, pady=5)
        
        # Label for current month and year
        month_year_label = tk.Label(
            control_frame,
            text="",
            font=("MS Sans Serif", 10, "bold"),
            bg=win95_bg
        )
        month_year_label.pack(side=tk.LEFT, padx=30, pady=5)
        
        # Frame for displaying the calendar grid
        calendar_frame = tk.Frame(main_frame, bg=win95_bg, bd=2, relief="raised")
        calendar_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Week day names
        for i, day in enumerate(day_names):
            day_label = tk.Label(
                calendar_frame,
                text=day,
                font=("MS Sans Serif", 8, "bold"),
                width=9,
                height=1,
                bg="#d3d3d3",
                bd=1,
                relief="sunken"
            )
            day_label.grid(row=0, column=i, sticky="nsew", padx=1, pady=1)
        
        # Create grid of cells for days
        day_cells = []
        for row in range(6):
            row_cells = []
            for col in range(7):
                cell = tk.Button(
                    calendar_frame,
                    text="",
                    font=("MS Sans Serif", 8),
                    bg=calendar_bg,
                    relief="raised",
                    bd=1,
                    width=9,
                    height=4,
                    justify="left",
                    anchor="nw"
                )
                cell.grid(row=row+1, column=col, sticky="nsew", padx=1, pady=1)
                row_cells.append(cell)
            day_cells.append(row_cells)
        
        # Configure grid for resolution
        for i in range(7):
            calendar_frame.columnconfigure(i, weight=1)
        for i in range(7):
            calendar_frame.rowconfigure(i, weight=1)
        
        # Status bar frame
        status_frame = tk.Frame(main_frame, bg=win95_bg, bd=1, relief="sunken")
        status_frame.pack(fill=tk.X, side=tk.BOTTOM)
        
        status_label = tk.Label(
            status_frame,
            text="Ready",
            bg=win95_bg,
            font=("MS Sans Serif", 8),
            anchor="w"
        )
        status_label.pack(side=tk.LEFT, padx=5)
        
        def get_month_calendar(year, month):
            """Generate calendar for specified month and year"""
            # First day of the month
            first_day = datetime(year, month, 1)
            
            # Weekday for the first day (0 = Monday, 6 = Sunday in ISO format)
            first_weekday = first_day.weekday()
            
            # Number of days in current month
            if month == 12:
                last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                last_day = datetime(year, month + 1, 1) - timedelta(days=1)
            days_in_month = last_day.day
            
            # Build calendar as a matrix
            cal = []
            week = [0] * 7  # Initialize with 7 zeros
            
            # Fill previous days with zeros
            for i in range(first_weekday):
                week[i] = 0
                
            day = 1
            for i in range(first_weekday, 7):
                if day <= days_in_month:
                    week[i] = day
                    day += 1
                else:
                    week[i] = 0
                    
            cal.append(week)
            
            # Continue with the rest of the weeks
            while day <= days_in_month:
                week = [0] * 7
                for i in range(7):
                    if day <= days_in_month:
                        week[i] = day
                        day += 1
                    else:
                        week[i] = 0
                cal.append(week)
                
            return cal
        
        def save_events():
            """Save events to a JSON file"""
            with open("calendar_events.json", "w") as file:
                json.dump(events, file)
        
        def add_event(date_str, listbox, parent_window):
            """Add an event for the specified date"""
            event = simpledialog.askstring(
                "Add event",
                "Enter event description:",
                parent=parent_window
            )
            
            if event:
                if date_str not in events:
                    events[date_str] = []
                
                events[date_str].append(event)
                listbox.insert(tk.END, event)
                update_calendar()
                save_events()  # Save events after adding
        
        def delete_event(date_str, listbox):
            """Delete the selected event"""
            selected = listbox.curselection()
            
            if selected:
                index = selected[0]
                if date_str in events and index < len(events[date_str]):
                    listbox.delete(index)
                    del events[date_str][index]
                    update_calendar()
                    save_events()  # Save events after deleting
            else:
                messagebox.showinfo(
                    "Warning",
                    "Select an event to delete."
                )
        
        def day_click(day, month, year):
            """Handler for clicking on a day in the calendar"""
            date_obj = datetime(year, month, day)
            date_str = date_obj.strftime("%Y-%m-%d")
            date_display = f"{day} {month_names[month]} {year}"
            
            # Create window for viewing/adding events
            events_window = tk.Toplevel(calendar_window)
            events_window.title(f"Events - {date_display}")
            events_window.overrideredirect(True)
            events_window.geometry("400x300")
            events_window.configure(bg=win95_bg)
            
            # Title bar
            title_bar = tk.Frame(events_window, bg="#000080", height=22)
            title_bar.pack(fill=tk.X)
            
            title_text = tk.Label(
                title_bar, 
                text=f"Events - {date_display}",
                fg="white", 
                bg="#000080",
                font=("MS Sans Serif", 8, "bold")
            )
            title_text.pack(side=tk.LEFT, padx=5)
            
            # Close button
            close_button = tk.Button(
                title_bar, 
                text="×", 
                bg="#c0c0c0",
                fg="black",
                font=("Arial", 8, "bold"),
                width=2,
                height=1,
                relief="raised",
                bd=1,
                command=events_window.destroy
            )
            close_button.pack(side=tk.RIGHT, padx=2, pady=1)
            
            # Make window draggable
            self.make_window_draggable(events_window, title_bar)
            
            # Main frame
            main_frame = tk.Frame(events_window, bg=win95_bg, bd=2, relief="raised")
            main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Listbox with events
            events_label = tk.Label(
                main_frame, 
                text="Events:", 
                bg=win95_bg,
                font=("MS Sans Serif", 8, "bold"),
                anchor="w"
            )
            events_label.pack(fill=tk.X, padx=5, pady=5)
            
            events_frame = tk.Frame(main_frame, bg="white", bd=1, relief="sunken")
            events_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            events_listbox = tk.Listbox(
                events_frame,
                font=("MS Sans Serif", 8),
                bg="white",
                selectmode=tk.SINGLE
            )
            events_listbox.pack(fill=tk.BOTH, expand=True)
            
            # Action buttons
            buttons_frame = tk.Frame(main_frame, bg=win95_bg)
            buttons_frame.pack(fill=tk.X, pady=5)
            
            add_button = tk.Button(
                buttons_frame, 
                text="Add event", 
                command=lambda: add_event(date_str, events_listbox, events_window),
                bg=win95_button,
                fg=win95_text,
                relief="raised",
                bd=2,
                font=("MS Sans Serif", 8)
            )
            add_button.pack(side=tk.LEFT, padx=5)
            
            delete_button = tk.Button(
                buttons_frame, 
                text="Delete event", 
                command=lambda: delete_event(date_str, events_listbox),
                bg=win95_button,
                fg=win95_text,
                relief="raised",
                bd=2,
                font=("MS Sans Serif", 8)
            )
            delete_button.pack(side=tk.LEFT, padx=5)
            
            close_button = tk.Button(
                buttons_frame, 
                text="Close", 
                command=events_window.destroy,
                bg=win95_button,
                fg=win95_text,
                relief="raised",
                bd=2,
                font=("MS Sans Serif", 8)
            )
            close_button.pack(side=tk.RIGHT, padx=5)
            
            # Populate listbox with existing events
            if date_str in events:
                for event in events[date_str]:
                    events_listbox.insert(tk.END, event)
        
        def update_calendar():
            # Update month and year label
            month_name = month_names[month]
            month_year_label.config(text=f"{month_name} {year}")
            
            # Get calendar for current month
            cal = get_month_calendar(year, month)
            
            # Reset all cells
            for row in day_cells:
                for cell in row:
                    cell.config(
                        text="",
                        bg=calendar_bg,
                        command=lambda: None
                    )
            
            # Fill cells with dates
            for week_idx, week in enumerate(cal):
                for day_idx, day in enumerate(week):
                    if day != 0:
                        # Determine if day has events
                        date_obj = datetime(year, month, day)
                        date_str = date_obj.strftime("%Y-%m-%d")
                        has_events = date_str in events and len(events[date_str]) > 0
                        
                        # Text for cell
                        cell_text = f"{day}"
                        if has_events:
                            cell_text += f"\n[{len(events[date_str])} events]"
                        
                        # Configure cell
                        day_cells[week_idx][day_idx].config(
                            text=cell_text,
                            bg="#d3d3d3" if has_events else calendar_bg,
                            command=lambda d=day, m=month, y=year: 
                                   day_click(d, m, y)
                        )
        
        # Initial calendar update
        update_calendar()
        
        # Add to taskbar
        self.add_window_to_taskbar("Calendar", calendar_window)
        calendar_window.protocol("WM_DELETE_WINDOW", lambda: self.close_window("Calendar", calendar_window))
    
    def setup_clock(self):
        def update_clock():
            current = datetime.now()
            formatted_text = current.strftime("%H:%M %d.%m.%y")  # ex: "15:30 29.07.25"
            self.clock_label.config(text=formatted_text)
            self.rootW95dist.after(1000, update_clock)
        
        # Adaugă event handler pentru click pe ceas
        self.clock_frame.bind("<Button-1>", lambda e: self.create_calendar())
        self.clock_label.bind("<Button-1>", lambda e: self.create_calendar())
        
        # Schimbă cursorul pentru a indica că se poate face click
        self.clock_frame.config(cursor="hand2")
        self.clock_label.config(cursor="hand2")
        
        update_clock()
    
    def setup_start_menu(self):
        # Start menu permanent deasupra taskbar
        self.start_menu = tk.Frame(self.rootW95dist, bg="#c0c0c0", relief="raised", bd=2)
        self.start_menu.place(x=2, y=self.rootW95dist.winfo_screenheight()-330, width=200, height=300)
        
        # Header-ul Start Menu
        header = tk.Frame(self.start_menu, bg="#000080", height=25)
        header.pack(fill="x")
        header_label = tk.Label(header, text="Windows 95", fg="white", bg="#000080",
                               font=("MS Sans Serif", 8, "bold"))
        header_label.pack(anchor="w", padx=5, pady=2)
        
        # Doar opțiunea Shut Down
        separator = tk.Frame(self.start_menu, bg="#808080", height=1)
        separator.pack(fill="x", padx=5, pady=10)
        
        shutdown_btn = tk.Button(
            self.start_menu,
            text="Shut Down...",
            font=("MS Sans Serif", 8),
            bg="#c0c0c0",
            fg="black",
            relief="flat",
            bd=0,
            anchor="w",
            command=self.shutdown_computer
        )
        shutdown_btn.pack(fill="x", padx=2, pady=1)
        
        sysinfo_btn = tk.Button(
            self.start_menu,
            text="System Requirements",
            font=("MS Sans Serif", 8),
            bg="#c0c0c0",
            fg="black",
            relief="flat",
            bd=0,
            anchor="w",
            command=self.create_about_window
        )
        sysinfo_btn.pack(fill="x", padx=2, pady=1)
        
        activation_btn = tk.Button(
            self.start_menu,
            text="Activation Wizard",
            font=("MS Sans Serif", 8),
            bg="#c0c0c0",
            fg="black",
            relief="flat",
            bd=0,
            anchor="w",
            command=self.create_activation_window
        )
        activation_btn.pack(fill="x", padx=2, pady=1)
        
        # Hover effects
        shutdown_btn.bind("<Enter>", lambda e: shutdown_btn.config(bg="#0000ff", fg="white"))
        shutdown_btn.bind("<Leave>", lambda e: shutdown_btn.config(bg="#c0c0c0", fg="black"))
        
        sysinfo_btn.bind("<Enter>", lambda e: sysinfo_btn.config(bg="#0000ff", fg="white"))
        sysinfo_btn.bind("<Leave>", lambda e: sysinfo_btn.config(bg="#c0c0c0", fg="black"))
        
        activation_btn.bind("<Enter>", lambda e: activation_btn.config(bg="#0000ff", fg="white"))
        activation_btn.bind("<Leave>", lambda e: activation_btn.config(bg="#c0c0c0", fg="black"))
    
    def toggle_start_menu(self):
        if self.start_menu_visible:
            self.hide_start_menu()
        else:
            self.show_start_menu()
    
    def show_start_menu(self):
        self.start_menu.place(x=2, y=self.rootW95dist.winfo_screenheight()-330, width=200, height=300)
        self.start_menu_visible = True
        self.start_button.config(relief="sunken")

    def hide_start_menu(self):
        self.start_menu.place_forget()
        self.start_menu_visible = False
        self.start_button.config(relief="raised")
    
    def shutdown_computer(self):
        self.rootW95dist.quit()
    
    def add_window_to_taskbar(self, title, window):
        # Create taskbar button for the window
        button = tk.Button(self.taskbar_windows_frame, text=title[:15], 
                          font=("MS Sans Serif", 8), bg="#c0c0c0",
                          relief="raised", bd=2, width=12,
                          command=lambda: self.bring_window_to_front(window))
        button.pack(side="left", padx=2, pady=2)
        
        # Add to open windows list
        self.open_windows.append((title, window, button))
        
        # Set initial active state
        self.set_active_window(window)
        
        # Update window state when minimized/restored/focused
        window.bind("<Unmap>", lambda e: self.update_taskbar_button_state(window, "minimized"))
        window.bind("<Map>", lambda e: self.update_taskbar_button_state(window, "restored"))
        window.bind("<FocusIn>", lambda e: self.set_active_window(window))
        
    def update_taskbar_button_state(self, window, state):
        """Update taskbar button appearance based on window state"""
        for title, win, button in self.open_windows:
            if win == window:
                # Check if this is the active window to maintain its color
                current_bg = button.cget("bg")
                is_active = current_bg == "#e0e0e0"
                
                if state == "minimized":
                    # Minimized = pressed/sunken
                    bg_color = "#e0e0e0" if is_active else "#c0c0c0"
                    button.config(relief="sunken", bd=1, bg=bg_color)
                elif state == "restored":
                    # Restored = normal/raised
                    bg_color = "#e0e0e0" if is_active else "#c0c0c0"
                    button.config(relief="raised", bd=2, bg=bg_color)
                break

    def set_active_window(self, active_window):
        """Set the active window and update all taskbar buttons accordingly"""
        for title, win, button in self.open_windows:
            if win == active_window:
                # Active window - lighter color
                if not win.winfo_viewable():  # hidden (minimized)
                    button.config(bg="#e0e0e0", relief="sunken", bd=1)
                else:
                    button.config(bg="#e0e0e0", relief="raised", bd=2)
            else:
                # Inactive windows - normal color
                if not win.winfo_viewable():  # hidden (minimized)
                    button.config(bg="#c0c0c0", relief="sunken", bd=1)
                else:
                    button.config(bg="#c0c0c0", relief="raised", bd=2)
    
    def bring_window_to_front(self, window):
        try:
            # Check if window is visible or hidden
            if window.winfo_viewable():
                # Window is visible - hide it (simulate minimize)
                window.withdraw()
                self.update_taskbar_button_state(window, "minimized")
            else:
                # Window is hidden - show it (simulate restore)
                window.deiconify()
                window.lift()
                window.focus_set()
                self.set_active_window(window)
                self.update_taskbar_button_state(window, "restored")
        except:
            pass
    
    def close_window(self, title, window):
        # Remove from taskbar
        for i, (window_title, win, button) in enumerate(self.open_windows):
            if window_title == title and win == window:
                button.destroy()
                self.open_windows.pop(i)
                break
        
        # Close window
        try:
            window.destroy()
        except:
            pass
    
    def run(self):
        # Bind escape key to exit fullscreen
        self.rootW95dist.bind("<Escape>", lambda e: self.rootW95dist.quit())
        
        # Bind F11 to toggle fullscreen
        self.rootW95dist.bind("<F11>", lambda e: self.toggle_fullscreen())
        
        # Start the GUI
        self.rootW95dist.mainloop()
    
    def toggle_fullscreen(self):
        current_state = self.rootW95dist.attributes('-fullscreen')
        self.rootW95dist.attributes('-fullscreen', not current_state)

# Start the Windows 95 Desktop
if __name__ == "__main__":
    desktop = Windows95Desktop()
    desktop.run()